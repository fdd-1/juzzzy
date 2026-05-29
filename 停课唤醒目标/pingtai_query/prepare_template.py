#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""把 filtered_*.xlsx 里的学员 ID 列提取出来，按平台模板格式写入 xlsx：
   A1 = "导入用户id"
   A2.. = 学员 ID（一行一个）
   输出：pingtai_query/upload_*.xlsx
"""
import sys, io, argparse, datetime as dt
from pathlib import Path

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

SCRIPT_DIR = Path(__file__).parent
ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = ROOT / "output"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", help="filtered xlsx 路径，默认 output/filtered_{today}.xlsx")
    ap.add_argument("--id-col", default="学员id", help="学员 ID 列名（默认 学员id）")
    args = ap.parse_args()

    import pandas as pd
    from openpyxl import Workbook

    if args.input:
        src = Path(args.input)
    else:
        today = dt.date.today().strftime("%Y%m%d")
        src = OUTPUT_DIR / f"filtered_{today}.xlsx"

    if not src.exists():
        print(f"[ERROR] 找不到 {src}", flush=True)
        sys.exit(1)

    df = pd.read_excel(str(src))
    if args.id_col not in df.columns:
        # 容错：列名带空格或大小写
        candidate = [c for c in df.columns if str(c).strip().lower() == args.id_col.lower()]
        if not candidate:
            print(f"[ERROR] 找不到列 '{args.id_col}'，现有列: {list(df.columns)}", flush=True)
            sys.exit(2)
        col = candidate[0]
    else:
        col = args.id_col

    ids = df[col].dropna().astype("int64").tolist()
    ids = [int(x) for x in ids if int(x) > 0]
    print(f"[OK] 提取到 {len(ids)} 个学员 ID", flush=True)

    today = dt.date.today().strftime("%Y%m%d")
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out = SCRIPT_DIR / f"upload_{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "导入用户id"
    for i, sid in enumerate(ids, start=2):
        ws.cell(row=i, column=1, value=sid)
    wb.save(str(out))

    print(f"[OK] 已生成上传文件: {out}", flush=True)
    print(f"     条数: {len(ids)}", flush=True)


if __name__ == "__main__":
    main()
