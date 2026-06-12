#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""把 filtered_*.xlsx 的学员id 和 dadou_mapping_*.xlsx 的豌豆大账号id 各转成
   六一工作台「按导入用户id」标签所需的 xlsx：单列 A=「导入用户id」。

输出：
  liuyi_tag/user_ids_{stamp}.xlsx        ← filtered_*.xlsx 第一列「学员id」
  liuyi_tag/dadou_ids_{stamp}.xlsx       ← dadou_mapping_*.xlsx 第二列「豌豆大账号id」
"""
import sys, io, argparse, datetime as dt
from pathlib import Path

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

SCRIPT_DIR = Path(__file__).parent
ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = ROOT / "output"


def write_id_xlsx(ids, out_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "用户id"   # 跟官方模板 usertag_template.csv 保持一致
    for i, sid in enumerate(ids, start=2):
        ws.cell(row=i, column=1, value=int(sid))
    wb.save(str(out_path))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--filtered", help="filtered xlsx 路径，默认 output/filtered_{today}.xlsx")
    ap.add_argument("--dadou", help="dadou_mapping xlsx 路径，默认 output/dadou_mapping_{today}.xlsx")
    args = ap.parse_args()

    import pandas as pd

    today = dt.date.today().strftime("%Y%m%d")
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")

    # 学员 ID
    f_path = Path(args.filtered) if args.filtered else OUTPUT_DIR / f"filtered_{today}.xlsx"
    if not f_path.exists():
        print(f"[ERROR] 找不到 {f_path}", flush=True)
        sys.exit(1)
    df_f = pd.read_excel(str(f_path))
    user_id_col = None
    for c in df_f.columns:
        if str(c).strip().lower() in ("学员id", "user_id", "userid"):
            user_id_col = c
            break
    if user_id_col is None:
        user_id_col = df_f.columns[0]
        print(f"[WARN] 没有列名「学员id」，用第一列 {user_id_col!r}", flush=True)
    user_ids = [int(x) for x in df_f[user_id_col].dropna().astype("int64").tolist() if int(x) > 0]
    user_out = SCRIPT_DIR / f"user_ids_{stamp}.xlsx"
    write_id_xlsx(user_ids, user_out)
    print(f"[OK] 学员id 文件: {user_out} ({len(user_ids)} 条)", flush=True)

    # 豌豆大账号 ID
    d_path = Path(args.dadou) if args.dadou else OUTPUT_DIR / f"dadou_mapping_{today}.xlsx"
    if not d_path.exists():
        print(f"[ERROR] 找不到 {d_path}", flush=True)
        sys.exit(2)
    df_d = pd.read_excel(str(d_path))
    # 第 2 列是豌豆大账号 id
    if len(df_d.columns) < 2:
        print(f"[ERROR] {d_path} 列数不足 2", flush=True)
        sys.exit(3)
    dadou_col = df_d.columns[1]
    dadou_ids = [int(x) for x in df_d[dadou_col].dropna().astype("int64").tolist() if int(x) > 0]
    dadou_out = SCRIPT_DIR / f"dadou_ids_{stamp}.xlsx"
    write_id_xlsx(dadou_ids, dadou_out)
    print(f"[OK] 豌豆大账号id 文件: {dadou_out} ({len(dadou_ids)} 条)", flush=True)

    # 写一个清单方便后面用
    manifest = SCRIPT_DIR / "latest_inputs.json"
    import json
    manifest.write_text(json.dumps({
        "user_ids_xlsx": str(user_out),
        "user_ids_count": len(user_ids),
        "dadou_ids_xlsx": str(dadou_out),
        "dadou_ids_count": len(dadou_ids),
        "stamp": stamp,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] 清单: {manifest}", flush=True)


if __name__ == "__main__":
    main()
