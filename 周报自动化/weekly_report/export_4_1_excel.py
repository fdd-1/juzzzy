"""4.1 — 把合并宽表导出为格式化 Excel
- 多级表头：分组标题 + 列名
- 数值列：百分比/2位小数/整数自动格式化
- 团队级（汇总）行加底色
- 列宽自适应
"""
from __future__ import annotations
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
MERGED = ROOT.parent / "exports" / "4_1" / "_merged_4_1.xlsx"
OUT = ROOT.parent / "exports" / "4_1" / "4_1_格式化.xlsx"

# 列分组定义（和图片一致）— (分组名, [列名...], 分组色)
GROUPS = [
    ("基础", ["层级", "团队/小组", "LP", "在职数"], "D9D9D9"),
    ("首通语义分析执行", [
        "命中首通新生数",
        "SOP_首通_邀请添加企微执行率",
        "SOP_首通_一家多娃问询执行率",
        "SOP_首通_转介绍执行率",
        "SOP_首通_执行率加和",
    ], "FFE699"),
    ("首通", [
        "首通_新生数", "首通_拨通数",
        "首通_及时跟进率", "首通_生均接通时长",
        "首通_企微绑定率", "首通_秒挂占比",
        "首通_follow率", "首通_生均外呼次数",
    ], "BDD7EE"),
    ("首课SOP执行", ["SOP_首课_执行率加和"], "FFE699"),
    ("首课", ["首课_学员数", "首课_跟进率", "首课_及时跟进率", "首课_作业完成率"], "C6E0B4"),
    ("首专", ["首专_学员数", "首专_跟进率", "首专_及时跟进率", "首专_作业完成率"], "F4B084"),
    ("LP架构", ["入职时间", "入职月份", "状态"], "D9D9D9"),
]

# 显示列名简化
COL_LABELS = {
    "层级": "层级",
    "团队/小组": "团队/小组",
    "LP": "LP姓名",
    "在职数": "到岗",
    "命中首通新生数": "拨通且命中首通场景新生数",
    "SOP_首通_邀请添加企微执行率": "邀请添加企微/WS/Line执行率",
    "SOP_首通_一家多娃问询执行率": "一家多娃问询执行率",
    "SOP_首通_转介绍执行率": "转介绍执行率",
    "SOP_首通_执行率加和": "执行率加和",
    "首通_新生数": "新生数",
    "首通_拨通数": "拨通数",
    "首通_及时跟进率": "及时跟进率",
    "首通_生均接通时长": "生均接通时长",
    "首通_企微绑定率": "企微绑定率",
    "首通_秒挂占比": "秒挂占比",
    "首通_follow率": "follow率",
    "首通_生均外呼次数": "生均外呼次数",
    "SOP_首课_执行率加和": "执行率加和",
    "首课_学员数": "学员数",
    "首课_跟进率": "跟进率",
    "首课_及时跟进率": "及时跟进率",
    "首课_作业完成率": "作业完成率",
    "首专_学员数": "学员数",
    "首专_跟进率": "跟进率",
    "首专_及时跟进率": "及时跟进率",
    "首专_作业完成率": "作业完成率",
    "入职时间": "入职时间",
    "入职月份": "LP工龄(月)",
    "状态": "人员状态",
}

PCT_COLS = {
    "SOP_首通_邀请添加企微执行率", "SOP_首通_一家多娃问询执行率", "SOP_首通_转介绍执行率",
    "首通_及时跟进率", "首通_企微绑定率", "首通_秒挂占比", "首通_follow率",
    "首课_跟进率", "首课_及时跟进率", "首课_作业完成率",
    "首专_跟进率", "首专_及时跟进率", "首专_作业完成率",
}
DEC2_COLS = {
    "SOP_首通_执行率加和", "SOP_首课_执行率加和",
    "首通_生均接通时长", "首通_生均外呼次数",
}
INT_COLS = {
    "在职数", "命中首通新生数", "首通_新生数", "首通_拨通数",
    "首课_学员数", "首专_学员数", "入职月份",
}


def main():
    df = pd.read_excel(MERGED)
    # 按 GROUPS 顺序展开列
    cols = []
    for _, gcols, _ in GROUPS:
        for c in gcols:
            if c in df.columns:
                cols.append(c)
    df = df[cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "4.1合并宽表"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_white = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
    bold_black = Font(bold=True, name="微软雅黑", size=10)
    normal = Font(name="微软雅黑", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # === 第 1 行：分组标题 ===
    col_idx = 1
    for gname, gcols, gcolor in GROUPS:
        present = [c for c in gcols if c in df.columns]
        if not present:
            continue
        span = len(present)
        ws.cell(row=1, column=col_idx, value=gname)
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_white if gname not in ("基础", "LP架构") else bold_black
        cell.fill = PatternFill("solid", fgColor=("4472C4" if gname not in ("基础", "LP架构") else gcolor))
        cell.alignment = center
        cell.border = border
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + span - 1)
        for j in range(span):
            ws.cell(row=1, column=col_idx + j).border = border
        col_idx += span

    # === 第 2 行：列名 ===
    for j, c in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=j, value=COL_LABELS.get(c, c))
        # 找分组色
        gcolor = "FFFFFF"
        for gname, gcols, gc in GROUPS:
            if c in gcols:
                gcolor = gc
                break
        cell.font = bold_black
        cell.fill = PatternFill("solid", fgColor=gcolor)
        cell.alignment = center
        cell.border = border

    # === 数据行 ===
    summary_fill = PatternFill("solid", fgColor="FFF2CC")
    for ri, row in enumerate(df.itertuples(index=False), start=3):
        is_summary = (getattr(row, "_1", None) == "汇总") if hasattr(row, "_1") else False
        # 用 df.iloc 更安全
        is_summary = df.iloc[ri - 3]["层级"] == "汇总"
        for j, c in enumerate(df.columns, start=1):
            val = df.iloc[ri - 3][c]
            cell = ws.cell(row=ri, column=j)
            if pd.isna(val):
                cell.value = ""
            elif c in PCT_COLS:
                try:
                    cell.value = float(val)
                    cell.number_format = "0.0%"
                except Exception:
                    cell.value = str(val)
            elif c in DEC2_COLS:
                try:
                    cell.value = float(val)
                    cell.number_format = "0.00"
                except Exception:
                    cell.value = str(val)
            elif c in INT_COLS:
                try:
                    cell.value = int(val)
                except Exception:
                    cell.value = val
            elif c == "入职时间":
                try:
                    cell.value = pd.to_datetime(val).strftime("%Y-%m-%d")
                except Exception:
                    cell.value = str(val) if pd.notna(val) else ""
            else:
                cell.value = val
            cell.font = bold_black if is_summary else normal
            cell.alignment = center
            cell.border = border
            if is_summary:
                cell.fill = summary_fill

    # 冻结前两行 + 前 3 列
    ws.freeze_panes = "D3"

    # 列宽
    widths = {
        "层级": 8, "团队/小组": 14, "LP": 10, "在职数": 8,
        "命中首通新生数": 12,
    }
    for j, c in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(c, 14)

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 38

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] 输出: {OUT}")
    print(f"  rows={ws.max_row}, cols={ws.max_column}")


if __name__ == "__main__":
    main()
