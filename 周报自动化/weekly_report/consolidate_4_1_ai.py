"""4.1 AI 学情助手 — 单独成表
- 团队级汇总（9 组 + 海外团队合）：任务总数、覆盖学员数、首课/首专 干预中占比、人均任务数
- 个人级明细：每个 LP 的 AI 数据
- 格式化 Excel：多级表头、百分比格式、汇总行加底色
"""
from __future__ import annotations
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
EXPORT_BASE = ROOT.parent / "exports" / "4_1"
AI_FILE = EXPORT_BASE / "07_ai_summary" / "AI学情助手完成情况汇总（验收中）.xlsx"
ARCH_FILE = EXPORT_BASE / "06_lp_arch" / "海外思维LP架构表.xlsx"
OUT_DATA = EXPORT_BASE / "_merged_4_1_ai.xlsx"
OUT_FMT = EXPORT_BASE / "4_1_AI学情_格式化.xlsx"

TEAM_ORDER = ["港澳1组", "港澳2组", "港澳组",
              "美澳1组", "美澳2组", "美澳3组", "美澳4组", "美澳5组", "台湾组"]


def load_ai():
    df = pd.read_excel(AI_FILE, header=3)
    df = df.rename(columns={"班主任姓名": "LP", "LP小组": "小组"})
    df["首课_干预中占比"] = np.where(df["首课_任务总数"] > 0, df["首课_干预中"] / df["首课_任务总数"], np.nan)
    df["首专_干预中占比"] = np.where(df["首专_任务总数"] > 0, df["首专_干预中"] / df["首专_任务总数"], np.nan)
    return df


def build_merged():
    ai = load_ai()
    arch = pd.read_excel(ARCH_FILE, header=3).rename(
        columns={"姓名": "LP", "入职时长（月份）": "入职月份", "是否在职": "状态"})

    # ── 团队级汇总 ──
    team_rows = []
    # 海外团队合
    rec = {
        "层级": "汇总", "团队/小组": "海外团队（合）", "LP": "总计",
        "任务总数": int(ai["任务总数"].sum()),
        "覆盖学员数": int(ai["覆盖学员数"].sum()),
        "首课_任务总数": int(ai["首课_任务总数"].sum()),
        "首课_干预中": int(ai["首课_干预中"].sum()),
        "首专_任务总数": int(ai["首专_任务总数"].sum()),
        "首专_干预中": int(ai["首专_干预中"].sum()),
    }
    rec["首课_干预中占比"] = rec["首课_干预中"] / rec["首课_任务总数"] if rec["首课_任务总数"] > 0 else np.nan
    rec["首专_干预中占比"] = rec["首专_干预中"] / rec["首专_任务总数"] if rec["首专_任务总数"] > 0 else np.nan
    team_rows.append(rec)

    for team in TEAM_ORDER:
        sub = ai[ai["小组"] == team]
        if sub.empty:
            continue
        rec = {
            "层级": "汇总", "团队/小组": team, "LP": "总计",
            "任务总数": int(sub["任务总数"].sum()),
            "覆盖学员数": int(sub["覆盖学员数"].sum()),
            "首课_任务总数": int(sub["首课_任务总数"].sum()),
            "首课_干预中": int(sub["首课_干预中"].sum()),
            "首专_任务总数": int(sub["首专_任务总数"].sum()),
            "首专_干预中": int(sub["首专_干预中"].sum()),
        }
        rec["首课_干预中占比"] = rec["首课_干预中"] / rec["首课_任务总数"] if rec["首课_任务总数"] > 0 else np.nan
        rec["首专_干预中占比"] = rec["首专_干预中"] / rec["首专_任务总数"] if rec["首专_任务总数"] > 0 else np.nan
        team_rows.append(rec)
    summary = pd.DataFrame(team_rows)

    # ── 个人级 ──
    detail = ai[ai["小组"].isin(TEAM_ORDER)].copy()
    detail = detail.merge(arch[["小组", "LP", "入职月份", "状态"]], on=["小组", "LP"], how="left")
    detail.insert(0, "层级", "个人")
    detail = detail.rename(columns={"小组": "团队/小组"})
    detail["_o"] = detail["团队/小组"].map({t: i for i, t in enumerate(TEAM_ORDER)})
    detail = detail.sort_values(["_o", "入职月份"], ascending=[True, False], na_position="last").drop(columns=["_o"])

    cols = ["层级", "团队/小组", "LP", "状态", "入职月份",
            "任务总数", "覆盖学员数",
            "首课_任务总数", "首课_干预中", "首课_干预中占比",
            "首专_任务总数", "首专_干预中", "首专_干预中占比"]
    detail = detail[[c for c in cols if c in detail.columns]]

    merged = pd.concat([summary, detail], ignore_index=True)
    merged = merged[[c for c in cols if c in merged.columns]]
    return merged


# ── 格式化 Excel ──
GROUPS = [
    ("基础", ["层级", "团队/小组", "LP", "状态", "入职月份"], "D9D9D9"),
    ("总览", ["任务总数", "覆盖学员数"], "BDD7EE"),
    ("首课AI干预", ["首课_任务总数", "首课_干预中", "首课_干预中占比"], "C6E0B4"),
    ("首专AI干预", ["首专_任务总数", "首专_干预中", "首专_干预中占比"], "F4B084"),
]
COL_LABELS = {
    "层级": "层级", "团队/小组": "团队/小组", "LP": "LP姓名",
    "状态": "状态", "入职月份": "工龄(月)",
    "任务总数": "任务总数", "覆盖学员数": "覆盖学员数",
    "首课_任务总数": "任务总数", "首课_干预中": "干预中",
    "首课_干预中占比": "干预中占比",
    "首专_任务总数": "任务总数", "首专_干预中": "干预中",
    "首专_干预中占比": "干预中占比",
}
PCT_COLS = {"首课_干预中占比", "首专_干预中占比"}
INT_COLS = {"入职月份", "任务总数", "覆盖学员数",
            "首课_任务总数", "首课_干预中", "首专_任务总数", "首专_干预中"}


def export_excel(df: pd.DataFrame):
    cols = []
    for _, gcols, _ in GROUPS:
        for c in gcols:
            if c in df.columns:
                cols.append(c)
    df = df[cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "4.1_AI学情"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_white = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
    bold_black = Font(bold=True, name="微软雅黑", size=10)
    normal = Font(name="微软雅黑", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 第 1 行：分组
    col_idx = 1
    for gname, gcols, gcolor in GROUPS:
        present = [c for c in gcols if c in df.columns]
        if not present:
            continue
        span = len(present)
        cell = ws.cell(row=1, column=col_idx, value=gname)
        cell.font = bold_white if gname != "基础" else bold_black
        cell.fill = PatternFill("solid", fgColor=("4472C4" if gname != "基础" else gcolor))
        cell.alignment = center
        cell.border = border
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + span - 1)
        for j in range(span):
            ws.cell(row=1, column=col_idx + j).border = border
        col_idx += span

    # 第 2 行：列名
    for j, c in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=j, value=COL_LABELS.get(c, c))
        gcolor = "FFFFFF"
        for _, gcols, gc in GROUPS:
            if c in gcols:
                gcolor = gc
                break
        cell.font = bold_black
        cell.fill = PatternFill("solid", fgColor=gcolor)
        cell.alignment = center
        cell.border = border

    summary_fill = PatternFill("solid", fgColor="FFF2CC")
    for ri in range(len(df)):
        is_summary = df.iloc[ri]["层级"] == "汇总"
        for j, c in enumerate(df.columns, start=1):
            val = df.iloc[ri][c]
            cell = ws.cell(row=ri + 3, column=j)
            if pd.isna(val) or val == "":
                cell.value = ""
            elif c in PCT_COLS:
                try:
                    cell.value = float(val)
                    cell.number_format = "0.0%"
                except Exception:
                    cell.value = str(val)
            elif c in INT_COLS:
                try:
                    cell.value = int(val)
                except Exception:
                    cell.value = val
            else:
                cell.value = val
            cell.font = bold_black if is_summary else normal
            cell.alignment = center
            cell.border = border
            if is_summary:
                cell.fill = summary_fill

    ws.freeze_panes = "D3"
    widths = {"层级": 8, "团队/小组": 14, "LP": 10, "状态": 8, "入职月份": 8}
    for j, c in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(c, 12)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 30

    OUT_FMT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FMT)
    print(f"[OK] 格式化: {OUT_FMT}")


def main():
    merged = build_merged()
    EXPORT_BASE.mkdir(parents=True, exist_ok=True)
    merged.to_excel(OUT_DATA, index=False)
    print(f"[OK] 数据: {OUT_DATA}  rows={len(merged)}, cols={len(merged.columns)}")
    export_excel(merged)
    pd.set_option("display.max_columns", 30)
    pd.set_option("display.width", 250)
    print()
    print("=== 团队级汇总 ===")
    print(merged[merged["层级"] == "汇总"].to_string(index=False))


if __name__ == "__main__":
    main()
