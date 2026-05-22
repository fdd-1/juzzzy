"""把 Excel 表格数据通过剪贴板粘贴到飞书文档中。

流程：
1. 用 openpyxl 读取 Excel 转为 HTML
2. 用 Playwright 打开飞书文档
3. 通过 CDP 将 HTML 写入浏览器剪贴板，再触发粘贴
"""
from __future__ import annotations
import sys, io, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
import pandas as pd

ROOT = Path(__file__).parent
EXPORT_BASE = ROOT.parent / "exports" / "4_1"
MAIN_EXCEL = EXPORT_BASE / "4_1_格式化.xlsx"
AI_EXCEL = EXPORT_BASE / "4_1_AI学情_格式化.xlsx"

FEISHU_DOC_URL = "https://hcnig43mb8gp.feishu.cn/docx/ZFb3d1CZFobHnTxSSnMcgcVanyg"
BROWSER_STATE = Path.home() / ".workbuddy" / "skills" / "bi_skill" / "feishu_state.json"


def excel_to_html(excel_path: Path) -> str:
    """读取 Excel 并转为 HTML table"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    html = '<meta charset="utf-8"><table border="1" cellspacing="0" cellpadding="2" style="border-collapse:collapse;font-size:12px;">'
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        html += "<tr>"
        for cell in row:
            val = cell.value if cell.value is not None else ""
            if isinstance(val, float):
                fmt = cell.number_format or ""
                if "%" in fmt:
                    val = f"{val*100:.1f}%"
                elif "0.00" in fmt:
                    val = f"{val:.2f}"
                else:
                    val = f"{val:g}"
            elif isinstance(val, pd.Timestamp) or hasattr(val, 'strftime'):
                try:
                    val = val.strftime("%Y-%m-%d")
                except Exception:
                    val = str(val)
            else:
                val = str(val) if val != "" else ""
            tag = "th" if cell.row <= 2 else "td"
            html += f"<{tag}>{val}</{tag}>"
        html += "</tr>"
    html += "</table>"
    return html


def paste_html_via_cdp(page, html: str):
    """通过 execCommand 直接插入 HTML 到编辑器"""
    # 方法1: execCommand insertHTML（触发编辑器的 input 事件）
    result = page.evaluate("""(html) => {
        try {
            // 先尝试 execCommand
            const ok = document.execCommand('insertHTML', false, html);
            return {method: 'execCommand', ok: ok};
        } catch(e) {
            return {method: 'execCommand', ok: false, error: e.message};
        }
    }""", html)
    print(f"  execCommand result: {result}")
    if result.get("ok"):
        return True

    # 方法2: 通过 InputEvent 模拟输入
    result2 = page.evaluate("""(html) => {
        try {
            const target = document.querySelector('[data-content-editable-root="true"]')
                        || document.querySelector('[contenteditable="true"]')
                        || document.activeElement;
            const evt = new InputEvent('beforeinput', {
                inputType: 'insertFromPaste',
                data: null,
                dataTransfer: (() => {
                    const dt = new DataTransfer();
                    dt.setData('text/html', html);
                    return dt;
                })(),
                bubbles: true,
                cancelable: true,
                composed: true
            });
            target.dispatchEvent(evt);
            return {method: 'InputEvent', ok: true};
        } catch(e) {
            return {method: 'InputEvent', ok: false, error: e.message};
        }
    }""", html)
    print(f"  InputEvent result: {result2}")
    return result2.get("ok", False)


def click_placeholder_and_paste(page, html: str, label: str):
    """定位占位文本并粘贴"""
    placeholder = page.locator('text=（数据表格从 Excel 粘贴）')
    count = placeholder.count()
    print(f"  占位文本数量: {count}")

    if count > 0:
        placeholder.first.click()
        time.sleep(0.5)
        page.keyboard.press("Home")
        page.keyboard.press("Shift+End")
        time.sleep(0.3)
    else:
        page.keyboard.press("Control+End")
        page.keyboard.press("Enter")
        time.sleep(0.3)

    # 通过多种方式尝试粘贴
    paste_html_via_cdp(page, html)
    print(f"  [OK] 已尝试粘贴 {label}")
    time.sleep(10)

    # 验证
    new_count = page.locator('text=（数据表格从 Excel 粘贴）').count()
    if new_count < count:
        print(f"  [OK] 占位文本已替换 ({count} → {new_count})")
    else:
        print(f"  [WARN] 占位文本未变化 ({count} → {new_count})")


def main():
    print("=" * 50)
    print("4.1 Excel → 飞书文档 粘贴流程")
    print("=" * 50)

    print("\n[1/4] 读取 Excel 数据...")
    main_html = excel_to_html(MAIN_EXCEL)
    ai_html = excel_to_html(AI_EXCEL)
    print(f"  主表 HTML: {len(main_html)} chars")
    print(f"  AI 表 HTML: {len(ai_html)} chars")

    print("\n[2/4] 打开飞书文档...")
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False, channel="msedge",
            args=["--disable-features=RendererCodeIntegrity"]
        )
        context = browser.new_context(
            storage_state=str(BROWSER_STATE) if BROWSER_STATE.exists() else None,
            permissions=["clipboard-read", "clipboard-write"],
        )
        page = context.new_page()
        page.goto(FEISHU_DOC_URL, wait_until="networkidle")
        time.sleep(5)

        # 检查登录
        if "login" in page.url or "passport" in page.url:
            print("  [INFO] 需要登录飞书，请手动登录...")
            page.wait_for_url("**/docx/**", timeout=120000)
            time.sleep(5)

        # 进入编辑模式
        editable = page.locator('[data-content-editable-root="true"]')
        if editable.count() > 0:
            editable.first.click()
            time.sleep(1)
            print("  [OK] 已进入编辑模式")
        else:
            page.locator('.doc-block-wrapper, .docx-editor').first.click()
            time.sleep(1)

        # ── 粘贴主表 ──
        print("\n[3/4] 粘贴主表...")
        click_placeholder_and_paste(page, main_html, "主表")

        time.sleep(3)

        # ── 粘贴 AI 表 ──
        print("\n[4/4] 粘贴 AI 表...")
        click_placeholder_and_paste(page, ai_html, "AI 表")

        # 等待保存
        print("\n  等待文档保存...")
        time.sleep(10)

        # 保存浏览器状态
        context.storage_state(path=str(BROWSER_STATE))
        browser.close()

    print("\n[DONE] 完成！请检查飞书文档：")
    print(f"  {FEISHU_DOC_URL}")


if __name__ == "__main__":
    main()


if __name__ == "__main__":
    main()
