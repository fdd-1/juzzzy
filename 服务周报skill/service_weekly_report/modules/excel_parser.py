"""通用多行表头 Excel 解析器"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


def parse_multi_header_excel(path: Path, key_column: str = None) -> pd.DataFrame:
    """解析多行表头的 Excel。

    Args:
        path: Excel 文件路径
        key_column: 关键列名（用于定位表头行，如 "团队"、"LP" 等）

    Returns:
        解析后的 DataFrame
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # 读取所有行
    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # 找表头：找到含关键列名的行（如果指定），否则找第一个含实际列名的行
    header_idx = None
    if key_column:
        for i, row in enumerate(all_rows):
            if any(key_column in str(cell) for cell in row if cell):
                header_idx = i
                break

    # 如果没找到或没指定关键列，尝试找第一个非空且看起来像表头的行
    if header_idx is None:
        for i, row in enumerate(all_rows):
            non_empty = [c for c in row if c is not None and str(c).strip()]
            # 表头特征：至少3个非空单元格，且包含中文或常见列名
            if len(non_empty) >= 3:
                header_idx = i
                break

    if header_idx is None:
        raise ValueError("找不到表头行")

    # 判断是单行还是多行表头
    # 如果下一行也有很多非空单元格，可能是二级表头
    row1 = all_rows[header_idx]
    row2 = all_rows[header_idx + 1] if header_idx + 1 < len(all_rows) else []

    # 统计下一行非空单元格数
    row2_non_empty = len([c for c in row2 if c is not None and str(c).strip()])
    row1_non_empty = len([c for c in row1 if c is not None and str(c).strip()])

    is_multi_header = row2_non_empty >= row1_non_empty * 0.5  # 如果第二行非空单元格数 >= 第一行的50%，认为是多行表头

    if is_multi_header:
        # 多行表头：合并 row1 和 row2
        columns = []
        current_group = None
        for i, (v1, v2) in enumerate(zip(row1, row2)):
            v1_str = str(v1).strip() if v1 is not None else ""
            v2_str = str(v2).strip() if v2 is not None else ""

            if v1_str and v1_str not in ['nan', 'None']:
                current_group = v1_str
                if v2_str and v2_str not in ['nan', 'None']:
                    columns.append(f"{v1_str}_{v2_str}")
                else:
                    columns.append(v1_str)
            elif v2_str and v2_str not in ['nan', 'None']:
                if current_group:
                    columns.append(f"{current_group}_{v2_str}")
                else:
                    columns.append(v2_str)
            else:
                columns.append(f"col_{i}")

        data_start = header_idx + 2
    else:
        # 单行表头
        columns = []
        for i, v in enumerate(row1):
            v_str = str(v).strip() if v is not None else ""
            if v_str and v_str not in ['nan', 'None']:
                columns.append(v_str)
            else:
                columns.append(f"col_{i}")

        data_start = header_idx + 1

    # 提取数据
    rows = []
    for r in range(data_start, len(all_rows)):
        row = all_rows[r]
        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        # 跳过明显的口径说明行
        first_cell = str(row[0]) if row[0] is not None else ""
        if any(kw in first_cell for kw in ["口径", "说明", "注：", "备注"]):
            break
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # 过滤掉所有列都为空的行
    df = df.dropna(how='all').reset_index(drop=True)

    return df


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--key-col", help="关键列名（用于定位表头）")
    args = parser.parse_args()

    df = parse_multi_header_excel(Path(args.input), args.key_col)

    print(f"解析结果: {df.shape}")
    print(f"列名: {list(df.columns)[:10]}")

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(args.output, index=False)
    print(f"已保存: {args.output}")
