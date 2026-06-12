"""处理 4.5 服务池 SOP 数据

需求：
1. 只保留"服务池"部分的数据（列34-40）
2. 列序：小组 → 负责人 → LP → 语义点执行率加和 → 其他服务池指标
3. 数据转为数字格式
4. 百分比列转百分比格式
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass


def process_fuwuchi_sop(input_path: Path, output_path: Path) -> pd.DataFrame:
    """提取服务池SOP数据。"""
    print("\n=== 4.5 服务池SOP数据处理 ===")

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    # 读取所有行
    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # 找到"服务池"在row2中的位置
    row2 = all_rows[1]  # row2 = index 1
    fuwuchi_start = None
    for i, v in enumerate(row2):
        if v == '服务池':
            fuwuchi_start = i
            break

    if fuwuchi_start is None:
        raise ValueError("找不到'服务池'列")

    print(f"  服务池起始列: {fuwuchi_start + 1}")

    # 服务池的列：34-40 (index 33-39)
    # row4 是最后一层表头
    row4 = all_rows[3]

    # 构造列名：团队列(1-3) + 服务池列(33-39)
    base_cols = ['col_0', '小组', '负责人', 'LP']
    fuwuchi_cols = []
    for i in range(fuwuchi_start, min(fuwuchi_start + 10, len(row2))):
        v = all_rows[2][i]  # row3
        if v:
            fuwuchi_cols.append(f"服务池_{v}")

    columns = base_cols + fuwuchi_cols

    # 数据从row5开始
    data_start = 4
    rows = []
    for r in range(data_start, len(all_rows)):
        row = all_rows[r]
        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        # 口径行
        first_cell = str(row[1]) if len(row) > 1 and row[1] else ""
        if any(kw in first_cell for kw in ["口径", "说明", "注："]):
            break

        # 提取：列1-3（小组/负责人/LP） + 列33-39（服务池）
        rec = row[0:4] + row[fuwuchi_start:fuwuchi_start + len(fuwuchi_cols)]
        rows.append(rec)

    df = pd.DataFrame(rows, columns=columns)

    # 过滤空行
    df = df[df['小组'].notna()].reset_index(drop=True)

    # 找到"语义点执行率加和"列
    jiahuo_col = None
    for col in df.columns:
        if '语义点执行率加和' in str(col) or '执行率加和' in str(col):
            jiahuo_col = col
            break

    # 重排列：小组 → 负责人 → LP → 语义点执行率加和 → 其他
    if jiahuo_col:
        other_cols = [c for c in df.columns if c not in ['col_0', '小组', '负责人', 'LP', jiahuo_col]]
        final_cols = ['小组', '负责人', 'LP', jiahuo_col] + other_cols
    else:
        final_cols = ['小组', '负责人', 'LP'] + [c for c in df.columns if c not in ['col_0', '小组', '负责人', 'LP']]

    df = df[[c for c in final_cols if c in df.columns]]

    # 数据转数字
    for col in df.columns:
        if col not in ['小组', '负责人', 'LP']:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # 百分比列转百分比格式
    for col in df.columns:
        if '执行率' in str(col) or '占比' in str(col):
            df[col] = df[col].apply(lambda v: f"{v*100:.2f}%" if pd.notna(v) and isinstance(v, (int, float)) else v)

    # 输出
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)

    print(f"\n  整合输出: {output_path}")
    print(f"  rows={len(df)}, cols={len(df.columns)}")

    return df


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    process_fuwuchi_sop(Path(args.input), Path(args.output))
