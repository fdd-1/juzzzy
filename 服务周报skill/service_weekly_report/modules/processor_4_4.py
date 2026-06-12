"""处理 4.4 停课唤醒数据

需求：
1. 正确解析表头和数据
2. 数据转为数字格式
3. 百分比列转百分比格式（保留两位小数）
4. 识别"停课占比"和"唤醒率"列应用色阶
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass


def process_tingke_huanxing(input_path: Path, output_path: Path) -> pd.DataFrame:
    """处理停课唤醒数据。"""
    print("\n=== 4.4 停课唤醒数据处理 ===")

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    # 读取所有行
    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # row 4 是表头
    header_row = 3  # index 3 = row 4
    headers = all_rows[header_row]

    # 清理表头
    columns = []
    for i, h in enumerate(headers):
        if h and str(h).strip():
            columns.append(str(h).strip())
        else:
            columns.append(f"col_{i}")

    print(f"  列数: {len(columns)}")
    print(f"  表头示例: {columns[:10]}")

    # 数据从 row 7 开始 (index 6)
    data_start = 6
    rows = []
    for r in range(data_start, len(all_rows)):
        row = all_rows[r]
        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        # 口径行
        if len(row) > 1 and row[1]:
            first_cell = str(row[1])
            if any(kw in first_cell for kw in ["口径", "说明", "注："]):
                break
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # 找到组别和LP列
    group_col = 'lp组别'
    lp_col = 'LP个人'

    # 过滤空行
    if group_col in df.columns:
        df = df[df[group_col].notna()].reset_index(drop=True)

    print(f"  数据行数: {len(df)}")

    # 数据转数字
    for col in df.columns:
        if col not in [group_col, lp_col, 'col_0']:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # 删除全空列
    df = df.dropna(axis=1, how='all')

    # 删除数据全为空的列（只有表头没数据）
    for col in df.columns:
        if col not in [group_col, lp_col, 'col_0']:
            if df[col].notna().sum() == 0:
                df = df.drop(columns=[col])

    print(f"  删除空列后: {len(df.columns)} 列")

    # 百分比列转百分比格式
    percent_keywords = ['占比', '率', '比例']
    for col in df.columns:
        if any(kw in str(col) for kw in percent_keywords):
            # 检查是否已经是百分比形式（0-1范围）
            df[col] = df[col].apply(
                lambda v: f"{v*100:.2f}%" if pd.notna(v) and isinstance(v, (int, float)) and 0 <= v <= 1
                else (f"{v:.2f}%" if pd.notna(v) and isinstance(v, (int, float)) else v)
            )

    # 输出
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)

    print(f"\n  整合输出: {output_path}")
    print(f"  rows={len(df)}, cols={len(df.columns)}")

    return df


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    process_tingke_huanxing(Path(args.input), Path(args.output))
