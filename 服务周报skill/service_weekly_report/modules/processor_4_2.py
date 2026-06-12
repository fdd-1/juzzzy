"""4.2 板块数据处理 — 组班意向

输入: 思维LP组班意向提交播报.xlsx
输出: _merged_4_2.xlsx
处理逻辑: 在LP左侧新增3列汇总数据
"""
from __future__ import annotations
import sys
from pathlib import Path
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass

import pandas as pd
import numpy as np


def process_4_2(input_path: Path, output_path: Path) -> pd.DataFrame:
    """处理组班意向报表。"""
    print("\n=== 4.2 组班意向数据处理 ===")

    # 用 openpyxl 手动解析
    from openpyxl import load_workbook
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    # 读取所有行
    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # 找到表头行（含'团队/小组'）
    header_idx = None
    for i, row in enumerate(all_rows):
        if '团队/小组' in str(row):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("找不到表头行")

    # row1 = 一级表头, row2 = 二级表头
    row1 = all_rows[header_idx]
    row2 = all_rows[header_idx + 1]

    # 构造列名
    columns = []
    current_group = None
    for i, (v1, v2) in enumerate(zip(row1, row2)):
        if v1 and str(v1).strip():
            current_group = str(v1).strip()
            if v2 and str(v2).strip():
                columns.append(f"{current_group}_{v2}")
            else:
                columns.append(current_group)
        elif v2 and str(v2).strip():
            if current_group:
                columns.append(f"{current_group}_{v2}")
            else:
                columns.append(str(v2).strip())
        else:
            columns.append(f"col_{i}")

    # 数据从 header_idx + 2 开始
    data_start = header_idx + 2
    rows = []
    for r in range(data_start, len(all_rows)):
        row = all_rows[r]
        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    print(f"  原始数据: {df.shape}")
    print(f"  列名前10: {list(df.columns)[:10]}")

    # 找到团队和LP列
    team_col = '团队/小组'
    lp_col = 'LP'

    if team_col not in df.columns or lp_col not in df.columns:
        print(f"  [错误] 找不到必需列, 所有列名: {df.columns.tolist()}")
        raise ValueError("找不到团队/小组或LP列")

    # 找到所有"当前意向等待学员数"和"2个意向及以上学员占比"列
    waiting_cols = [c for c in df.columns if '当前意向等待学员数' in str(c)]
    ratio_cols = [c for c in df.columns if '2个意向及以上学员占比' in str(c)]

    print(f"  等待学员数列: {len(waiting_cols)} 个")
    print(f"  占比列: {len(ratio_cols)} 个")

    # 转为数值
    for col in waiting_cols + ratio_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    # 计算汇总
    df['汇总_当前意向等待学员数'] = df[waiting_cols].sum(axis=1)

    # 加权占比汇总
    weighted = pd.Series(0.0, index=df.index)
    for wait_col, ratio_col in zip(waiting_cols, ratio_cols):
        weighted += df[wait_col].fillna(0) * df[ratio_col].fillna(0)

    df['汇总_2个意向及以上占比加权和'] = weighted
    df['汇总_多意向占比'] = df.apply(
        lambda row: row['汇总_2个意向及以上占比加权和'] / row['汇总_当前意向等待学员数']
        if row['汇总_当前意向等待学员数'] > 0 else 0,
        axis=1
    )

    # 重排列
    summary_cols = ['汇总_当前意向等待学员数', '汇总_2个意向及以上占比加权和', '汇总_多意向占比']
    other_cols = [c for c in df.columns if c not in [team_col, lp_col] + summary_cols]
    final_cols = [team_col] + summary_cols + [lp_col] + other_cols
    df = df[final_cols]

    # 过滤空行
    df = df[df[team_col].notna()].reset_index(drop=True)

    # 输出
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)

    print(f"\n  整合输出: {output_path}")
    print(f"  rows={len(df)}, cols={len(df.columns)}")

    return df


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="输入报表路径")
    parser.add_argument("--output", required=True, help="输出路径")
    args = parser.parse_args()

    process_4_2(Path(args.input), Path(args.output))
