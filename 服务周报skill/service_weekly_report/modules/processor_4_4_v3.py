"""4.4 停课唤醒数据处理 - 正确处理多级表头

表头结构:
- row 4: 一级表头（lp组别/LP个人/停课占比/.../停课90天内唤醒/停课90天以上唤醒/当月新增停课待唤醒）
- row 5: 二级表头（在合并单元格下面，比如停课90天内唤醒-停课数/唤醒率等）
- 数据从 row 7 开始
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
import warnings
warnings.filterwarnings('ignore')

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass

sys.path.insert(0, str(Path(__file__).parent))
from data_formatter import format_dataframe, remove_empty_and_caliber

from openpyxl import load_workbook


def process_4_4_v3(input_path: Path, output_path: Path) -> pd.DataFrame:
    """4.4 停课唤醒（多级表头正确解析）"""
    print("\n=== 4.4 停课唤醒（v3）===")

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # 找到 'lp组别' 行
    header_idx = None
    for i, row in enumerate(all_rows):
        if 'lp组别' in [str(v) for v in row if v]:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("找不到 lp组别 表头")

    # row index = header_idx (一级), header_idx+1 (二级)
    row1 = all_rows[header_idx]
    row2 = all_rows[header_idx + 1] if header_idx + 1 < len(all_rows) else []

    # 构造列名: 一级 + (一级_二级)
    columns = []
    current_group = None
    for i in range(len(row1)):
        v1 = str(row1[i]).strip() if row1[i] is not None else ""
        v2 = str(row2[i]).strip() if i < len(row2) and row2[i] is not None else ""

        # 跳过 nan
        v1 = "" if v1 == "nan" else v1
        v2 = "" if v2 == "nan" else v2

        if v1:
            current_group = v1
            if v2:
                # 同时有一级和二级 → 一级_二级（如"停课占比"和子列同名时取一级）
                columns.append(f"{v1}_{v2}" if v1 != v2 else v1)
            else:
                columns.append(v1)
        elif v2:
            # 只有二级 → current_group_二级
            columns.append(f"{current_group}_{v2}" if current_group else v2)
        else:
            columns.append(f"col_{i}")

    # 数据从 header_idx + 2 开始（跳过两行表头之后可能有的空行 row 6）
    data_start = header_idx + 2
    rows = []
    for r in range(data_start, len(all_rows)):
        row = all_rows[r]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        # 检查是否口径行
        first_cell = str(row[1]) if len(row) > 1 and row[1] else ""
        if any(kw in first_cell for kw in ["口径", "说明", "注："]):
            break
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # 删除 col_X
    df = df.drop(columns=[c for c in df.columns if c.startswith('col_')])

    # 排除 LP个人 是 "停课数"/"唤醒率" 等子表头残留
    if 'LP个人' in df.columns:
        invalid = ['停课数', '唤醒率', '唤醒数', '停课占比', '外呼', '企微信息']
        df = df[~df['LP个人'].astype(str).isin(invalid)]

    # forward-fill lp组别（合并单元格场景：个人 LP 行的 lp组别 是 None）
    if 'lp组别' in df.columns:
        df['lp组别'] = df['lp组别'].ffill()

    # 删除 LP个人 为空的行
    if 'LP个人' in df.columns:
        df = df[df['LP个人'].notna()].reset_index(drop=True)

    # 删除空白列、口径
    df = remove_empty_and_caliber(df, key_col='lp组别')

    # 应用格式化
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return df


if __name__ == "__main__":
    import argparse
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from _paths import PROJECT_ROOT  # noqa: E402

    parser = argparse.ArgumentParser(description="4.4 停课唤醒 v3 处理器（默认使用本周项目目录下数据）")
    parser.add_argument(
        "--input",
        default=str(PROJECT_ROOT.parent / "6.1-.6.7周报数据" / "思维停课学员执行监控.xlsx"),
        help="输入 xlsx",
    )
    parser.add_argument(
        "--output",
        default=str(PROJECT_ROOT / "exports" / "weekly_20260601_20260607" / "4_4" / "_merged_4_4_v3.xlsx"),
        help="输出 xlsx",
    )
    args = parser.parse_args()

    df = process_4_4_v3(Path(args.input), Path(args.output))
    print('\n列名（前30）:')
    for c in list(df.columns)[:30]:
        print(f'  {c}')
