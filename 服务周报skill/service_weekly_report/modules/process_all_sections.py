"""统一处理所有板块的数据，应用格式化规则后保存到 Excel"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
import warnings
warnings.filterwarnings('ignore')

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from data_formatter import format_dataframe, remove_empty_and_caliber
from excel_parser import parse_multi_header_excel
from openpyxl import load_workbook
from _paths import PROJECT_ROOT  # noqa: E402


# 默认源数据 / 输出目录（基于项目根推算，可被 CLI 覆盖）
SOURCE_DIR = PROJECT_ROOT.parent / "6.1-.6.7周报数据"
OUTPUT_BASE = PROJECT_ROOT / "exports" / "weekly_20260601_20260607"


def process_4_2_组班意向(input_path: Path, output_path: Path) -> pd.DataFrame:
    """4.2 组班意向"""
    print("\n=== 4.2 组班意向 ===")

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # 找表头（含'团队/小组'）
    header_idx = None
    for i, row in enumerate(all_rows):
        if '团队/小组' in str(row):
            header_idx = i
            break

    row1 = all_rows[header_idx]
    row2 = all_rows[header_idx + 1]

    # 构造列名
    columns = []
    current_group = None
    for i, (v1, v2) in enumerate(zip(row1, row2)):
        v1_s = str(v1).strip() if v1 else ""
        v2_s = str(v2).strip() if v2 else ""
        if v1_s and v1_s != "nan":
            current_group = v1_s
            columns.append(f"{v1_s}_{v2_s}" if v2_s and v2_s != "nan" else v1_s)
        elif v2_s and v2_s != "nan":
            columns.append(f"{current_group}_{v2_s}" if current_group else v2_s)
        else:
            columns.append(f"col_{i}")

    # 数据
    rows = []
    for r in range(header_idx + 2, len(all_rows)):
        row = all_rows[r]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # 找team和lp列
    team_col = '团队/小组'
    lp_col = 'LP'

    # 计算汇总列
    waiting_cols = [c for c in df.columns if '当前意向等待学员数' in str(c)]
    ratio_cols = [c for c in df.columns if '2个意向及以上学员占比' in str(c)]

    for col in waiting_cols + ratio_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df['汇总_当前意向等待学员数'] = df[waiting_cols].sum(axis=1)
    weighted = pd.Series(0.0, index=df.index)
    for wc, rc in zip(waiting_cols, ratio_cols):
        weighted += df[wc].fillna(0) * df[rc].fillna(0)
    df['汇总_2个意向及以上学员占比'] = df.apply(
        lambda r: weighted[r.name] / r['汇总_当前意向等待学员数'] if r['汇总_当前意向等待学员数'] > 0 else 0,
        axis=1
    )
    df['汇总_多意向占比'] = df['汇总_2个意向及以上学员占比']

    # 重排
    summary_cols = ['汇总_当前意向等待学员数', '汇总_2个意向及以上学员占比', '汇总_多意向占比']
    other_cols = [c for c in df.columns if c not in [team_col, lp_col] + summary_cols and not c.startswith('col_')]
    df = df[[team_col] + summary_cols + [lp_col] + other_cols]

    # 删除空白行/列/口径
    df = remove_empty_and_caliber(df, key_col=team_col)

    # 应用格式化
    df = format_dataframe(df)

    # 输出
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return df


def process_4_3_群发消息(input_path: Path, output_path: Path) -> pd.DataFrame:
    """4.3 群发消息"""
    print("\n=== 4.3 群发消息 ===")
    df = parse_multi_header_excel(input_path, key_column='小组')
    # 删除 col_0
    df = df.drop(columns=[c for c in df.columns if c.startswith('col_')])
    # forward-fill 小组
    if '小组' in df.columns:
        df['小组'] = df['小组'].ffill()
    df = remove_empty_and_caliber(df, key_col='小组')
    df = format_dataframe(df)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return df


def process_4_4_停课唤醒(input_path: Path, output_path: Path) -> pd.DataFrame:
    """4.4 停课唤醒"""
    print("\n=== 4.4 停课唤醒 ===")

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # 多级表头：找 'lp组别'
    header_idx = None
    for i, row in enumerate(all_rows):
        if 'lp组别' in [str(v) for v in row if v]:
            header_idx = i
            break

    # 上面的行可能有大组（停课90天内唤醒等）
    # 检查 row above
    row_above = all_rows[header_idx - 1] if header_idx > 0 else []
    row_2above = all_rows[header_idx - 2] if header_idx > 1 else []

    # 用 row_above 作为大组，header_idx 作为子列
    headers_main = all_rows[header_idx]

    columns = []
    current_group = None
    for i, h in enumerate(headers_main):
        h_s = str(h).strip() if h else ""
        # 大组：row_above 中的非空值
        big = str(row_above[i]).strip() if i < len(row_above) and row_above[i] else ""
        if big and big != "nan":
            current_group = big
        if h_s and h_s != "nan":
            if current_group and current_group != h_s:
                columns.append(f"{current_group}-{h_s}")
            else:
                columns.append(h_s)
        else:
            columns.append(f"col_{i}")

    # 数据从 header_idx + 1 开始（跳过空行）
    data_start = header_idx + 1
    rows = []
    for r in range(data_start, len(all_rows)):
        row = all_rows[r]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # 删 col_X
    df = df.drop(columns=[c for c in df.columns if c.startswith('col_')])
    df = remove_empty_and_caliber(df, key_col='lp组别')
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return df


def process_4_5_服务月跟进(input_path: Path, output_path: Path) -> pd.DataFrame:
    """4.5 服务月跟进 - 只取服务池数据"""
    print("\n=== 4.5 服务月跟进 ===")

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    row3 = all_rows[2]
    row4 = all_rows[3]

    fuwuchi_start = None
    for i, v in enumerate(row3):
        if v == '服务池':
            fuwuchi_start = i
            break

    # 服务池一直到下一个大组
    fuwuchi_end = fuwuchi_start
    for i in range(fuwuchi_start + 1, len(row3)):
        if row3[i] is not None:
            fuwuchi_end = i
            break

    if fuwuchi_end == fuwuchi_start:
        fuwuchi_end = fuwuchi_start + 13  # 默认服务池有13列

    # 列名
    columns = ['小组', 'LP']
    for i in range(fuwuchi_start, fuwuchi_end):
        v = row4[i]
        if v:
            columns.append(f"服务池-{v}")

    # 数据
    rows = []
    for r in range(4, len(all_rows)):
        row = all_rows[r]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        if len(row) > 1 and row[1]:
            first = str(row[1])
            if any(kw in first for kw in ["口径", "说明", "注："]):
                break
        rec = [row[1], row[2]] + row[fuwuchi_start:fuwuchi_start + len(columns) - 2]
        rows.append(rec)

    df = pd.DataFrame(rows, columns=columns)

    # 排除大区总计
    df = df[df['小组'].notna() | df['LP'].notna()]
    # forward-fill 小组（合并单元格）
    df['小组'] = df['小组'].ffill()
    exclude = ['台湾', '欧美澳']
    df = df[~df['小组'].astype(str).isin(exclude)]
    df = df[df['LP'].notna()]
    df = df.reset_index(drop=True)

    df = remove_empty_and_caliber(df, key_col='小组')
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return df


def process_4_5_服务池SOP(input_path: Path, output_path: Path) -> pd.DataFrame:
    """4.5 服务池SOP - 只保留服务池部分，加和放在LP右边"""
    print("\n=== 4.5 服务池SOP ===")

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    row2 = all_rows[1]  # 大组
    row3 = all_rows[2]  # 子列

    fuwuchi_start = None
    for i, v in enumerate(row2):
        if v == '服务池':
            fuwuchi_start = i
            break

    # 服务池一直到下一个大组
    fuwuchi_end = len(row2)
    for i in range(fuwuchi_start + 1, len(row2)):
        if row2[i] is not None:
            fuwuchi_end = i
            break

    # 列名: 小组 / 负责人 / LP / 服务池子列
    columns = ['小组', '负责人', 'LP']
    fuwuchi_cols = []
    for i in range(fuwuchi_start, fuwuchi_end):
        v = row3[i]
        if v:
            fuwuchi_cols.append(f"服务池-{v}")
    columns.extend(fuwuchi_cols)

    # 数据
    rows = []
    for r in range(4, len(all_rows)):
        row = all_rows[r]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        if len(row) > 1 and row[1]:
            first = str(row[1])
            if any(kw in first for kw in ["口径", "说明", "注："]):
                break
        rec = [row[1], row[2], row[3]] + row[fuwuchi_start:fuwuchi_start + len(fuwuchi_cols)]
        rows.append(rec)

    df = pd.DataFrame(rows, columns=columns)

    # 把"语义点执行率加和"列移到LP右边
    jiahuo_col = None
    for c in df.columns:
        if '语义点执行率加和' in str(c) or '执行率加和' in str(c):
            jiahuo_col = c
            break

    if jiahuo_col:
        other_cols = [c for c in df.columns if c not in ['小组', '负责人', 'LP', jiahuo_col]]
        df = df[['小组', '负责人', 'LP'] + [jiahuo_col] + other_cols]

    # forward-fill 小组和负责人（合并单元格）
    df['小组'] = df['小组'].ffill()
    df['负责人'] = df['负责人'].ffill()

    df = df[df['LP'].notna()].reset_index(drop=True)
    df = remove_empty_and_caliber(df, key_col='小组')
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return df


def process_4_6_外呼监控(input_path: Path, output_path: Path) -> pd.DataFrame:
    """4.6 外呼监控"""
    print("\n=== 4.6 外呼监控 ===")
    df = parse_multi_header_excel(input_path, key_column='小组')
    df = df.drop(columns=[c for c in df.columns if c.startswith('col_')])
    # forward-fill 小组（合并单元格）
    if '小组' in df.columns:
        df['小组'] = df['小组'].ffill()
    df = remove_empty_and_caliber(df, key_col='小组')
    df = format_dataframe(df)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return df


def process_4_6_企微回复(input_path: Path, output_path: Path) -> pd.DataFrame:
    """4.6 企微回复"""
    print("\n=== 4.6 企微回复 ===")
    df = parse_multi_header_excel(input_path, key_column='小组')
    df = df.drop(columns=[c for c in df.columns if c.startswith('col_')])
    # forward-fill 当前小组
    if '当前小组' in df.columns:
        df['当前小组'] = df['当前小组'].ffill()
    if 'LP姓名' in df.columns:
        df['LP姓名'] = df['LP姓名'].ffill()
    df = remove_empty_and_caliber(df, key_col='当前小组')
    df = format_dataframe(df)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return df


if __name__ == "__main__":
    # 4.2
    process_4_2_组班意向(
        SOURCE_DIR / "思维LP组班意向提交播报.xlsx",
        OUTPUT_BASE / "4_2" / "_merged_4_2_v2.xlsx"
    )

    # 4.3
    process_4_3_群发消息(
        SOURCE_DIR / "思维海外群发消息汇总数据播报.xlsx",
        OUTPUT_BASE / "4_3" / "_merged_4_3_v2.xlsx"
    )

    # 4.4
    process_4_4_停课唤醒(
        SOURCE_DIR / "思维停课学员执行监控.xlsx",
        OUTPUT_BASE / "4_4" / "_merged_4_4_v2.xlsx"
    )

    # 4.5 服务月跟进
    process_4_5_服务月跟进(
        SOURCE_DIR / "思维转介绍过程跟进报表_末次渠道.xlsx",
        OUTPUT_BASE / "4_5" / "_merged_4_5_fuwuyue_v2.xlsx"
    )

    # 4.5 服务池SOP
    process_4_5_服务池SOP(
        SOURCE_DIR / "海外思维服务SOP执行情况.xlsx",
        OUTPUT_BASE / "4_5" / "_merged_4_5_sop_v2.xlsx"
    )

    # 4.6 外呼
    process_4_6_外呼监控(
        SOURCE_DIR / "LP系统外呼监控-分池子.xlsx",
        OUTPUT_BASE / "4_6" / "_merged_4_6_waihu_v2.xlsx"
    )

    # 4.6 企微
    process_4_6_企微回复(
        SOURCE_DIR / "LP企微回复比监控-分池子.xlsx",
        OUTPUT_BASE / "4_6" / "_merged_4_6_qiwei_v2.xlsx"
    )

    print("\n✓ 所有板块数据处理完成")
