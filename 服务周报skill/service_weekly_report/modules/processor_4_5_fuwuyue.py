"""处理 4.5 服务月跟进（转介绍报表-服务池部分）

需求：
1. 只保留"服务池"部分的数据（列113-125）
2. 只保留海外团队、小组、LP 三个维度
3. 数据转数字格式
4. 百分比列转百分比格式
5. 删除空白列和空白行
6. 删除口径说明
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass


def process_fuwuyue_genjin(input_path: Path, output_path: Path) -> pd.DataFrame:
    """提取服务月跟进-服务池数据。"""
    print("\n=== 4.5 服务月跟进数据处理 ===")

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    # 读取所有行
    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # row3 是一级表头，row4 是二级表头
    row3 = all_rows[2]  # index 2 = row 3
    row4 = all_rows[3]  # index 3 = row 4

    # 服务池从列113开始（index 112）
    fuwuchi_start = 112  # 列113 = index 112

    # 构造列名：小组、LP + 服务池列
    base_cols = ['小组', 'LP']
    fuwuchi_cols = []

    # 服务池列：113-125 (index 112-124)
    for i in range(fuwuchi_start, min(fuwuchi_start + 13, len(row4))):
        v = row4[i]
        if v:
            fuwuchi_cols.append(f"服务池_{v}")

    columns = base_cols + fuwuchi_cols

    # 数据从 row5 开始 (index 4)
    data_start = 4
    rows = []
    for r in range(data_start, len(all_rows)):
        row = all_rows[r]
        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        # 口径行
        if len(row) > 1 and row[1]:
            first_cell = str(row[1])
            if any(kw in first_cell for kw in ["口径", "说明", "注：", "备注"]):
                break

        # 提取：列1-2（小组、LP）+ 列112-124（服务池）
        rec = [row[1], row[2]] + row[fuwuchi_start:fuwuchi_start + len(fuwuchi_cols)]
        rows.append(rec)

    df = pd.DataFrame(rows, columns=columns)

    # 过滤：只保留海外团队、小组总计、个人LP（排除大区总计）
    df = df[df['小组'].notna()].reset_index(drop=True)

    # 排除大区总计行（台湾、欧美澳等）
    exclude_keywords = ['台湾总计', '欧美澳', '港澳总计']
    for keyword in exclude_keywords:
        df = df[~df['小组'].astype(str).str.contains(keyword, na=False)]

    print(f"  过滤后数据行数: {len(df)}")

    # 数据转数字
    for col in df.columns:
        if col not in ['小组', 'LP']:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # 删除全空列
    df = df.dropna(axis=1, how='all')

    # 百分比列转百分比格式
    percent_keywords = ['率', '占比']
    for col in df.columns:
        if any(kw in str(col) for kw in percent_keywords):
            df[col] = df[col].apply(
                lambda v: f"{v*100:.2f}%" if pd.notna(v) and isinstance(v, (int, float)) and 0 <= v <= 1
                else (f"{v:.2f}%" if pd.notna(v) and isinstance(v, (int, float)) else v)
            )

    # 输出
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)

    print(f"\n  整合输出: {output_path}")
    print(f"  rows={len(df)}, cols={len(df.columns)}")

    return df


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    process_fuwuyue_genjin(Path(args.input), Path(args.output))
