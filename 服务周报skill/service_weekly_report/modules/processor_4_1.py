"""4.1 板块数据处理 — 5 份报表整合成一张宽表

输入: exports/weekly_{开始}_{结束}/4_1/<报表id>/*.xlsx
输出: exports/weekly_{开始}_{结束}/4_1/_merged_4_1.xlsx
"""
from __future__ import annotations
import sys
from pathlib import Path
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# UTF-8 输出
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass

import pandas as pd
import numpy as np
from openpyxl import load_workbook


# 9 个有效团队 (报表中可能用'海外教学服务部'代替'海外团队')
VALID_TEAMS = {
    "海外团队", "海外教学服务部", "台湾组",
    "港澳1组", "港澳2组", "港澳组",
    "美澳1组", "美澳2组", "美澳3组", "美澳4组", "美澳5组",
}

# 团队名称映射: 标准化为统一名称
TEAM_NAME_MAP = {
    "海外教学服务部": "海外团队",
}

TEAM_ORDER = ["海外团队", "港澳1组", "港澳2组", "港澳组",
              "美澳1组", "美澳2组", "美澳3组", "美澳4组", "美澳5组", "台湾组"]


def _is_valid_team(team) -> bool:
    if pd.isna(team) or not isinstance(team, str):
        return False
    return team.strip() in VALID_TEAMS


def _is_caliber_row(text) -> bool:
    """识别口径/注释行: 含'口径说明'、'数字)'、'数字、'等字符串。"""
    if not isinstance(text, str):
        return False
    s = text.strip()
    if not s:
        return False
    if s in VALID_TEAMS:
        return False
    # 口径关键词
    if any(kw in s for kw in ["口径", "说明", "注：", "备注"]):
        return True
    # 形如"1)"、"2)"、"1、"开头的注释行
    import re
    if re.match(r"^\d+[\)\、)）.]", s):
        return True
    return False


def _normalize_team(team):
    """统一团队名称(海外教学服务部 → 海外团队)。"""
    if pd.isna(team) or not isinstance(team, str):
        return team
    return TEAM_NAME_MAP.get(team.strip(), team.strip())


def process_report_1_首通(path: Path) -> pd.DataFrame:
    """报表1: 益智海外新生首通监控
    - 找到含'LP小组'的行作为表头, 下一行子列名
    - 仅保留月度大组的列
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    header_idx = None
    for i, row in enumerate(all_rows):
        if "LP小组" in row:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("找不到 LP小组")

    row1 = all_rows[header_idx]
    row2 = all_rows[header_idx + 1]

    month_start = None
    month_end = None
    for i, v in enumerate(row1):
        if v == "月度":
            month_start = i
        elif v is not None and month_start is not None and month_end is None:
            month_end = i
    if month_start is None:
        raise ValueError("找不到月度大组")
    if month_end is None:
        month_end = len(row2)

    sub_cols = row2[month_start:month_end]
    base_idx = {}
    for i, v in enumerate(row1):
        if v == "LP小组":
            base_idx["团队"] = i
        elif v == "LP姓名":
            base_idx["LP"] = i

    data_start = header_idx + 2
    rows = []
    for r in range(data_start, len(all_rows)):
        row = all_rows[r]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rec = {
            "团队": row[base_idx["团队"]] if "团队" in base_idx else None,
            "LP": row[base_idx["LP"]] if "LP" in base_idx else None,
        }
        for j, sub_name in enumerate(sub_cols):
            if sub_name and isinstance(sub_name, str):
                rec[f"首通_{sub_name}"] = row[month_start + j]
        rows.append(rec)

    df = pd.DataFrame(rows)
    # forward-fill 团队列(个人LP行继承所属小组)
    if not df.empty:
        # 截断到第一个口径行之前
        first_caliber = None
        for i, row in df.iterrows():
            if _is_caliber_row(row["团队"]) or _is_caliber_row(row.get("LP")):
                first_caliber = i
                break
        if first_caliber is not None:
            df = df.iloc[:first_caliber].copy()
        df["团队"] = df["团队"].ffill()
        df["团队"] = df["团队"].apply(_normalize_team)
    df = df[df["团队"].apply(_is_valid_team) & df["LP"].notna()].reset_index(drop=True)
    return df


def process_service_report(path: Path, prefix: str) -> pd.DataFrame:
    """报表2/3: 学管服务指标统计表 - row1 是表头(单行)
    - prefix: '首课' 或 '首专'
    - 仅保留以 prefix 开头的列
    - 团队列需 forward-fill (个人LP行的团队列继承上一行小组)
    """
    df = pd.read_excel(path, header=1)
    if "LP名称" in df.columns:
        df = df.rename(columns={"LP名称": "LP"})

    # 先 forward-fill 团队列(个人LP行继承所属小组)
    if "团队" in df.columns:
        # 截断到第一个口径行之前
        first_caliber = None
        for i, row in df.iterrows():
            if _is_caliber_row(row["团队"]):
                first_caliber = i
                break
        if first_caliber is not None:
            df = df.iloc[:first_caliber].copy()
        df["团队"] = df["团队"].ffill()
        df["团队"] = df["团队"].apply(_normalize_team)

    keep_cols = ["团队", "LP"]
    for c in df.columns:
        if isinstance(c, str) and c.startswith(prefix):
            keep_cols.append(c)
    keep_cols = [c for c in keep_cols if c in df.columns]
    df = df[keep_cols]
    # 过滤: 团队是有效团队 + LP 不为空
    df = df[df["团队"].apply(_is_valid_team) & df["LP"].notna()].reset_index(drop=True)
    return df


def process_report_4_SOP(path: Path) -> dict[str, pd.DataFrame]:
    """报表4: SOP执行情况
    - 多行表头: row1=大组(团队/首通/首课/...), row2=子列名, row3=小组/负责人/LP
    - 数据从 row4 起
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    header_idx = None
    for i, row in enumerate(all_rows):
        if "团队" in row:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("找不到 SOP 表头")

    row1 = all_rows[header_idx]
    row2 = all_rows[header_idx + 1]
    row3 = all_rows[header_idx + 2]

    groups = {}
    current = None
    for i, v in enumerate(row1):
        if v in ("团队", "首通", "首课", "停课唤醒", "服务池"):
            if current is not None:
                groups[current[0]] = (current[1], i)
            current = (v, i)
    if current is not None:
        groups[current[0]] = (current[1], len(row1))

    team_start, team_end = groups.get("团队", (0, 0))
    base_idx = {}
    for i in range(team_start, team_end):
        v = row3[i]
        if v == "小组":
            base_idx["小组"] = i
        elif v == "LP":
            base_idx["LP"] = i

    data_start = header_idx + 3

    def extract(group_name: str, prefix: str) -> pd.DataFrame:
        if group_name not in groups:
            return pd.DataFrame()
        gs, ge = groups[group_name]
        sub_cols = row2[gs:ge]
        rows = []
        for r in range(data_start, len(all_rows)):
            row = all_rows[r]
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            rec = {
                "团队": row[base_idx["小组"]] if "小组" in base_idx else None,
                "LP": row[base_idx["LP"]] if "LP" in base_idx else None,
            }
            for j, sub_name in enumerate(sub_cols):
                if sub_name and isinstance(sub_name, str):
                    rec[f"{prefix}_{sub_name}"] = row[gs + j]
            rows.append(rec)
        df = pd.DataFrame(rows)
        if not df.empty:
            # 截断口径行前
            first_caliber = None
            for i, row in df.iterrows():
                if _is_caliber_row(row["团队"]) or _is_caliber_row(row.get("LP")):
                    first_caliber = i
                    break
            if first_caliber is not None:
                df = df.iloc[:first_caliber].copy()
            df["团队"] = df["团队"].ffill()
            df["团队"] = df["团队"].apply(_normalize_team)
            df = df[df["团队"].apply(_is_valid_team) & df["LP"].notna()].reset_index(drop=True)
        return df

    return {
        "首通语义点执行": extract("首通", "首通语义点执行"),
        "首课语义点执行": extract("首课", "首课语义点执行"),
    }


def process_report_5_LP架构(path: Path) -> pd.DataFrame:
    """报表5: LP架构表 - row 3 为表头(大区/团队/.../小组/.../姓名/.../入职时间/.../入职时长分组)
    用'小组' as 团队, '姓名' as LP
    """
    df = pd.read_excel(path, header=3)
    # 先去掉原来的"团队"列(我们用"小组"作为团队)
    if "团队" in df.columns:
        df = df.drop(columns=["团队"])

    rename = {}
    for c in df.columns:
        s = str(c).strip()
        if s == "小组":
            rename[c] = "团队"
        elif s == "姓名":
            rename[c] = "LP"
        elif s == "入职时间":
            rename[c] = "入职时间"
        elif s == "离职日期":
            rename[c] = "离职时间"
        elif s == "入职时长分组":
            rename[c] = "入职时间分组"
    df = df.rename(columns=rename)

    keep = [c for c in ["团队", "LP", "入职时间", "离职时间", "入职时间分组"] if c in df.columns]
    df = df[keep]
    if "LP" in df.columns:
        df = df[df["LP"].notna()].reset_index(drop=True)
    if "团队" in df.columns:
        df["团队"] = df["团队"].apply(_normalize_team)
    # 去重
    df = df.drop_duplicates(subset=["团队", "LP"]).reset_index(drop=True)
    return df


def merge_4_1(report_paths: dict[str, Path], output_path: Path) -> pd.DataFrame:
    """整合5份报表数据,以(团队+LP)为复合主键避免笛卡尔积。"""
    print("\n=== 4.1 数据整合 ===")

    df1 = process_report_1_首通(report_paths["1_首通监控"])
    print(f"  报表1 (首通): {df1.shape}")

    df2 = process_service_report(report_paths["2_服务指标_首课"], "首课")
    print(f"  报表2 (首课): {df2.shape}")

    df3 = process_service_report(report_paths["3_服务指标_首专"], "首专")
    print(f"  报表3 (首专): {df3.shape}")

    sop = process_report_4_SOP(report_paths["4_SOP执行"])
    df4_1 = sop["首通语义点执行"]
    df4_2 = sop["首课语义点执行"]
    print(f"  报表4 (首通语义): {df4_1.shape}")
    print(f"  报表4 (首课语义): {df4_2.shape}")

    # 报表5 (LP架构) - 可选，如果没有则跳过
    if "5_LP架构" in report_paths and report_paths["5_LP架构"]:
        df5 = process_report_5_LP架构(report_paths["5_LP架构"])
        print(f"  报表5 (LP架构): {df5.shape}")
    else:
        df5 = pd.DataFrame()
        print(f"  报表5 (LP架构): 未提供，跳过")

    # 期望列序
    cols_首通语义 = [
        "首通语义点执行_执行率加和",
        "首通语义点执行_拨通且命中首通场景新生数",
        "首通语义点执行_邀请添加企微/WS/Line执行率",
        "首通语义点执行_一家多娃问询执行率",
        "首通语义点执行_转介绍执行率",
    ]
    cols_首通 = [
        "首通_新生数", "首通_勿扰新生数", "首通_一家多娃占比",
        "首通_跟进率", "首通_及时跟进率", "首通_企微绑定率", "首通_秒挂占比",
    ]
    cols_首课语义 = [
        "首课语义点执行_执行率加和",
        "首课语义点执行_询问上课感受执行率",
    ]
    cols_首课 = ["首课学员数", "首课勿扰学员数", "首课跟进率", "首课及时跟进率"]
    cols_首专 = ["首专学员数", "首专勿扰学员数", "首专跟进率", "首专及时跟进率"]
    cols_LP入职 = ["入职时间", "离职时间", "入职时间分组"]

    # 以报表2为基础,LEFT JOIN 其他(用 团队+LP 联合主键)
    def safe_merge(left, right, on, how="left"):
        if right.empty:
            return left
        # 去重 right 的主键 (避免 LP 在多个团队重名导致重复)
        right_dedup = right.drop_duplicates(subset=on)
        # 只取 right 中的新列
        right_cols = [c for c in right_dedup.columns if c in on or c not in left.columns]
        return left.merge(right_dedup[right_cols], on=on, how=how)

    base = df2.copy()
    base = safe_merge(base, df3, on=["团队", "LP"])
    base = safe_merge(base, df1, on=["团队", "LP"])
    base = safe_merge(base, df4_1, on=["团队", "LP"])
    base = safe_merge(base, df4_2, on=["团队", "LP"])
    base = safe_merge(base, df5, on=["团队", "LP"])

    # 重排
    final_cols = ["团队", "LP"] + cols_首通语义 + cols_首通 \
                 + cols_首课语义 + cols_首课 + cols_首专 + cols_LP入职
    for c in final_cols:
        if c not in base.columns:
            base[c] = pd.NA
    base = base[final_cols]

    # 列格式化规则：
    # - 含"加和"的列：保留两位小数，不转百分比（如执行率加和 = 5.35）
    # - 含"执行率/占比/跟进率/绑定率"且不含"加和"：转百分比格式（保留两位小数）
    for col in base.columns:
        if not isinstance(col, str):
            continue
        if "加和" in col:
            # 加和列：保留两位小数（数字格式）
            base[col] = base[col].apply(
                lambda v: round(float(v), 2) if pd.notna(v) and isinstance(v, (int, float)) else v
            )
        elif any(kw in col for kw in ["执行率", "占比", "跟进率", "绑定率"]):
            # 百分比列：转为 "85.73%" 字符串
            base[col] = base[col].apply(
                lambda v: f"{v*100:.2f}%" if pd.notna(v) and isinstance(v, (int, float)) else v
            )

    # 排序: 海外团队总计 → 所有小组总计 → 各小组个人LP
    # 逻辑：
    # 1. 海外团队 + LP==总计 → (0, 0)
    # 2. 所有小组 + LP==总计 → (1, team_order)
    # 3. 各小组个人LP → (2, team_order, LP名)
    def sort_key(row):
        team = row["团队"]
        lp = row["LP"]

        # 海外团队总计 = (0, 0, "")
        if team == "海外团队" and lp == "总计":
            return (0, 0, "")

        # 所有小组总计 = (1, team_order, "")
        if lp == "总计":
            team_idx = TEAM_ORDER.index(team) if team in TEAM_ORDER else 99
            return (1, team_idx, "")

        # 各小组个人LP = (2, team_order, LP名)
        team_idx = TEAM_ORDER.index(team) if team in TEAM_ORDER else 99
        return (2, team_idx, lp)

    base["_sort"] = base.apply(sort_key, axis=1)
    base = base.sort_values("_sort").reset_index(drop=True)
    base = base.drop(columns=["_sort"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    base.to_excel(output_path, index=False)
    print(f"\n  整合输出: {output_path}")
    print(f"  rows={len(base)}, cols={len(base.columns)}")
    return base


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--exports-dir", required=True)
    args = parser.parse_args()

    base_dir = Path(args.exports_dir)
    report_paths = {}
    for sub in base_dir.iterdir():
        if sub.is_dir():
            xlsx = list(sub.glob("*.xlsx"))
            if xlsx:
                report_paths[sub.name] = max(xlsx, key=lambda p: p.stat().st_mtime)

    print("[报表路径]")
    for k, v in report_paths.items():
        print(f"  {k}: {v}")

    output_path = base_dir / "_merged_4_1.xlsx"
    merge_4_1(report_paths, output_path)
