"""服务周报数据处理 - 从 SmartBI 下载目录读取并整合

输入：downloads/smartbi_reports/{run_date}/4_X_xxx/*.xlsx
输出：exports/weekly_{start}_{end}/_merged_4_X.xlsx
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
import warnings
warnings.filterwarnings('ignore')

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass

# 模块路径
MODULES_DIR = Path(__file__).resolve().parent.parent / "modules"
sys.path.insert(0, str(MODULES_DIR))

from data_formatter import format_dataframe, remove_empty_and_caliber
from excel_parser import parse_multi_header_excel
from openpyxl import load_workbook


def find_xlsx(folder: Path) -> Path:
    """从文件夹中找到第一个 xlsx 文件"""
    files = list(folder.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"找不到 xlsx 文件: {folder}")
    return files[0]


def process_4_1_整合(downloads_dir: Path, output_dir: Path) -> Path:
    """4.1 整合 5 份报表（首通+首课+首专+SOP+LP架构）"""
    print("\n=== 4.1 服务指标跟进 & 语义分析 ===")

    # 5 份报表路径
    paths = {
        "1_首通监控": find_xlsx(downloads_dir / "4_1_shoutong"),
        "2_服务指标_首课": find_xlsx(downloads_dir / "4_1_shouke"),
        "3_服务指标_首专": find_xlsx(downloads_dir / "4_1_shouzhuan"),
        "4_SOP执行": find_xlsx(downloads_dir / "4_1_sop"),
    }

    # LP 架构表（如有，可选）
    lp_arch_dir = downloads_dir / "4_1_lp_arch"
    if lp_arch_dir.exists() and list(lp_arch_dir.glob("*.xlsx")):
        paths["5_LP架构"] = find_xlsx(lp_arch_dir)

    # 调用现有的 processor_4_1
    from processor_4_1 import merge_4_1
    output_path = output_dir / "_merged_4_1.xlsx"
    merge_4_1(paths, output_path)

    return output_path


def process_4_2_组班意向(downloads_dir: Path, output_dir: Path) -> Path:
    """4.2 组班意向"""
    print("\n=== 4.2 组班意向 ===")

    input_path = find_xlsx(downloads_dir / "4_2")
    output_path = output_dir / "_merged_4_2.xlsx"

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    # 找表头
    header_idx = None
    for i, row in enumerate(all_rows):
        if '团队/小组' in str(row):
            header_idx = i
            break

    row1 = all_rows[header_idx]
    row2 = all_rows[header_idx + 1]

    # 构造列名
    columns = []
    current_group = None
    for i, (v1, v2) in enumerate(zip(row1, row2)):
        v1_s = str(v1).strip() if v1 else ""
        v2_s = str(v2).strip() if v2 else ""
        if v1_s and v1_s != "nan":
            current_group = v1_s
            columns.append(f"{v1_s}_{v2_s}" if v2_s and v2_s != "nan" else v1_s)
        elif v2_s and v2_s != "nan":
            columns.append(f"{current_group}_{v2_s}" if current_group else v2_s)
        else:
            columns.append(f"col_{i}")

    rows = []
    for r in range(header_idx + 2, len(all_rows)):
        row = all_rows[r]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # 计算汇总列
    waiting_cols = [c for c in df.columns if '当前意向等待学员数' in str(c)]
    ratio_cols = [c for c in df.columns if '2个意向及以上学员占比' in str(c)]

    for col in waiting_cols + ratio_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df['汇总_当前意向等待学员数'] = df[waiting_cols].sum(axis=1)
    weighted = pd.Series(0.0, index=df.index)
    for wc, rc in zip(waiting_cols, ratio_cols):
        weighted += df[wc].fillna(0) * df[rc].fillna(0)
    df['汇总_2个意向及以上学员占比'] = df.apply(
        lambda r: weighted[r.name] / r['汇总_当前意向等待学员数'] if r['汇总_当前意向等待学员数'] > 0 else 0,
        axis=1
    )
    df['汇总_多意向占比'] = df['汇总_2个意向及以上学员占比']

    summary_cols = ['汇总_当前意向等待学员数', '汇总_2个意向及以上学员占比', '汇总_多意向占比']
    other_cols = [c for c in df.columns if c not in ['团队/小组', 'LP'] + summary_cols and not c.startswith('col_')]
    df = df[['团队/小组'] + summary_cols + ['LP'] + other_cols]

    df = remove_empty_and_caliber(df, key_col='团队/小组')
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return output_path


def process_4_3_群发消息(downloads_dir: Path, output_dir: Path) -> Path:
    """4.3 群发消息"""
    print("\n=== 4.3 群发消息 ===")

    input_path = find_xlsx(downloads_dir / "4_3")
    output_path = output_dir / "_merged_4_3.xlsx"

    df = parse_multi_header_excel(input_path, key_column='小组')
    df = df.drop(columns=[c for c in df.columns if c.startswith('col_')])
    if '小组' in df.columns:
        df['小组'] = df['小组'].ffill()
    df = remove_empty_and_caliber(df, key_col='小组')
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return output_path


def process_4_4_停课唤醒(downloads_dir: Path, output_dir: Path) -> Path:
    """4.4 停课唤醒"""
    print("\n=== 4.4 停课唤醒 ===")

    input_path = find_xlsx(downloads_dir / "4_4")
    output_path = output_dir / "_merged_4_4.xlsx"

    # 调用现有的 processor_4_4_v3
    from processor_4_4_v3 import process_4_4_v3
    process_4_4_v3(input_path, output_path)

    return output_path


def process_4_5_服务月跟进(downloads_dir: Path, output_dir: Path) -> Path:
    """4.5 服务月跟进 - 从转介绍报表提取服务池数据"""
    print("\n=== 4.5 服务月跟进 ===")

    input_path = find_xlsx(downloads_dir / "4_5_fuwuyue")
    output_path = output_dir / "_merged_4_5_fuwuyue.xlsx"

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    row3 = all_rows[2]
    row4 = all_rows[3]

    fuwuchi_start = None
    for i, v in enumerate(row3):
        if v == '服务池':
            fuwuchi_start = i
            break

    fuwuchi_end = fuwuchi_start
    for i in range(fuwuchi_start + 1, len(row3)):
        if row3[i] is not None:
            fuwuchi_end = i
            break

    if fuwuchi_end == fuwuchi_start:
        fuwuchi_end = fuwuchi_start + 13

    columns = ['小组', 'LP']
    for i in range(fuwuchi_start, fuwuchi_end):
        v = row4[i]
        if v:
            columns.append(f"服务池-{v}")

    rows = []
    for r in range(4, len(all_rows)):
        row = all_rows[r]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        if len(row) > 1 and row[1]:
            first = str(row[1])
            if any(kw in first for kw in ["口径", "说明", "注："]):
                break
        rec = [row[1], row[2]] + row[fuwuchi_start:fuwuchi_start + len(columns) - 2]
        rows.append(rec)

    df = pd.DataFrame(rows, columns=columns)

    df = df[df['小组'].notna() | df['LP'].notna()]
    df['小组'] = df['小组'].ffill()
    exclude = ['台湾', '欧美澳']
    df = df[~df['小组'].astype(str).isin(exclude)]
    df = df[df['LP'].notna()]
    df = df.reset_index(drop=True)

    df = remove_empty_and_caliber(df, key_col='小组')
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return output_path


def process_4_5_服务池SOP(downloads_dir: Path, output_dir: Path) -> Path:
    """4.5 服务池SOP - 复用 4.1 SOP 文件，提取服务池部分"""
    print("\n=== 4.5 服务池SOP（复用 4.1 SOP）===")

    # 4.5 SOP 与 4.1 SOP 同源
    input_path = find_xlsx(downloads_dir / "4_1_sop")
    output_path = output_dir / "_merged_4_5_sop.xlsx"

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    all_rows = []
    for r in range(1, ws.max_row + 1):
        all_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    row2 = all_rows[1]
    row3 = all_rows[2]

    fuwuchi_start = None
    for i, v in enumerate(row2):
        if v == '服务池':
            fuwuchi_start = i
            break

    fuwuchi_end = len(row2)
    for i in range(fuwuchi_start + 1, len(row2)):
        if row2[i] is not None:
            fuwuchi_end = i
            break

    columns = ['小组', '负责人', 'LP']
    fuwuchi_cols = []
    for i in range(fuwuchi_start, fuwuchi_end):
        v = row3[i]
        if v:
            fuwuchi_cols.append(f"服务池-{v}")
    columns.extend(fuwuchi_cols)

    rows = []
    for r in range(4, len(all_rows)):
        row = all_rows[r]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        if len(row) > 1 and row[1]:
            first = str(row[1])
            if any(kw in first for kw in ["口径", "说明", "注："]):
                break
        rec = [row[1], row[2], row[3]] + row[fuwuchi_start:fuwuchi_start + len(fuwuchi_cols)]
        rows.append(rec)

    df = pd.DataFrame(rows, columns=columns)

    # 把"语义点执行率加和"列移到LP右边
    jiahuo_col = None
    for c in df.columns:
        if '语义点执行率加和' in str(c) or '执行率加和' in str(c):
            jiahuo_col = c
            break

    if jiahuo_col:
        other_cols = [c for c in df.columns if c not in ['小组', '负责人', 'LP', jiahuo_col]]
        df = df[['小组', '负责人', 'LP'] + [jiahuo_col] + other_cols]

    df['小组'] = df['小组'].ffill()
    df['负责人'] = df['负责人'].ffill()

    df = df[df['LP'].notna()].reset_index(drop=True)
    df = remove_empty_and_caliber(df, key_col='小组')
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return output_path


def process_4_6_外呼监控(downloads_dir: Path, output_dir: Path) -> Path:
    """4.6 外呼监控"""
    print("\n=== 4.6 外呼监控 ===")

    input_path = find_xlsx(downloads_dir / "4_6_waihu")
    output_path = output_dir / "_merged_4_6_waihu.xlsx"

    df = parse_multi_header_excel(input_path, key_column='小组')
    df = df.drop(columns=[c for c in df.columns if c.startswith('col_')])
    if '小组' in df.columns:
        df['小组'] = df['小组'].ffill()
    df = remove_empty_and_caliber(df, key_col='小组')
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return output_path


def process_4_6_企微回复(downloads_dir: Path, output_dir: Path) -> Path:
    """4.6 企微回复"""
    print("\n=== 4.6 企微回复 ===")

    input_path = find_xlsx(downloads_dir / "4_6_qiwei")
    output_path = output_dir / "_merged_4_6_qiwei.xlsx"

    df = parse_multi_header_excel(input_path, key_column='小组')
    df = df.drop(columns=[c for c in df.columns if c.startswith('col_')])
    if '当前小组' in df.columns:
        df['当前小组'] = df['当前小组'].ffill()
    if 'LP姓名' in df.columns:
        df['LP姓名'] = df['LP姓名'].ffill()
    df = remove_empty_and_caliber(df, key_col='当前小组')
    df = format_dataframe(df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    print(f"  rows={len(df)}, cols={len(df.columns)}")
    print(f"  → {output_path}")
    return output_path


def process_all(downloads_dir: Path, output_dir: Path) -> dict:
    """处理所有板块的数据"""
    print(f"\n{'='*70}")
    print(f"数据整合与格式化")
    print(f"{'='*70}")
    print(f"输入: {downloads_dir}")
    print(f"输出: {output_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)

    results = {}

    try:
        results['4.1'] = process_4_1_整合(downloads_dir, output_dir)
    except Exception as e:
        print(f"  ⚠ 4.1 失败: {e}")

    try:
        results['4.2'] = process_4_2_组班意向(downloads_dir, output_dir)
    except Exception as e:
        print(f"  ⚠ 4.2 失败: {e}")

    try:
        results['4.3'] = process_4_3_群发消息(downloads_dir, output_dir)
    except Exception as e:
        print(f"  ⚠ 4.3 失败: {e}")

    try:
        results['4.4'] = process_4_4_停课唤醒(downloads_dir, output_dir)
    except Exception as e:
        print(f"  ⚠ 4.4 失败: {e}")

    try:
        results['4.5_fuwuyue'] = process_4_5_服务月跟进(downloads_dir, output_dir)
    except Exception as e:
        print(f"  ⚠ 4.5 服务月失败: {e}")

    try:
        results['4.5_sop'] = process_4_5_服务池SOP(downloads_dir, output_dir)
    except Exception as e:
        print(f"  ⚠ 4.5 SOP失败: {e}")

    try:
        results['4.6_waihu'] = process_4_6_外呼监控(downloads_dir, output_dir)
    except Exception as e:
        print(f"  ⚠ 4.6 外呼失败: {e}")

    try:
        results['4.6_qiwei'] = process_4_6_企微回复(downloads_dir, output_dir)
    except Exception as e:
        print(f"  ⚠ 4.6 企微失败: {e}")

    print(f"\n✓ 处理完成: {len(results)}/8 板块")
    return results


if __name__ == "__main__":
    import argparse
    from datetime import datetime

    parser = argparse.ArgumentParser()
    parser.add_argument("--downloads-dir", required=True, help="SmartBI 下载目录")
    parser.add_argument("--output-dir", required=True, help="处理后的输出目录")
    args = parser.parse_args()

    results = process_all(Path(args.downloads_dir), Path(args.output_dir))
