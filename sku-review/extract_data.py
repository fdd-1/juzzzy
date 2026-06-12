# -*- coding: utf-8 -*-
"""SKU复盘自动化 - 数据提取模块"""

import sys
import re
from pathlib import Path
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from config import FILTERS, PACKAGE_CATEGORIES, POOL_SHEET_NAME, POOL_NODE2_TO_SKU


def map_pool_node_to_sku(node2_value):
    """池子节点2 → SKU节点（升舱/早鸟/其余）"""
    if node2_value is None:
        return "其余"
    s = str(node2_value).strip()
    return POOL_NODE2_TO_SKU.get(s, "其余")


def find_header_row(ws, keyword):
    """自动搜索包含指定关键字的表头行"""
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, min(100, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val and str(val).strip() == keyword:
                return row
    return None


def get_headers(ws, header_row):
    """获取表头字段名到列号的映射"""
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val:
            headers[str(val).strip()] = col
    return headers


def normalize_id(val):
    """标准化ID值（处理float/int/str差异）"""
    if val is None:
        return None
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def classify_package(package_name):
    """根据套餐名称分类"""
    for keyword, category in PACKAGE_CATEGORIES:
        if keyword in package_name:
            return category
    return "其他"


def extract_bi_data(bi_file, filters=None):
    """从BI报表提取订单数据，应用筛选条件"""
    print(f"  读取BI报表: {bi_file}")
    wb = openpyxl.load_workbook(bi_file, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws, "用户ID")
    if not header_row:
        raise ValueError("BI报表中未找到'用户ID'表头行")

    headers = get_headers(ws, header_row)
    print(f"  表头行: {header_row}, 总行数: {ws.max_row}")

    required = ["用户ID", "套餐名称", "用户实际支付金额",
                "课时数（不含积分）", "课时数（含积分）"]
    for field in required:
        if field not in headers:
            raise ValueError(f"BI报表缺少字段: {field}")

    orders = []
    skipped = 0
    for row in range(header_row + 1, ws.max_row + 1):
        if filters:
            skip = False
            for field_name, expected_value in filters.items():
                # 排除语义：字段_排除=[值1, 值2] → 命中任一则跳过
                if field_name.endswith("_排除"):
                    real_field = field_name[:-3]
                    if real_field in headers:
                        val = ws.cell(row, headers[real_field]).value
                        if str(val).strip() in [str(v).strip() for v in expected_value]:
                            skip = True
                            break
                else:
                    if field_name in headers:
                        val = ws.cell(row, headers[field_name]).value
                        if str(val).strip() != expected_value:
                            skip = True
                            break
            if skip:
                skipped += 1
                continue

        user_id = ws.cell(row, headers["用户ID"]).value
        pkg = ws.cell(row, headers["套餐名称"]).value
        amt = ws.cell(row, headers["用户实际支付金额"]).value
        hrs_no = ws.cell(row, headers["课时数（不含积分）"]).value
        hrs_w = ws.cell(row, headers["课时数（含积分）"]).value

        if user_id and pkg:
            orders.append({
                "user_id": normalize_id(user_id),
                "package": str(pkg).strip(),
                "amount": float(amt) if amt else 0,
                "hours_no_integ": float(hrs_no) if hrs_no else 0,
                "hours_with_integ": float(hrs_w) if hrs_w else 0,
                "category": classify_package(str(pkg)),
            })

    wb.close()
    print(f"  筛选后订单: {len(orders)} 条 (排除: {skipped})")
    return orders


def extract_pool_data(pool_file):
    """从正式池提取学员人群属性 + 池子节点2"""
    print(f"  读取正式池: {pool_file}")
    wb = openpyxl.load_workbook(pool_file, data_only=True)

    # 优先使用指定sheet；否则模糊匹配"池内"+"剔"+"不可续"（兼容3月/4月/6月命名差异）
    if POOL_SHEET_NAME in wb.sheetnames:
        ws = wb[POOL_SHEET_NAME]
        print(f"  使用Sheet: {POOL_SHEET_NAME}")
    else:
        fuzzy = [s for s in wb.sheetnames
                 if "池内" in s and "剔" in s and "不可续" in s]
        if fuzzy:
            ws = wb[fuzzy[0]]
            print(f"  使用Sheet(模糊匹配): {fuzzy[0]}")
        else:
            ws = wb.active
            print(f"  使用默认Sheet: {ws.title}")

    header_row = find_header_row(ws, "学员ID")
    if not header_row:
        raise ValueError("正式池中未找到'学员ID'表头行")

    headers = get_headers(ws, header_row)
    print(f"  表头行: {header_row}, 总行数: {ws.max_row}")

    if "当前课包顺序" not in headers:
        raise ValueError("正式池缺少字段: 当前课包顺序")
    if "池子节点2" not in headers:
        print("  ⚠ 正式池缺少'池子节点2'字段，节点将无法识别")

    pool = {}
    for row in range(header_row + 1, ws.max_row + 1):
        sid = ws.cell(row, headers["学员ID"]).value
        order = ws.cell(row, headers["当前课包顺序"]).value
        node2 = ws.cell(row, headers["池子节点2"]).value if "池子节点2" in headers else None
        if sid:
            pool[normalize_id(sid)] = {
                "course_order": order,
                "pool_node2": node2,
            }

    wb.close()
    print(f"  学员数: {len(pool)}")
    return pool


def match_orders_with_pool(orders, pool):
    """匹配订单与学员属性。
    - 续池=0 或 未在正式池 → 剔除
    - SKU节点取自正式池'池子节点2'映射（升舱/早鸟池→升舱/早鸟，其他→其余）
    """
    matched = []
    unmatched = 0
    pool_out = 0

    for order in orders:
        info = pool.get(order["user_id"])
        if info is not None:
            course_order = info["course_order"]
            if course_order == 1:
                cohort = "一续"
            elif course_order is not None and course_order > 1:
                cohort = "多续"
            else:
                pool_out += 1
                continue
            order["cohort"] = cohort
            order["pool_node2"] = info["pool_node2"]
            order["sku_node"] = map_pool_node_to_sku(info["pool_node2"])
            matched.append(order)
        else:
            unmatched += 1

    print(f"  匹配成功: {len(matched)} (一续+多续，用于GMV/课时计算)")
    print(f"  剔除-池外(续池=0): {pool_out}")
    print(f"  剔除-未在正式池: {unmatched}")
    return matched
