# -*- coding: utf-8 -*-
"""SKU复盘自动化 - 分析模块"""

import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from extract_data import find_header_row, get_headers


def aggregate_by_node(matched_orders):
    """按【人群×SKU节点】聚合（节点=升舱/早鸟/其余，来自正式池池子节点2）
    返回包含'综合'汇总（同节点跨人群+同人群跨节点+全量综合）"""
    agg = defaultdict(lambda: {
        "count": 0, "total_amt": 0,
        "total_hrs_no": 0, "total_hrs_w": 0
    })

    for order in matched_orders:
        node = order.get("sku_node", "其余")
        cohort = order["cohort"]
        # (cohort, node)
        agg[(cohort, node)]["count"] += 1
        agg[(cohort, node)]["total_amt"] += order["amount"]
        agg[(cohort, node)]["total_hrs_no"] += order["hours_no_integ"]
        agg[(cohort, node)]["total_hrs_w"] += order["hours_with_integ"]
        # (综合, node) 跨人群同节点
        agg[("综合", node)]["count"] += 1
        agg[("综合", node)]["total_amt"] += order["amount"]
        agg[("综合", node)]["total_hrs_no"] += order["hours_no_integ"]
        agg[("综合", node)]["total_hrs_w"] += order["hours_with_integ"]
        # (cohort, 综合) 同人群跨节点
        agg[(cohort, "综合")]["count"] += 1
        agg[(cohort, "综合")]["total_amt"] += order["amount"]
        agg[(cohort, "综合")]["total_hrs_no"] += order["hours_no_integ"]
        agg[(cohort, "综合")]["total_hrs_w"] += order["hours_with_integ"]
        # (综合, 综合) 全量
        agg[("综合", "综合")]["count"] += 1
        agg[("综合", "综合")]["total_amt"] += order["amount"]
        agg[("综合", "综合")]["total_hrs_no"] += order["hours_no_integ"]
        agg[("综合", "综合")]["total_hrs_w"] += order["hours_with_integ"]

    results = {}
    for (cohort, node), data in agg.items():
        asp = data["total_amt"] / data["count"] if data["count"] else 0
        price_w = data["total_amt"] / data["total_hrs_w"] if data["total_hrs_w"] else 0
        price_no = data["total_amt"] / data["total_hrs_no"] if data["total_hrs_no"] else 0
        results[(cohort, node)] = {
            "count": data["count"],
            "total_amt": round(data["total_amt"], 2),
            "total_hrs_no": round(data["total_hrs_no"], 2),
            "total_hrs_w": round(data["total_hrs_w"], 2),
            "asp": round(asp, 2),
            "price_with_integ": round(price_w, 2),
            "price_no_integ": round(price_no, 2),
        }
    return results


def aggregate_by_category(matched_orders):
    """按【人群×类别】聚合计算指标"""
    agg = defaultdict(lambda: {
        "count": 0, "total_amt": 0,
        "total_hrs_no": 0, "total_hrs_w": 0
    })

    for order in matched_orders:
        key = (order["cohort"], order["category"])
        agg[key]["count"] += 1
        agg[key]["total_amt"] += order["amount"]
        agg[key]["total_hrs_no"] += order["hours_no_integ"]
        agg[key]["total_hrs_w"] += order["hours_with_integ"]

    results = {}
    for (cohort, category), data in agg.items():
        asp = data["total_amt"] / data["count"] if data["count"] else 0
        price_w = data["total_amt"] / data["total_hrs_w"] if data["total_hrs_w"] else 0
        price_no = data["total_amt"] / data["total_hrs_no"] if data["total_hrs_no"] else 0
        results[(cohort, category)] = {
            "count": data["count"],
            "total_amt": data["total_amt"],
            "asp": round(asp, 2),
            "price_with_integ": round(price_w, 2),
            "price_no_integ": round(price_no, 2),
        }

    return results


def aggregate_by_package(matched_orders):
    """按【人群×具体套餐】聚合计算指标和占比"""
    agg = defaultdict(lambda: {
        "count": 0, "total_amt": 0, "category": "", "sku_node": "",
        "total_hrs_no": 0, "total_hrs_w": 0
    })
    cohort_totals = defaultdict(int)

    for order in matched_orders:
        key = (order["cohort"], order["package"])
        agg[key]["count"] += 1
        agg[key]["total_amt"] += order["amount"]
        agg[key]["total_hrs_no"] += order["hours_no_integ"]
        agg[key]["total_hrs_w"] += order["hours_with_integ"]
        agg[key]["category"] = order.get("category", "")
        agg[key]["sku_node"] = order.get("sku_node", "")
        cohort_totals[order["cohort"]] += 1

    results = []
    for (cohort, package), data in agg.items():
        asp = data["total_amt"] / data["count"] if data["count"] else 0
        price_w = data["total_amt"] / data["total_hrs_w"] if data["total_hrs_w"] else 0
        price_no = data["total_amt"] / data["total_hrs_no"] if data["total_hrs_no"] else 0
        ratio = data["count"] / cohort_totals[cohort] if cohort_totals[cohort] else 0

        results.append({
            "cohort": cohort,
            "package": package,
            "category": data["category"],
            "sku_node": data["sku_node"],
            "count": data["count"],
            "total_amt": round(data["total_amt"], 2),
            "asp": round(asp, 2),
            "price_with_integ": round(price_w, 2),
            "price_no_integ": round(price_no, 2),
            "ratio": round(ratio, 4),
        })

    results.sort(key=lambda x: (x["cohort"], -x["count"]))
    return results, cohort_totals


def extract_budget_from_sku(sku_file):
    """从SKU测算文档提取预算指标（自适应两种表结构）。

    结构A（旧 — 4月模板 "SKU测算&实际对比" sheet）:
      含积分块 R5-R8、不含积分块 R13-R16，列 col4/5/8/9/12/13

    结构B（新 — 6月+ "港澳SKU&测算" sheet）:
      测算数据散落在右侧区域（col38-54），通过搜索 "预估ASP" 定位行标题，
      数据行按 升舱/早鸟/其余/全量(综合)/综合 排列。
      col47/48=一续/多续ASP, col49/50=一续/多续含积分单课时价,
      col51/52=一续/多续不含积分单课时价(不含商务)

    返回:
    {
        ('含积分', cohort, node): {'asp': ..., 'price': ...},
        ('不含积分', cohort, node): {...}
    }
    """
    import openpyxl

    # 优先用 read_only 模式快速加载（v3 文件接近 30MB，普通模式可能数分钟）
    wb = openpyxl.load_workbook(sku_file, data_only=True, read_only=True)
    target_sheet = None
    for name in wb.sheetnames:
        if "测算" in name and "对比" in name:
            target_sheet = name
            break
    if not target_sheet:
        for name in wb.sheetnames:
            if "SKU" in name and "测算" in name:
                target_sheet = name
                break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]
    print(f"  SKU测算Sheet: {target_sheet}")

    # read_only 不支持 ws.cell(r,c) 重复访问，缓存前 80 行到 dict
    cache = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True, max_row=80), 1):
        for ci, v in enumerate(row, 1):
            if v is not None:
                cache[(ri, ci)] = v

    def cell_v(r, c):
        return cache.get((r, c))

    budget = {}

    # === 尝试结构B: 搜索 "预估ASP" 关键字定位 ===
    asp_col = None
    asp_row = None
    for row in range(1, 80):
        for col in range(30, 70):
            v = cell_v(row, col)
            if v and "预估ASP" in str(v):
                asp_col = col
                asp_row = row
                break
        if asp_col:
            break

    if asp_col:
        print(f"  [结构B] 找到'预估ASP'在 R{asp_row}C{asp_col}")

        # 搜索两个区域标记：含积分 / 不含积分（在 col38 附近）
        label_col = asp_col - 9  # col38 区域存节点名
        kind_positions = []
        for row in range(1, 80):
            v = cell_v(row, label_col)
            if v and isinstance(v, str) and v.strip() in ("含积分", "不含积分"):
                kind_positions.append((row, v.strip()))

        if not kind_positions:
            kind_positions = [(-1, "含积分")]  # fallback: 只有一块

        node_map = {"升舱": "升舱", "早鸟": "早鸟", "其余": "其余",
                    "全量": "综合", "汇总": "综合", "综合": "综合"}

        for kind_start_row, kind_label in kind_positions:
            # 每块从 kind_start_row 往下找 "预估ASP" header
            search_start = kind_start_row + 1 if kind_start_row > 0 else asp_row
            block_asp_row = None
            for r in range(search_start, min(search_start + 5, 80)):
                v = cell_v(r, asp_col)
                if v and "预估ASP" in str(v):
                    block_asp_row = r
                    break
            if not block_asp_row:
                block_asp_row = search_start  # fallback

            # 数据行在 header 之后 +2（跳过一续/多续 sub-header）
            data_start = block_asp_row + 2
            found_综合 = False
            for r in range(data_start, min(data_start + 10, 80)):
                label = cell_v(r, label_col)
                if not label or not isinstance(label, str):
                    # 综合汇总行: label可能为空但ASP有值
                    asp_v = cell_v(r, asp_col)
                    if isinstance(asp_v, (int, float)) and asp_v > 100 and not found_综合:
                        pr_v = cell_v(r, asp_col + 2)
                        budget[(kind_label, "综合", "综合")] = {
                            "asp": round(float(asp_v), 2),
                            "price": round(float(pr_v), 2) if isinstance(pr_v, (int, float)) else 0,
                        }
                        found_综合 = True
                    continue

                node = None
                for key, mapped in node_map.items():
                    if key in str(label):
                        node = mapped
                        break
                if not node:
                    continue

                asp_1 = cell_v(r, asp_col)
                asp_2 = cell_v(r, asp_col + 1)
                pr_1 = cell_v(r, asp_col + 2)
                pr_2 = cell_v(r, asp_col + 3)

                if isinstance(asp_1, (int, float)) and asp_1 > 100:
                    budget[(kind_label, "一续", node)] = {
                        "asp": round(float(asp_1), 2),
                        "price": round(float(pr_1), 2) if isinstance(pr_1, (int, float)) else 0,
                    }
                if isinstance(asp_2, (int, float)) and asp_2 > 100:
                    budget[(kind_label, "多续", node)] = {
                        "asp": round(float(asp_2), 2),
                        "price": round(float(pr_2), 2) if isinstance(pr_2, (int, float)) else 0,
                    }

    # === 结构A fallback: 固定行位置 ===
    if not budget:
        print(f"  [结构A] 使用固定行列映射")
        node_rows_with = {5: "升舱", 6: "早鸟", 7: "其余", 8: "综合"}
        for row, node in node_rows_with.items():
            cell = cell_v(row, 1)
            if not cell or node not in str(cell):
                continue
            a = cell_v(row, 4)
            p = cell_v(row, 5)
            if isinstance(a, (int, float)) and a > 0:
                budget[("含积分", "一续", node)] = {
                    "asp": round(float(a), 2),
                    "price": round(float(p), 2) if isinstance(p, (int, float)) else 0,
                }
            a = cell_v(row, 8)
            p = cell_v(row, 9)
            if isinstance(a, (int, float)) and a > 0:
                budget[("含积分", "多续", node)] = {
                    "asp": round(float(a), 2),
                    "price": round(float(p), 2) if isinstance(p, (int, float)) else 0,
                }
            a = cell_v(row, 12)
            p = cell_v(row, 13)
            if isinstance(a, (int, float)) and a > 0:
                budget[("含积分", "综合", node)] = {
                    "asp": round(float(a), 2),
                    "price": round(float(p), 2) if isinstance(p, (int, float)) else 0,
                }

        node_rows_no = {13: "升舱", 14: "早鸟", 15: "其余", 16: "综合"}
        for row, node in node_rows_no.items():
            cell = cell_v(row, 1)
            if not cell or node not in str(cell):
                continue
            a = cell_v(row, 4)
            p = cell_v(row, 5)
            if isinstance(a, (int, float)) and a > 0:
                budget[("不含积分", "一续", node)] = {
                    "asp": round(float(a), 2),
                    "price": round(float(p), 2) if isinstance(p, (int, float)) else 0,
                }
            a = cell_v(row, 8)
            p = cell_v(row, 9)
            if isinstance(a, (int, float)) and a > 0:
                budget[("不含积分", "多续", node)] = {
                    "asp": round(float(a), 2),
                    "price": round(float(p), 2) if isinstance(p, (int, float)) else 0,
                }
            a = cell_v(row, 12)
            p = cell_v(row, 13)
            if isinstance(a, (int, float)) and a > 0:
                budget[("不含积分", "综合", node)] = {
                    "asp": round(float(a), 2),
                    "price": round(float(p), 2) if isinstance(p, (int, float)) else 0,
                }

    wb.close()
    print(f"  预算数据: {len(budget)} 个组合")
    return budget


def extract_package_budget_ratio(sku_file):
    """从SKU测算文档提取套餐占比预算（行21开始，列23-34）
    返回 {(cohort, node, package_name): budget_ratio}
    """
    import openpyxl
    wb = openpyxl.load_workbook(sku_file, data_only=True)
    ws = wb['SKU测算&实际对比'] if 'SKU测算&实际对比' in wb.sheetnames else wb.active

    # 列23-34: 一续升舱测/实, 多续升舱测/实, 一续早鸟测/实, 多续早鸟测/实, 一续其余测/实, 多续其余测/实
    col_map = {
        23: ('一续', '升舱'), 24: ('一续', '升舱'),
        25: ('多续', '升舱'), 26: ('多续', '升舱'),
        27: ('一续', '早鸟'), 28: ('一续', '早鸟'),
        29: ('多续', '早鸟'), 30: ('多续', '早鸟'),
        31: ('一续', '其余'), 32: ('一续', '其余'),
        33: ('多续', '其余'), 34: ('多续', '其余'),
    }
    budget_col = [23, 25, 27, 29, 31, 33]  # 测算列

    ratios = {}
    for row in range(24, ws.max_row + 1):
        pkg_name = ws.cell(row, 1).value
        if not pkg_name or not isinstance(pkg_name, str):
            continue
        if '【VIPTHINK】' not in pkg_name:
            continue
        for col in budget_col:
            val = ws.cell(row, col).value
            if isinstance(val, (int, float)) and val > 0:
                cohort, node = col_map[col]
                ratios[(cohort, node, pkg_name)] = float(val)

    wb.close()
    print(f"  套餐占比预算: {len(ratios)} 个")
    return ratios
