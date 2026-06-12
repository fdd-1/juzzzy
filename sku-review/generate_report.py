# -*- coding: utf-8 -*-
"""SKU复盘自动化 - 报告生成模块（按复盘模板格式）"""

import sys
import csv
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


BOLD = Font(bold=True)
BOLD12 = Font(bold=True, size=12)
BOLD14 = Font(bold=True, size=14)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))
YELLOW = PatternFill(start_color="FFF2CC", fill_type="solid")
BLUE = PatternFill(start_color="D9E2F3", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="E2EFDA", fill_type="solid")
LIGHT_RED = PatternFill(start_color="FCE4D6", fill_type="solid")


def write_asp_block(ws, start_row, title, node_results, budget_data, integ_key):
    """写入一个ASP&单价对比块（含积分 或 不含积分）。
    布局参考模板：
    R0: 标题（合并）
    R1: 节点 | 一续(实际) | 一续(测算) | 多续(实际) | 多续(测算) | 综合(实际) | 综合(测算)
    R2: ASP|单价 重复
    R3-6: 升舱/早鸟/其余/综合
    """
    # 标题行
    ws.cell(start_row, 1, title).font = BOLD12
    ws.cell(start_row, 1).fill = YELLOW
    ws.cell(start_row, 1).alignment = CENTER
    ws.merge_cells(start_row=start_row, end_row=start_row,
                   start_column=1, end_column=13)

    # 第一层表头
    r1 = start_row + 1
    ws.cell(r1, 1, "节点")
    headers_l1 = [(2, "一续（实际）"), (4, "一续（测算）"),
                  (6, "多续（实际）"), (8, "多续（测算）"),
                  (10, "综合（实际）"), (12, "综合（测算）")]
    for col, label in headers_l1:
        ws.cell(r1, col, label)
        ws.merge_cells(start_row=r1, end_row=r1,
                       start_column=col, end_column=col + 1)

    # 第二层表头
    r2 = start_row + 2
    for col in [2, 4, 6, 8, 10, 12]:
        ws.cell(r2, col, "ASP")
        ws.cell(r2, col + 1, "单课时价格")

    for c in range(1, 14):
        ws.cell(r1, c).font = BOLD
        ws.cell(r1, c).alignment = CENTER
        ws.cell(r1, c).fill = BLUE
        ws.cell(r1, c).border = THIN
        ws.cell(r2, c).font = BOLD
        ws.cell(r2, c).alignment = CENTER
        ws.cell(r2, c).fill = BLUE
        ws.cell(r2, c).border = THIN
    ws.merge_cells(start_row=r1, end_row=r2, start_column=1, end_column=1)

    # 数据行：升舱/早鸟/其余/综合
    nodes = ["升舱", "早鸟", "其余", "综合"]
    for i, node in enumerate(nodes):
        r = start_row + 3 + i
        ws.cell(r, 1, node)
        ws.cell(r, 1).font = BOLD
        ws.cell(r, 1).alignment = CENTER
        ws.cell(r, 1).border = THIN

        cohorts_cols = [("一续", 2, 4), ("多续", 6, 8), ("综合", 10, 12)]
        for cohort, actual_col, budget_col in cohorts_cols:
            actual = node_results.get((cohort, node), {})
            asp_a = actual.get("asp", 0) if actual else 0
            if integ_key == "含积分":
                price_a = actual.get("price_with_integ", 0) if actual else 0
            else:
                price_a = actual.get("price_no_integ", 0) if actual else 0

            ws.cell(r, actual_col, asp_a if asp_a else "-")
            ws.cell(r, actual_col + 1, price_a if price_a else "-")

            b = budget_data.get((integ_key, cohort, node), {})
            ws.cell(r, budget_col, b.get("asp", "-"))
            ws.cell(r, budget_col + 1, b.get("price", "-"))

        for c in range(2, 14):
            cell = ws.cell(r, c)
            cell.alignment = CENTER
            cell.border = THIN
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    return start_row + 7  # 占用7行


def generate_template_excel(node_results, budget_data, pkg_results,
                            cohort_totals, pkg_budget_ratio, month_label, output_dir, filters):
    """以复盘模板格式生成Excel报告"""
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / f"SKU精细对标_{month_label}.xlsx"

    wb = openpyxl.Workbook()

    # ============ Sheet 1: SKU测算&实际对比 (按模板) ============
    ws = wb.active
    ws.title = "SKU测算&实际对比"

    # 含积分块
    end_row = write_asp_block(ws, 1, "含积分单课时&ASP",
                              node_results, budget_data, "含积分")
    # 空一行
    # 不含积分块
    end_row = write_asp_block(ws, end_row + 2, "不含积分单课时&ASP",
                              node_results, budget_data, "不含积分")

    # 列宽
    ws.column_dimensions["A"].width = 12
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # ============ Sheet 2: 套餐出单明细 ============
    ws2 = wb.create_sheet("套餐出单明细")
    headers2 = ["人群", "SKU节点", "套餐名称", "订单数", "占比",
                "总GMV", "ASP", "含积分单价", "不含积分单价"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(1, col, h)
        c.font = BOLD
        c.alignment = CENTER
        c.fill = BLUE
        c.border = THIN

    for i, p in enumerate(pkg_results, 2):
        ws2.cell(i, 1, p["cohort"]).border = THIN
        ws2.cell(i, 2, p.get("sku_node", p.get("category", ""))).border = THIN
        ws2.cell(i, 3, p["package"]).border = THIN
        ws2.cell(i, 4, p["count"]).border = THIN
        ws2.cell(i, 5, p["ratio"]).border = THIN
        ws2.cell(i, 5).number_format = "0.00%"
        ws2.cell(i, 6, p["total_amt"]).border = THIN
        ws2.cell(i, 7, p["asp"]).border = THIN
        ws2.cell(i, 8, p["price_with_integ"]).border = THIN
        ws2.cell(i, 9, p["price_no_integ"]).border = THIN

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 50
    for c in "DEFGHI":
        ws2.column_dimensions[c].width = 14
    ws2.freeze_panes = "A2"

    # ============ Sheet 3: 计算口径 ============
    ws3 = wb.create_sheet("计算口径")
    filter_text = " AND ".join(
        f"{k}∈{v}" if isinstance(v, list) or k.endswith("_排除")
        else f"{k}={v}"
        for k, v in filters.items()
    )
    notes = [
        "=== 口径说明 ===",
        f"1. BI报表筛选：{filter_text}",
        "2. 用户ID 关联 正式池[池内（剔已续不可续）].学员ID",
        "3. 人群划分：当前课包顺序=1→一续, >1→多续, =0→池外（剔除），未在正式池→剔除",
        "4. SKU节点划分：取自正式池'池子节点2'字段",
        "   - 升舱 → 升舱",
        "   - 早鸟池 → 早鸟",
        "   - 其他（当月结课/次月结课/次次月结课/活跃低课时等） → 其余",
        "5. ASP = 总GMV / 订单数",
        "6. 含积分单课时价 = 总GMV / 总课时数(含积分)",
        "7. 不含积分单课时价 = 总GMV / 总课时数(不含积分)",
        "",
        f"用于计算的订单数: {sum(cohort_totals.values())}",
        f"  一续: {cohort_totals.get('一续', 0)} 单",
        f"  多续: {cohort_totals.get('多续', 0)} 单",
    ]
    for i, line in enumerate(notes, 1):
        ws3.cell(i, 1, line)
        if i == 1:
            ws3.cell(i, 1).font = BOLD14
    ws3.column_dimensions["A"].width = 80

    # ============ Sheet 4: 课包出单比例对比 ============
    ws4 = wb.create_sheet("课包出单比例对比")
    ws4.cell(1, 1, "课包出单比例（测算 vs 实际）").font = BOLD14
    ws4.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
    ws4.cell(1, 1).alignment = CENTER
    ws4.cell(1, 1).fill = YELLOW

    # 表头
    headers4 = ["人群", "SKU节点", "套餐名称", "测算占比", "实际占比", "差异"]
    for col, h in enumerate(headers4, 1):
        c = ws4.cell(2, col, h)
        c.font = BOLD
        c.alignment = CENTER
        c.fill = BLUE
        c.border = THIN

    # 数据：按人群x节点分组
    # 先按人群x节点聚合总订单数
    node_totals = {}
    for p in pkg_results:
        key = (p["cohort"], p.get("sku_node", ""))
        node_totals[key] = node_totals.get(key, 0) + p["count"]

    r = 3
    for cohort in ["一续", "多续"]:
        for node in ["升舱", "早鸟", "其余"]:
            total = node_totals.get((cohort, node), 0)
            if total == 0:
                continue
            # 该组下的所有套餐
            group_pkgs = [p for p in pkg_results
                         if p["cohort"] == cohort and p.get("sku_node", "") == node]
            if not group_pkgs:
                continue
            # 组标题行
            ws4.cell(r, 1, f"{cohort} {node}").font = BOLD
            ws4.cell(r, 1).fill = LIGHT_GREEN
            ws4.merge_cells(start_row=r, end_row=r, start_column=1, end_column=6)
            r += 1

            for p in sorted(group_pkgs, key=lambda x: -x["count"]):
                actual_ratio = p["count"] / total if total else 0
                budget_ratio = pkg_budget_ratio.get((cohort, node, p["package"]), None)
                diff = (actual_ratio - budget_ratio) if budget_ratio else None

                ws4.cell(r, 1, cohort)
                ws4.cell(r, 2, node)
                ws4.cell(r, 3, p["package"])
                ws4.cell(r, 4, budget_ratio if budget_ratio else "-")
                ws4.cell(r, 5, actual_ratio)
                ws4.cell(r, 6, diff if diff is not None else "-")

                for col in range(1, 7):
                    ws4.cell(r, col).border = THIN
                if budget_ratio:
                    ws4.cell(r, 4).number_format = "0.00%"
                ws4.cell(r, 5).number_format = "0.00%"
                if diff is not None:
                    ws4.cell(r, 6).number_format = "0.00%"
                    # 差异超过5%标红，-5%~5%绿色
                    if abs(diff) > 0.05:
                        ws4.cell(r, 6).fill = LIGHT_RED
                    else:
                        ws4.cell(r, 6).fill = LIGHT_GREEN
                r += 1

    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 50
    for c in "DEF":
        ws4.column_dimensions[c].width = 14

    wb.save(xlsx_path)
    print(f"  Excel报告: {xlsx_path}")
    return xlsx_path


def generate_html_report(node_results, budget_data, pkg_results,
                         cohort_totals, pkg_budget_ratio, month_label, output_dir, filters):
    """生成HTML分析报告（与Excel配套）"""
    output_dir.mkdir(parents=True, exist_ok=True)
    html_path = output_dir / f"SKU复盘分析_{month_label}.html"

    total = sum(cohort_totals.values())
    one = cohort_totals.get("一续", 0)
    multi = cohort_totals.get("多续", 0)
    filter_text = " &amp; ".join(f"{k} = {v}" for k, v in filters.items())
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    def diff_pct(a, b):
        if not b:
            return '<td>-</td>'
        d = (a - b) / b * 100
        cls = "pos" if abs(d) <= 3 else ("warn" if abs(d) <= 8 else "neg")
        sign = "+" if d > 0 else ""
        return f'<td class="{cls}">{sign}{d:.1f}%</td>'

    def fmt(v):
        if isinstance(v, (int, float)) and v:
            return f"{v:,.2f}"
        return "-"

    def build_block(title, integ_key):
        rows = ""
        for node in ["升舱", "早鸟", "其余", "综合"]:
            cells = [f"<td><b>{node}</b></td>"]
            for cohort in ["一续", "多续", "综合"]:
                actual = node_results.get((cohort, node), {})
                a_asp = actual.get("asp", 0)
                if integ_key == "含积分":
                    a_price = actual.get("price_with_integ", 0)
                else:
                    a_price = actual.get("price_no_integ", 0)
                b = budget_data.get((integ_key, cohort, node), {})
                b_asp = b.get("asp", 0)
                b_price = b.get("price", 0)
                cells.append(f"<td>{fmt(a_asp)}</td><td>{fmt(a_price)}</td>")
                cells.append(f"<td>{fmt(b_asp)}</td><td>{fmt(b_price)}</td>")
                cells.append(diff_pct(a_asp, b_asp))
                cells.append(diff_pct(a_price, b_price))
            rows += "<tr>" + "".join(cells) + "</tr>"
        return f"""
<div class="section">
<h2>{title}</h2>
<table>
<tr>
<th rowspan="2">节点</th>
<th colspan="6">一续</th>
<th colspan="6">多续</th>
<th colspan="6">综合</th>
</tr>
<tr>
<th>实际ASP</th><th>实际单价</th><th>测算ASP</th><th>测算单价</th><th>ASP差</th><th>单价差</th>
<th>实际ASP</th><th>实际单价</th><th>测算ASP</th><th>测算单价</th><th>ASP差</th><th>单价差</th>
<th>实际ASP</th><th>实际单价</th><th>测算ASP</th><th>测算单价</th><th>ASP差</th><th>单价差</th>
</tr>
{rows}
</table>
</div>"""

    block_with = build_block("一、含积分单课时&amp;ASP对标", "含积分")
    block_no = build_block("二、不含积分单课时&amp;ASP对标", "不含积分")

    # 套餐明细表
    one_rows = ""
    multi_rows = ""
    for p in pkg_results:
        sku_node = p.get("sku_node", p.get("category", ""))
        row_html = (f'<tr><td class="left">{p["package"]}</td>'
                    f'<td>{sku_node}</td>'
                    f'<td>{p["count"]}</td>'
                    f'<td>{p["ratio"]*100:.2f}%</td>'
                    f'<td>{p["asp"]:,.0f}</td>'
                    f'<td>{p["price_with_integ"]:.2f}</td>'
                    f'<td>{p["price_no_integ"]:.2f}</td></tr>')
        if p["cohort"] == "一续":
            one_rows += row_html
        else:
            multi_rows += row_html

    # 课包出单比例对比表 + 脱离预期分析
    node_totals = {}
    for p in pkg_results:
        key = (p["cohort"], p.get("sku_node", ""))
        node_totals[key] = node_totals.get(key, 0) + p["count"]

    ratio_rows = ""
    issues = []  # 收集脱离预期的套餐
    for cohort in ["一续", "多续"]:
        for node in ["升舱", "早鸟", "其余"]:
            total = node_totals.get((cohort, node), 0)
            if total == 0:
                continue
            group_pkgs = [p for p in pkg_results
                         if p["cohort"] == cohort and p.get("sku_node", "") == node]
            if not group_pkgs:
                continue

            ratio_rows += f'<tr class="group-header"><td colspan="6"><b>{cohort} {node}</b></td></tr>'
            for p in sorted(group_pkgs, key=lambda x: -x["count"]):
                actual = p["count"] / total if total else 0
                budget = pkg_budget_ratio.get((cohort, node, p["package"]), None)
                diff = (actual - budget) if budget else None
                diff_pct = diff * 100 if diff is not None else None

                cls = ""
                if diff is not None:
                    if abs(diff) > 0.08:  # 超8%
                        cls = "critical"
                        issues.append({
                            "cohort": cohort, "node": node, "package": p["package"],
                            "budget": budget, "actual": actual, "diff": diff
                        })
                    elif abs(diff) > 0.05:  # 5-8%
                        cls = "warn"

                ratio_rows += f'<tr class="{cls}">'
                ratio_rows += f'<td class="left">{p["package"]}</td>'
                ratio_rows += f'<td>{cohort}</td><td>{node}</td>'
                ratio_rows += f'<td>{budget*100:.1f}%</td>' if budget else '<td>-</td>'
                ratio_rows += f'<td>{actual*100:.1f}%</td>'
                if diff_pct is not None:
                    cls2 = "pos" if abs(diff) <= 0.05 else ("warn" if abs(diff) <= 0.08 else "neg")
                    sign = "+" if diff > 0 else ""
                    ratio_rows += f'<td class="{cls2}">{sign}{diff_pct:.1f}%</td>'
                else:
                    ratio_rows += '<td>-</td>'
                ratio_rows += '</tr>'

    # 脱离预期分析与建议
    analysis_html = ""
    if issues:
        issues.sort(key=lambda x: abs(x["diff"]), reverse=True)
        analysis_items = []
        for i in issues[:5]:  # 最严重的5个
            direction = "高于" if i["diff"] > 0 else "低于"
            analysis_items.append(
                f'<li><b>{i["cohort"]} {i["node"]} - {i["package"][:50]}...</b><br>'
                f'测算 {i["budget"]*100:.1f}%，实际 {i["actual"]*100:.1f}%，'
                f'<span class="{"pos" if i["diff"] > 0 else "neg"}">{direction}预期 {abs(i["diff"])*100:.1f}%</span></li>'
            )

        recommendations = []
        # 按节点汇总问题
        node_issues = {}
        for i in issues:
            k = (i["cohort"], i["node"])
            if k not in node_issues:
                node_issues[k] = []
            node_issues[k].append(i)

        for (cohort, node), group in node_issues.items():
            over = [x for x in group if x["diff"] > 0.08]
            under = [x for x in group if x["diff"] < -0.08]
            if over:
                recommendations.append(
                    f'<li><b>{cohort} {node}</b>：部分套餐出单超预期，建议分析高出单套餐的定价、权益和推广策略，复制到表现不佳套餐</li>'
                )
            if under:
                recommendations.append(
                    f'<li><b>{cohort} {node}</b>：部分套餐出单低于预期，建议检查套餐权益吸引力、定价竞争力，或优化话术与推送策略</li>'
                )

        analysis_html = f'''
<div class="section analysis">
<h2>五、脱离预期分析与改进建议</h2>
<h3>严重偏离套餐（差异&gt;8%）</h3>
<ul>{"".join(analysis_items)}</ul>
<h3>改进建议</h3>
<ul>{"".join(recommendations) if recommendations else "<li>无严重偏离情况</li>"}</ul>
</div>'''
    else:
        analysis_html = '''
<div class="section analysis">
<h2>五、脱离预期分析与改进建议</h2>
<p style="color:#10b981;font-weight:600;">✓ 所有套餐出单占比均在合理范围内（±8%以内），符合SKU测算预期。</p>
</div>'''

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>{month_label}SKU复盘分析报告</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,'Segoe UI',sans-serif;background:#f5f7fa;color:#333;padding:20px;line-height:1.6}}
.container{{max-width:1400px;margin:0 auto}}
.header{{background:linear-gradient(135deg,#667eea,#764ba2);color:white;padding:24px;border-radius:12px;margin-bottom:20px}}
.header h1{{font-size:24px;margin-bottom:6px}}
.header p{{opacity:0.9;font-size:13px}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin-bottom:20px}}
.card{{background:white;border-radius:10px;padding:14px;box-shadow:0 2px 8px rgba(0,0,0,0.08);text-align:center}}
.card .label{{font-size:11px;color:#888}}
.card .value{{font-size:22px;font-weight:700;margin-top:4px}}
.card .sub{{font-size:11px;color:#aaa}}
.section{{background:white;border-radius:10px;padding:20px;margin-bottom:18px;box-shadow:0 2px 8px rgba(0,0,0,0.08);overflow-x:auto}}
.section h2{{font-size:16px;margin-bottom:12px;padding-bottom:6px;border-bottom:2px solid #667eea}}
.section h3{{font-size:14px;margin:12px 0 8px;color:#555}}
table{{width:100%;border-collapse:collapse;font-size:11px}}
th{{background:#f0f2f5;padding:6px 4px;border:1px solid #e0e0e0;font-weight:600;white-space:nowrap}}
td{{padding:5px 4px;border:1px solid #e0e0e0;text-align:center}}
.pos{{color:#10b981;font-weight:600}}
.neg{{color:#ef4444;font-weight:600}}
.warn{{color:#f59e0b;font-weight:600}}
.left{{text-align:left}}
.filter-info{{background:#eff6ff;border-left:4px solid #3b82f6;padding:10px;border-radius:0 8px 8px 0;margin-bottom:16px;font-size:12px}}
.group-header{{background:#e2efda!important}}
.group-header td{{font-weight:600;text-align:left!important;padding:8px 6px}}
.critical{{background:#fef2f2!important}}
.analysis{{background:#fffbeb}}
.analysis ul{{margin-left:20px;margin-top:8px}}
.analysis li{{margin:8px 0;line-height:1.5}}
</style>
</head>
<body>
<div class="container">
<div class="header">
<h1>{month_label}SKU复盘分析报告</h1>
<p>分析对象：益智续费套餐 | 节点取自正式池"池子节点2" | 生成时间：{now}</p>
</div>
<div class="filter-info"><strong>筛选条件：</strong>{filter_text}</div>
<div class="cards">
<div class="card"><div class="label">订单总数</div><div class="value">{total}</div></div>
<div class="card"><div class="label">一续订单</div><div class="value">{one}</div><div class="sub">{one/total*100:.1f}%</div></div>
<div class="card"><div class="label">多续订单</div><div class="value">{multi}</div><div class="sub">{multi/total*100:.1f}%</div></div>
<div class="card"><div class="label">套餐组合</div><div class="value">{len(pkg_results)}</div></div>
</div>

{block_with}
{block_no}

<div class="section">
<h2>三、一续套餐出单明细</h2>
<table>
<tr><th class="left">套餐名称</th><th>SKU节点</th><th>订单数</th><th>占比</th><th>ASP</th><th>含积分单价</th><th>不含积分单价</th></tr>
{one_rows}
</table>
</div>

<div class="section">
<h2>四、多续套餐出单明细</h2>
<table>
<tr><th class="left">套餐名称</th><th>SKU节点</th><th>订单数</th><th>占比</th><th>ASP</th><th>含积分单价</th><th>不含积分单价</th></tr>
{multi_rows}
</table>
</div>

<div class="section">
<h2>五、课包出单比例对比（测算 vs 实际）</h2>
<table>
<tr><th class="left">套餐名称</th><th>人群</th><th>节点</th><th>测算占比</th><th>实际占比</th><th>差异</th></tr>
{ratio_rows}
</table>
</div>

{analysis_html}

</div>
</body>
</html>"""

    html_path.write_text(html, encoding="utf-8")
    print(f"  HTML报告: {html_path}")
    return html_path


def generate_csv_detail(pkg_results, month_label, output_dir):
    """导出详细CSV"""
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / f"套餐明细_{month_label}.csv"

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["人群", "SKU节点", "套餐名称", "订单数", "占比",
                         "总GMV", "ASP", "含积分单价", "不含积分单价"])
        for p in pkg_results:
            writer.writerow([
                p["cohort"], p.get("sku_node", p.get("category", "")),
                p["package"], p["count"], f"{p['ratio']*100:.2f}%",
                p["total_amt"], p["asp"],
                p["price_with_integ"], p["price_no_integ"]
            ])

    print(f"  CSV明细: {csv_path}")
    return csv_path
