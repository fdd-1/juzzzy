# -*- coding: utf-8 -*-
"""导出带匹配结果和计算过程的BI数据明细Excel
- 加入"池子节点2"和"SKU节点"列，方便核对
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict

from config import FILTERS, POOL_SHEET_NAME
from extract_data import (
    find_header_row, get_headers, normalize_id,
    classify_package, map_pool_node_to_sku
)


def export_audit_excel(bi_file, pool_file, output_file):
    """生成审计Excel：原始BI明细+匹配结果+计算过程"""

    print(f"读取BI报表: {bi_file}")
    bi_wb = openpyxl.load_workbook(bi_file, data_only=True)
    bi_ws = bi_wb.active

    bi_header_row = find_header_row(bi_ws, "用户ID")
    bi_headers = get_headers(bi_ws, bi_header_row)

    print(f"读取正式池: {pool_file}")
    pool_wb = openpyxl.load_workbook(pool_file, data_only=True)
    pool_ws = pool_wb[POOL_SHEET_NAME] if POOL_SHEET_NAME in pool_wb.sheetnames else pool_wb.active
    pool_header_row = find_header_row(pool_ws, "学员ID")
    pool_headers = get_headers(pool_ws, pool_header_row)

    pool = {}
    for row in range(pool_header_row + 1, pool_ws.max_row + 1):
        sid = pool_ws.cell(row, pool_headers["学员ID"]).value
        order = pool_ws.cell(row, pool_headers["当前课包顺序"]).value
        node2 = pool_ws.cell(row, pool_headers["池子节点2"]).value if "池子节点2" in pool_headers else None
        if sid:
            pool[normalize_id(sid)] = {
                "course_order": order, "pool_node2": node2,
            }
    pool_wb.close()

    print(f"创建审计Excel...")
    out_wb = openpyxl.Workbook()

    ws1 = out_wb.active
    ws1.title = "订单匹配明细"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    blue_fill = PatternFill(start_color="D9E2F3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", fill_type="solid")

    headers_out = [
        "用户ID", "套餐名称", "区域等级", "订单支付时业绩归属人五级部门",
        "用户实际支付金额", "课时数(不含积分)", "课时数(含积分)",
        "→ 当前课包顺序", "→ 人群", "→ 池子节点2", "→ SKU节点",
        "→ 是否纳入计算", "→ 排除原因"
    ]
    for col, h in enumerate(headers_out, 1):
        c = ws1.cell(1, col, h)
        c.font = bold
        c.alignment = center
        c.border = border
        c.fill = blue_fill if col <= 7 else yellow_fill

    matched_orders = []
    out_row = 2

    for row in range(bi_header_row + 1, bi_ws.max_row + 1):
        skip_filter = False
        for field, expected in FILTERS.items():
            if field in bi_headers:
                val = bi_ws.cell(row, bi_headers[field]).value
                if str(val).strip() != expected:
                    skip_filter = True
                    break
        if skip_filter:
            continue

        user_id = bi_ws.cell(row, bi_headers["用户ID"]).value
        pkg = bi_ws.cell(row, bi_headers["套餐名称"]).value
        region = bi_ws.cell(row, bi_headers["区域等级"]).value
        dept = bi_ws.cell(row, bi_headers["订单支付时业绩归属人五级部门"]).value
        amt = bi_ws.cell(row, bi_headers["用户实际支付金额"]).value
        hrs_no = bi_ws.cell(row, bi_headers["课时数（不含积分）"]).value
        hrs_w = bi_ws.cell(row, bi_headers["课时数（含积分）"]).value

        if not (user_id and pkg):
            continue

        uid = normalize_id(user_id)
        info = pool.get(uid)
        course_order = info["course_order"] if info else None
        pool_node2 = info["pool_node2"] if info else None
        sku_node = map_pool_node_to_sku(pool_node2) if info else "-"

        if course_order is None:
            cohort = "-"
            included = "否"
            reason = "未在正式池中"
            fill = red_fill
            sku_node = "-"
        elif course_order == 1:
            cohort = "一续"
            included = "是"
            reason = ""
            fill = green_fill
        elif course_order > 1:
            cohort = "多续"
            included = "是"
            reason = ""
            fill = green_fill
        else:
            cohort = "池外"
            included = "否"
            reason = "续池=0(池外)"
            fill = red_fill
            sku_node = "-"

        amt_v = float(amt) if amt else 0
        hrs_no_v = float(hrs_no) if hrs_no else 0
        hrs_w_v = float(hrs_w) if hrs_w else 0

        values = [
            uid, str(pkg), region, dept,
            amt_v, hrs_no_v, hrs_w_v,
            course_order if course_order is not None else "-",
            cohort,
            str(pool_node2) if pool_node2 else "-",
            sku_node,
            included, reason
        ]
        for col, v in enumerate(values, 1):
            c = ws1.cell(out_row, col, v)
            c.border = border
            if col >= 8:
                c.fill = fill

        if included == "是":
            matched_orders.append({
                "cohort": cohort, "sku_node": sku_node,
                "amount": amt_v, "hours_no_integ": hrs_no_v,
                "hours_with_integ": hrs_w_v,
            })

        out_row += 1

    widths = [12, 50, 10, 25, 14, 14, 14, 12, 8, 14, 12, 12, 18]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws1.freeze_panes = "A2"
    print(f"  订单明细行数: {out_row - 2}")

    bi_wb.close()

    # ============ Sheet 2: 计算过程汇总 ============
    ws2 = out_wb.create_sheet("计算过程")

    ws2.cell(1, 1, "口径说明").font = Font(bold=True, size=14)
    ws2.cell(2, 1, "1. 筛选条件: 订单支付时业绩归属人五级部门=港澳益智教学服务区 AND 区域等级=港澳")
    ws2.cell(3, 1, "2. 用户ID 关联 正式池[池内（剔已续不可续）].学员ID")
    ws2.cell(4, 1, "3. 当前课包顺序=1→一续, >1→多续, =0→池外（剔除）, 未匹配→剔除")
    ws2.cell(5, 1, "4. SKU节点 = 池子节点2 映射: 升舱→升舱, 早鸟池→早鸟, 其他→其余")
    ws2.cell(6, 1, "5. ASP = 总GMV / 订单数")
    ws2.cell(7, 1, "6. 含积分单课时价 = 总GMV / 总课时数(含积分)")
    ws2.cell(8, 1, "7. 不含积分单课时价 = 总GMV / 总课时数(不含积分)")

    # 按【人群×SKU节点】聚合
    agg = defaultdict(lambda: {
        "count": 0, "total_amt": 0, "total_hrs_no": 0, "total_hrs_w": 0
    })
    for o in matched_orders:
        key = (o["cohort"], o["sku_node"])
        agg[key]["count"] += 1
        agg[key]["total_amt"] += o["amount"]
        agg[key]["total_hrs_no"] += o["hours_no_integ"]
        agg[key]["total_hrs_w"] += o["hours_with_integ"]

    start_row = 10
    ws2.cell(start_row, 1, "按【人群×SKU节点】聚合").font = Font(bold=True, size=12)

    cat_headers = ["人群", "SKU节点", "订单数", "总GMV", "总课时(不含积分)",
                   "总课时(含积分)", "ASP=GMV/订单数", "含积分单价",
                   "不含积分单价"]
    for col, h in enumerate(cat_headers, 1):
        c = ws2.cell(start_row + 1, col, h)
        c.font = bold
        c.alignment = center
        c.border = border
        c.fill = blue_fill

    r = start_row + 2
    total_count = total_amt = total_hrs_no = total_hrs_w = 0
    for (cohort, node), d in sorted(agg.items()):
        asp = d["total_amt"] / d["count"] if d["count"] else 0
        pw = d["total_amt"] / d["total_hrs_w"] if d["total_hrs_w"] else 0
        pn = d["total_amt"] / d["total_hrs_no"] if d["total_hrs_no"] else 0
        vals = [cohort, node, d["count"],
                round(d["total_amt"], 2),
                round(d["total_hrs_no"], 2),
                round(d["total_hrs_w"], 2),
                round(asp, 2), round(pw, 2), round(pn, 2)]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(r, col, v)
            c.border = border
        total_count += d["count"]
        total_amt += d["total_amt"]
        total_hrs_no += d["total_hrs_no"]
        total_hrs_w += d["total_hrs_w"]
        r += 1

    asp_total = total_amt / total_count if total_count else 0
    pw_total = total_amt / total_hrs_w if total_hrs_w else 0
    pn_total = total_amt / total_hrs_no if total_hrs_no else 0
    totals = ["合计", "-", total_count, round(total_amt, 2),
              round(total_hrs_no, 2), round(total_hrs_w, 2),
              round(asp_total, 2), round(pw_total, 2), round(pn_total, 2)]
    for col, v in enumerate(totals, 1):
        c = ws2.cell(r, col, v)
        c.font = bold
        c.fill = yellow_fill
        c.border = border

    for i, w in enumerate([8, 12, 8, 14, 16, 16, 16, 16, 16], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    out_wb.save(output_file)
    print(f"\n✓ 审计Excel已生成: {output_file}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--month", default="4月")
    args = parser.parse_args()

    skill_dir = Path(__file__).parent
    data_dir = skill_dir / "data" / args.month
    output_dir = skill_dir / "output" / args.month
    output_dir.mkdir(parents=True, exist_ok=True)

    bi_files = [f for f in data_dir.glob("*主订单宽表*") if not f.name.startswith("~$")]
    pool_files = [f for f in data_dir.glob("*正式池*") if not f.name.startswith("~$")]
    if not bi_files or not pool_files:
        print("找不到BI报表或正式池文件")
        sys.exit(1)

    out = output_dir / f"审计明细_{args.month}.xlsx"
    export_audit_excel(bi_files[0], pool_files[0], out)

