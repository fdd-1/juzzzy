# -*- coding: utf-8 -*-
"""
执行前校验：Excel 列头、课包类型映射、必填字段、数值字段。
执行后验证：在 CRM 列表搜索课时包/套餐名称，确认创建成功。

返回 (ok: bool, errors: list[str], warnings: list[str])
"""

from pathlib import Path
import openpyxl


REQUIRED_LESSON_HEADERS = [
    "课时包名称", "课包类型", "课包分类",
    "有效期", "补课次数", "普通课时", "赠送课时",
    "原价", "优惠价", "试学期", "打卡次数", "停课次数",
    "适用课类",
]

INTEGER_FIELDS = [
    "有效期", "补课次数", "普通课时", "赠送课时",
    "原价", "优惠价", "试学期", "打卡次数", "停课次数",
]


def _iter_data_rows(ws):
    for raw in ws.iter_rows(values_only=True):
        if not raw or raw[0] is None:
            continue
        if str(raw[0]).strip() != "课时包名称":
            continue
        yield raw


def _row_to_kv(raw):
    kv = {}
    for i in range(0, len(raw), 2):
        label = raw[i]
        value = raw[i + 1] if i + 1 < len(raw) else None
        if label is None:
            break
        kv[str(label).strip()] = value
    return kv


def precheck_lesson_excel(excel_path, type_mapping, type_parent):
    """课时包脚本的执行前校验。

    type_mapping: {Excel 名称: CRM 名称} 已覆盖映射表
    type_parent : {CRM 名称: 父级} 父级映射表（确认叶子节点能落到合法父级）
    """
    errors = []
    warnings = []

    p = Path(excel_path)
    if not p.exists():
        return False, [f"Excel 不存在: {excel_path}"], []
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        warnings.append(f"文件后缀非 .xlsx：{p.suffix}")

    try:
        wb = openpyxl.load_workbook(p, data_only=True)
    except Exception as e:
        return False, [f"Excel 打不开: {e}"], []
    ws = wb.active

    rows = list(_iter_data_rows(ws))
    if not rows:
        return False, ["未找到行首为「课时包名称」的数据行"], warnings

    seen_types = set()
    seen_names = set()

    for idx, raw in enumerate(rows, 1):
        kv = _row_to_kv(raw)

        # 列头覆盖：每行都要含全部必填 label
        missing = [h for h in REQUIRED_LESSON_HEADERS if h not in kv]
        if missing:
            errors.append(f"第 {idx} 行缺列：{', '.join(missing)}")

        # 名称唯一性
        name = (kv.get("课时包名称") or "").strip() if kv.get("课时包名称") else ""
        if not name:
            errors.append(f"第 {idx} 行「课时包名称」为空")
        elif name in seen_names:
            errors.append(f"第 {idx} 行课时包名称重复：{name}")
        else:
            seen_names.add(name)

        # 数值字段
        for f in INTEGER_FIELDS:
            v = kv.get(f)
            if v is None or v == "":
                errors.append(f"第 {idx} 行「{f}」为空")
                continue
            try:
                int(v)
            except (TypeError, ValueError):
                errors.append(f"第 {idx} 行「{f}」非整数：{v!r}")

        # 课包类型收集
        t = kv.get("课包类型")
        if t is not None:
            seen_types.add(str(t).strip())

    # 课包类型映射覆盖
    valid_leaves = set(type_parent.keys())
    for t in sorted(seen_types):
        if not t:
            errors.append("出现空的「课包类型」值")
            continue
        mapped = type_mapping.get(t, t)
        if mapped not in valid_leaves:
            errors.append(
                f"课包类型未覆盖：Excel「{t}」→ 映射为「{mapped}」，"
                f"但不在 TYPE_PARENT（{', '.join(sorted(valid_leaves))}）中。"
                f"请在脚本里补 mapping 或 TYPE_PARENT。"
            )

    return (len(errors) == 0), errors, warnings


def precheck_package_excel(excel_path):
    """套餐脚本的执行前校验：只需课时包名称。"""
    errors = []
    warnings = []

    p = Path(excel_path)
    if not p.exists():
        return False, [f"Excel 不存在: {excel_path}"], []

    try:
        wb = openpyxl.load_workbook(p, data_only=True)
    except Exception as e:
        return False, [f"Excel 打不开: {e}"], []
    ws = wb.active

    rows = list(_iter_data_rows(ws))
    if not rows:
        return False, ["未找到行首为「课时包名称」的数据行"], warnings

    seen = set()
    for idx, raw in enumerate(rows, 1):
        kv = _row_to_kv(raw)
        name = (kv.get("课时包名称") or "").strip() if kv.get("课时包名称") else ""
        if not name:
            errors.append(f"第 {idx} 行「课时包名称」为空")
            continue
        if name in seen:
            errors.append(f"第 {idx} 行课时包名称重复：{name}")
        seen.add(name)

    return (len(errors) == 0), errors, warnings


def print_report(title, ok, errors, warnings):
    print(f"\n=== {title} ===")
    if warnings:
        for w in warnings:
            print(f"  [WARN] {w}")
    if errors:
        for e in errors:
            print(f"  [ERR ] {e}")
    print(f"  结果: {'通过' if ok else '未通过'}\n")


def verify_in_list(page, name, search_input_selector, query_button_text="查询",
                   table_selector=".el-table tr", wait_ms=1000):
    """在 CRM 列表页搜索名称，确认创建成功。

    返回 (found: bool, detail: str)
    """
    try:
        page.wait_for_timeout(800)
        search = page.locator(search_input_selector).first
        try:
            search.fill("")
        except Exception:
            pass
        search.fill(name)
        page.wait_for_timeout(300)

        # 点查询
        clicked = False
        for btn in page.locator(f"button:has-text('{query_button_text}')").all():
            try:
                if btn.is_visible():
                    btn.click()
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked:
            return False, f"未找到「{query_button_text}」按钮"

        page.wait_for_timeout(wait_ms)

        # 表格中是否有该名称
        try:
            row = page.locator(table_selector).filter(has_text=name).first
            if row.is_visible():
                return True, "列表已找到"
        except Exception:
            pass
        return False, "列表未找到"
    except Exception as e:
        return False, f"验证异常: {e}"
