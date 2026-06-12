#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""二次筛选导出的报表 + 在 BI 报表 xlsx 上追加 3 个 sheet（P0）

筛选条件：
  - 月初剩余总课时 in [1, 2, 3, ..., 12]
  - 是否可续学员 == 1
  - 月初是否续费 == 空（NaN）
  - 学员状态 == 执行中（剔除 结课/退费/停课/等班 等所有非执行中状态）

团队过滤（P0 规则）：
  - 续费归属老师6级部门含「台湾」→ 直接剔除（不进任何输出 / sheet）
  - 外教不再单独剔除：通过 sheet3/sheet4 按 3 级部门海外/国内自动可视化

BI 报表追加 sheet（在原 xlsx 上 in-place 加 3 个 sheet）：
  - sheet2「{X}月教学协作-1-12课时低活明细」：剔除台湾后的全部明细
  - sheet3「海外主讲团队」：sheet2 中续费归属老师3级部门含「海外」的明细
  - sheet4「国内主讲团队」：sheet2 中续费归属老师3级部门不含「海外」的明细

输出：
  - 原 BI 报表 xlsx 末尾追加上述 3 个 sheet
  - output/p0/dadou_ids_{YYYYMMDD}.xlsx → 普通标签/用户群（列：豌豆大账号ID）
  - output/p0/user_ids_{YYYYMMDD}.xlsx  → 益智标签/用户群（列：学员ID）
"""
import sys, io, argparse, datetime as dt
from pathlib import Path
import pandas as pd

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output" / "p0"

def log(m): print(m, flush=True)


def append_detail_sheets(input_path, filtered_df, sheet2_name):
    """在 BI 报表 xlsx 上 in-place 追加 3 个 sheet：
       sheet2 = filtered_df 全部
       sheet3「海外主讲团队」 = filtered_df 中 续费归属老师3级部门 含「海外」
       sheet4「国内主讲团队」 = filtered_df 中 续费归属老师3级部门 不含「海外」
       同名 sheet 已存在则先删后建（重跑幂等）。
    """
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    if "续费归属老师3级部门" not in filtered_df.columns:
        log("[WARN] 缺少 续费归属老师3级部门 列，跳过追加 sheet3/sheet4")
        return

    dept3 = filtered_df["续费归属老师3级部门"].astype(str)
    haiwai_mask = dept3.str.contains("海外", na=False)
    haiwai_df = filtered_df[haiwai_mask].copy()
    guonei_df = filtered_df[~haiwai_mask].copy()

    targets = [
        (sheet2_name, filtered_df),
        ("海外主讲团队", haiwai_df),
        ("国内主讲团队", guonei_df),
    ]

    log(f"[STEP 4] 追加 sheet 到 BI 报表：{input_path.name}")
    wb = load_workbook(input_path)
    for name, sub_df in targets:
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        for r in dataframe_to_rows(sub_df, index=False, header=True):
            ws.append(r)
        log(f"  -> sheet「{name}」: {len(sub_df)} 行")
    wb.save(input_path)
    log(f"[OK] BI 报表已更新（含 sheet1 原始数据 + 3 个新 sheet）")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="导出的报表文件路径（xlsx）")
    ap.add_argument("--month", help="月份 YYYY-MM，用于命名输出文件（可选）")
    ap.add_argument("--course-range", default="1-12", help="月初剩余总课时筛选区间（含端点），格式 'min-max'，默认 1-12")
    args = ap.parse_args()

    # 解析课时区间
    try:
        ks_min, ks_max = map(int, args.course_range.split("-"))
        if ks_min < 0 or ks_max < ks_min:
            raise ValueError
    except Exception:
        log(f"[ERROR] --course-range 格式错误，需 'min-max'：{args.course_range}")
        sys.exit(1)
    course_range_label = f"{ks_min}-{ks_max}"

    input_path = Path(args.input)
    if not input_path.exists():
        log(f"[ERROR] 找不到输入文件：{input_path}")
        sys.exit(1)

    log(f"[STEP 1] 读取 {input_path}")
    df = pd.read_excel(input_path)
    log(f"  -> 原始数据 {len(df)} 行")
    log(f"  -> 列名：{list(df.columns)}")

    # 确认列名存在
    required_cols = ["月初剩余总课时", "是否可续学员", "月初是否续费", "学员状态", "学员ID", "大账号ID", "续费归属老师", "续费归属老师6级部门", "续费归属老师3级部门"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        log(f"[ERROR] 缺少必需列：{missing}")
        sys.exit(2)

    # 筛选
    log(f"[STEP 2] 筛选条件：月初剩余总课时 {course_range_label} / 是否可续学员=1 / 月初是否续费=空 / 学员状态=执行中 / 续费归属老师非空")
    teacher_col = df["续费归属老师"].astype(str).str.strip()
    teacher_filled = df["续费归属老师"].notna() & (teacher_col != "") & (teacher_col.str.lower() != "nan")
    status_col = df["学员状态"].astype(str).str.strip()
    status_active = (status_col == "执行中")
    mask = (
        (df["月初剩余总课时"] >= ks_min) & (df["月初剩余总课时"] <= ks_max) &
        (df["是否可续学员"] == 1) &
        df["月初是否续费"].isna() &
        status_active &
        teacher_filled
    )
    filtered = df[mask].copy()
    excluded_status = df[
        (df["月初剩余总课时"] >= ks_min) & (df["月初剩余总课时"] <= ks_max) &
        (df["是否可续学员"] == 1) & df["月初是否续费"].isna() & teacher_filled & ~status_active
    ]
    if len(excluded_status) > 0:
        excluded_status_samples = sorted(set(excluded_status["学员状态"].astype(str).tolist()))
        log(f"  -> 学员状态非执行中剔除 {len(excluded_status)} 行，涉及状态：{excluded_status_samples}")
    log(f"  -> 筛选后 {len(filtered)} 行（其中续费归属老师为空被剔除）")

    # 团队过滤（P0 规则）：仅剔除台湾团队
    log("[STEP 2.1] 团队过滤：仅剔除续费归属老师6级部门含「台湾」")
    dept_col = filtered["续费归属老师6级部门"].astype(str)
    taiwan_mask = dept_col.str.contains("台湾", na=False)
    taiwan_count = int(taiwan_mask.sum())
    if taiwan_count > 0:
        taiwan_samples = sorted(set(dept_col[taiwan_mask].tolist()))
        log(f"  -> 剔除台湾团队 {taiwan_count} 行，涉及部门：{taiwan_samples}")
    else:
        log("  -> 无台湾团队明细")
    filtered = filtered[~taiwan_mask].copy()
    log(f"  -> 剔除后剩余 {len(filtered)} 行")

    if len(filtered) == 0:
        log("[WARN] 主输出为空，不生成 dadou/user 文件")

    # 拆分
    log("[STEP 3] 拆分主输出：豌豆大账号ID / 学员ID")
    # 普通群：大账号ID，重命名列为"用户id"以匹配模板
    dadou_df = filtered[["大账号ID"]].drop_duplicates().dropna()
    dadou_df.columns = ["用户id"]  # 保持模板原始列名
    # 益智群：学员ID，重命名列为"用户id"以匹配模板
    user_df = filtered[["学员ID"]].drop_duplicates().dropna()
    user_df.columns = ["用户id"]  # 保持模板原始列名

    log(f"  -> 豌豆大账号ID（普通群）：{len(dadou_df)} 条")
    log(f"  -> 学员ID（益智群）：{len(user_df)} 条")

    # 落盘
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    today_tag = dt.date.today().strftime("%Y%m%d")
    dadou_path = OUTPUT_DIR / f"dadou_ids_{today_tag}.xlsx"
    user_path = OUTPUT_DIR / f"user_ids_{today_tag}.xlsx"

    dadou_df.to_excel(dadou_path, index=False, engine="openpyxl")
    user_df.to_excel(user_path, index=False, engine="openpyxl")

    log(f"[OK] 主输出已生成:")
    log(f"  - {dadou_path}")
    log(f"  - {user_path}")

    # 在 BI 报表上追加 3 个 sheet（明细 + 海外/国内主讲团队）
    if args.month:
        try:
            _, m = map(int, args.month.split("-"))
            sheet2_name = f"{m}月教学协作-{course_range_label}课时低活明细"
        except Exception:
            sheet2_name = f"{course_range_label}课时低活明细"
    else:
        sheet2_name = f"{course_range_label}课时低活明细"
    append_detail_sheets(input_path, filtered, sheet2_name)

    # 写 latest_inputs.json 供后续脚本用（含课时区间，create_tag/create_group 命名需要）
    manifest_path = SCRIPT_DIR / "liuyi_tag" / "latest_inputs.json"
    manifest_path.parent.mkdir(exist_ok=True)
    manifest = {
        "dadou_ids_xlsx": str(dadou_path),
        "user_ids_xlsx": str(user_path),
        "course_range": course_range_label,
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"[OK] 已写入 {manifest_path}")


if __name__ == "__main__":
    import json
    main()
