#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""P1 服务池协作 - 二次筛选导出的报表，拆分为两份 xlsx + BI 报表追加 3 个 sheet

筛选条件（与P0不同，不限制课时）：
  - 是否可续学员 == 1
  - 月初是否续费 == 空（NaN）
  - 学员状态 == 执行中（剔除 结课/退费/停课/等班 等所有非执行中状态）

团队过滤（P1）：续费归属老师6级部门含「台湾」或「外教」→ 剔除

BI 报表追加 sheet（在原 xlsx 上 in-place 加 3 个 sheet）：
  - sheet2「{X}月教学协作-{X+1}月服务池明细」：剔除后剩余的全部明细（完整列）
  - sheet3「海外主讲团队」：sheet2 中续费归属老师3级部门含「海外」的明细
  - sheet4「国内主讲团队」：sheet2 中续费归属老师3级部门不含「海外」的明细

输出：
  - 原 BI 报表 xlsx 末尾追加上述 3 个 sheet
  - output/p1/p1_dadou_ids_{YYYYMMDD}.xlsx → 普通标签/用户群（列：豌豆大账号ID）
  - output/p1/p1_user_ids_{YYYYMMDD}.xlsx  → 益智标签/用户群（列：学员ID）
"""
import sys, io, argparse, datetime as dt
from pathlib import Path
import pandas as pd

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output" / "p1"

def log(m): print(m, flush=True)


def append_detail_sheets(input_path, filtered_df, sheet2_name):
    """在 BI 报表 xlsx 上 in-place 追加 3 个 sheet：
       sheet2 = filtered_df 全部
       sheet3 「海外主讲团队」 = filtered_df 中 续费归属老师3级部门 含「海外」
       sheet4 「国内主讲团队」 = filtered_df 中 续费归属老师3级部门 不含「海外」
       同名 sheet 已存在则先删后建（重跑幂等）。
    """
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    if "续费归属老师3级部门" not in filtered_df.columns:
        log("[WARN] 缺少 续费归属老师3级部门 列，跳过追加 sheet3/sheet4")
        return

    dept3 = filtered_df["续费归属老师3级部门"].astype(str)
    haiwai_mask = dept3.str.contains("海外", na=False)
    haiwai_df = filtered_df[haiwai_mask].copy()
    guonei_df = filtered_df[~haiwai_mask].copy()

    targets = [
        (sheet2_name, filtered_df),
        ("海外主讲团队", haiwai_df),
        ("国内主讲团队", guonei_df),
    ]

    log(f"[STEP 4] 追加 sheet 到 BI 报表：{input_path.name}")
    wb = load_workbook(input_path)
    for name, sub_df in targets:
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        for r in dataframe_to_rows(sub_df, index=False, header=True):
            ws.append(r)
        log(f"  -> sheet「{name}」: {len(sub_df)} 行")
    wb.save(input_path)
    log(f"[OK] BI 报表已更新（含 sheet1 原始数据 + 3 个新 sheet）")


def derive_user_month(internal_month_str):
    """从内部月份（X+1）推回用户输入的月份（X），用于 sheet2 命名。"""
    try:
        y, m = map(int, internal_month_str.split("-"))
    except Exception:
        return None, None
    if m == 1:
        return y - 1, 12
    return y, m - 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="导出的报表文件路径（xlsx）")
    ap.add_argument("--month", help="月份 YYYY-MM，用于命名输出文件（可选）")
    args = ap.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        log(f"[ERROR] 找不到输入文件：{input_path}")
        sys.exit(1)

    log(f"[STEP 1] 读取 {input_path}")
    # 自动定位表头：CLI 导出 header=13，browser 导出 header=6，按 "学员ID" 标记定位
    raw = pd.read_excel(input_path, header=None, nrows=20)
    header_row = None
    for i in range(len(raw)):
        row_vals = [str(v) for v in raw.iloc[i].tolist()]
        if "学员ID" in row_vals and "大账号ID" in row_vals:
            header_row = i
            break
    if header_row is None:
        log("[ERROR] 无法定位表头行（找不到 '学员ID' + '大账号ID'）")
        sys.exit(1)
    log(f"  -> 表头位于第 {header_row + 1} 行（0-indexed: {header_row}）")
    df = pd.read_excel(input_path, header=header_row)
    log(f"  -> 原始数据 {len(df)} 行")
    log(f"  -> 列名：{list(df.columns)}")

    # 确认列名存在
    required_cols = ["是否可续学员", "月初是否续费", "学员状态", "学员ID", "大账号ID", "续费归属老师", "续费归属老师6级部门", "续费归属老师3级部门"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        log(f"[ERROR] 缺少必需列：{missing}")
        sys.exit(2)

    # 筛选（P1：不限制课时）
    log("[STEP 2] 筛选条件：是否可续学员=1 / 月初是否续费=空 / 学员状态=执行中 / 续费归属老师非空（P1不限制课时）")
    teacher_col = df["续费归属老师"].astype(str).str.strip()
    teacher_filled = df["续费归属老师"].notna() & (teacher_col != "") & (teacher_col.str.lower() != "nan")
    status_col = df["学员状态"].astype(str).str.strip()
    status_active = (status_col == "执行中")
    mask = (
        (df["是否可续学员"] == 1) &
        df["月初是否续费"].isna() &
        status_active &
        teacher_filled
    )
    filtered = df[mask].copy()
    excluded_status = df[(df["是否可续学员"] == 1) & df["月初是否续费"].isna() & teacher_filled & ~status_active]
    if len(excluded_status) > 0:
        excluded_status_samples = sorted(set(excluded_status["学员状态"].astype(str).tolist()))
        log(f"  -> 学员状态非执行中剔除 {len(excluded_status)} 行，涉及状态：{excluded_status_samples}")
    log(f"  -> 筛选后 {len(filtered)} 行")

    # 二次剔除：续费归属老师6级部门含「台湾」或「外教」的明细
    log("[STEP 2.1] 剔除续费归属老师6级部门含「台湾」或「外教」的明细")
    dept_col = filtered["续费归属老师6级部门"].astype(str)
    exclude_mask = dept_col.str.contains("台湾", na=False) | dept_col.str.contains("外教", na=False)
    excluded_count = int(exclude_mask.sum())
    if excluded_count > 0:
        excluded_samples = sorted(set(dept_col[exclude_mask].tolist()))
        log(f"  -> 剔除 {excluded_count} 行，涉及部门：{excluded_samples}")
    else:
        log("  -> 无需剔除")
    filtered = filtered[~exclude_mask].copy()
    log(f"  -> 剔除后剩余 {len(filtered)} 行")

    if len(filtered) == 0:
        log("[WARN] 筛选结果为空，不生成输出文件")
        sys.exit(0)

    # 拆分
    log("[STEP 3] 拆分为两份：豌豆大账号ID / 学员ID")
    # 普通群：大账号ID，重命名列为"用户id"以匹配模板
    dadou_df = filtered[["大账号ID"]].drop_duplicates().dropna()
    dadou_df.columns = ["用户id"]  # 保持模板原始列名
    # 益智群：学员ID，重命名列为"用户id"以匹配模板
    user_df = filtered[["学员ID"]].drop_duplicates().dropna()
    user_df.columns = ["用户id"]  # 保持模板原始列名

    log(f"  -> 豌豆大账号ID（普通群）：{len(dadou_df)} 条")
    log(f"  -> 学员ID（益智群）：{len(user_df)} 条")

    # 落盘
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    today_tag = dt.date.today().strftime("%Y%m%d")
    dadou_path = OUTPUT_DIR / f"p1_dadou_ids_{today_tag}.xlsx"
    user_path = OUTPUT_DIR / f"p1_user_ids_{today_tag}.xlsx"

    dadou_df.to_excel(dadou_path, index=False, engine="openpyxl")
    user_df.to_excel(user_path, index=False, engine="openpyxl")

    log(f"[OK] 已生成:")
    log(f"  - {dadou_path}")
    log(f"  - {user_path}")

    # 在 BI 报表上追加 3 个 sheet（明细 + 海外/国内主讲团队）
    if args.month:
        uy, um = derive_user_month(args.month)
        iy, im = map(int, args.month.split("-"))
        if uy is not None:
            sheet2_name = f"{um}月教学协作-{im}月服务池明细"
        else:
            sheet2_name = "服务池明细"
    else:
        sheet2_name = "服务池明细"
    append_detail_sheets(input_path, filtered, sheet2_name)

    # 写 latest_inputs_p1.json 供后续脚本用
    manifest_path = SCRIPT_DIR / "liuyi_tag" / "latest_inputs_p1.json"
    manifest_path.parent.mkdir(exist_ok=True)
    manifest = {
        "dadou_ids_xlsx": str(dadou_path),
        "user_ids_xlsx": str(user_path),
    }
    import json
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"[OK] 已写入 {manifest_path}")


if __name__ == "__main__":
    main()
