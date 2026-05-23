#!/usr/bin/env python3
"""学情总结抽查分析 - 按老师维度抽样、模板匹配、同质化评估.

Usage:
    python3 xueqing_review.py --input <input.xlsx> --output <output.xlsx> [--sample 20] [--seed 42]
"""

import argparse
import pandas as pd
import re
import random
from difflib import SequenceMatcher
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# 默认配置
# ══════════════════════════════════════════════════════════════════
DEFAULT_INPUT = Path(r"C:/Users/fengjianyi/Desktop/教学协作/学情抽查/教学&学管协作跟进明细.xlsx")
DEFAULT_OUTPUT = Path(r"C:/Users/fengjianyi/Desktop/教学协作/学情抽查/学情总结抽查结果.xlsx")
DEFAULT_SAMPLE = 20
DEFAULT_SEED = 42
HEADER_ROW = 9  # 数据表头所在行（第10行，索引9）
XUEQING_COL = "学情总结"
TEACHER_COL = "主讲姓名"
ART_NAME_COL = "上课老师艺名"
STUDENT_ID_COL = "学员id"
XUEQING_TIME_COL = "学情总结创建时间"

# 模板关键词（三个阶段通用关键词）
TEMPLATE_KEYWORDS = [
    "记忆犹新", "较好知识点", "需要提升知识点", "提升知识点",
    "家长痛点", "未续费顾虑点", "顾虑点"
]

# 各阶段特有关键词
STAGE_KEYWORDS = {
    "s1-3": ["专注力", "学习兴趣", "课堂活跃度"],
    "s4-6": ["学习习惯", "笔记草稿", "三率"],
    "s7-9": ["校内成绩", "校内学习成绩", "学校考试"],
}

# 不符合模板的短文本特征（课后反馈、状态备注等）
INVALID_PATTERNS = [
    r'^.{0,15}$',  # 太短的文本（15字以内）
    r'^(已转走|已续费|未接|停课|退费|请假|已流失|不续费|已满|转介绍|已加微|待跟进|待沟通)',
    r'^(rhya|.+课上很认真|.+很认真|孩子.+专注|宝贝.+认真)',
]


# ══════════════════════════════════════════════════════════════════
# 模板匹配
# ══════════════════════════════════════════════════════════════════
def is_invalid_entry(text: str) -> bool:
    """判断是否为无效学情总结（非模板格式）."""
    text = str(text).strip()
    if not text:
        return True
    for pattern in INVALID_PATTERNS:
        if re.match(pattern, text, re.IGNORECASE):
            return True
    return False


def count_numbered_items(text: str) -> int:
    """统计文本中编号条目数量 (1. 2. 3. 等)."""
    matches = re.findall(r'(?:^|\n)\s*\d+[\.、．]\s*', text)
    return len(matches)


def match_template(text: str) -> dict:
    """评估学情总结是否匹配模板.

    返回:
        {
            "matched": bool,  # 是否符合模板
            "stage": str,     # 匹配的阶段
            "score": float,   # 匹配分数 0-1
            "reason": str,    # 评估说明
        }
    """
    text = str(text).strip()

    # 1. 先检查是否无效
    if is_invalid_entry(text):
        return {
            "matched": False,
            "stage": "无效",
            "score": 0.0,
            "reason": "不符合模板要求，疑似课后反馈或状态备注"
        }

    # 2. 检查是否有编号结构
    numbered_items = count_numbered_items(text)

    # 3. 检查关键词匹配
    matched_keywords = []
    for kw in TEMPLATE_KEYWORDS:
        if kw in text:
            matched_keywords.append(kw)

    # 4. 检查阶段匹配
    stage_matched = None
    for stage, keywords in STAGE_KEYWORDS.items():
        for kw in keywords:
            if kw in text:
                stage_matched = stage
                break
        if stage_matched:
            break

    # 5. 综合判断
    # 有编号结构 + 至少3个关键词 = 符合模板
    # 有编号结构 + 至少2个关键词 + 有阶段匹配 = 符合模板
    keyword_count = len(matched_keywords)

    if numbered_items >= 5 and keyword_count >= 3:
        stage = stage_matched or "未明确阶段"
        score = min(1.0, (numbered_items / 7) * 0.5 + (keyword_count / 5) * 0.5)
        return {
            "matched": True,
            "stage": stage,
            "score": round(score, 2),
            "reason": f"符合模板结构（{numbered_items}个编号条目，{keyword_count}个关键词命中：{', '.join(matched_keywords[:4])}）"
        }
    elif numbered_items >= 3 and keyword_count >= 2 and stage_matched:
        score = min(1.0, (numbered_items / 7) * 0.4 + (keyword_count / 5) * 0.6)
        return {
            "matched": True,
            "stage": stage_matched,
            "score": round(score, 2),
            "reason": f"基本符合模板（{numbered_items}个编号条目，阶段{stage_matched}，{keyword_count}个关键词）"
        }
    elif numbered_items >= 5 and keyword_count >= 1:
        return {
            "matched": True,
            "stage": stage_matched or "未明确阶段",
            "score": round(0.5, 2),
            "reason": f"有编号结构但关键词较少（{numbered_items}个编号条目，{keyword_count}个关键词），模板匹配度一般"
        }
    else:
        return {
            "matched": False,
            "stage": "不符合",
            "score": 0.0,
            "reason": f"不符合模板要求（{numbered_items}个编号条目，{keyword_count}个关键词），内容缺少标准结构"
        }


# ══════════════════════════════════════════════════════════════════
# 同质化评估
# ══════════════════════════════════════════════════════════════════
def text_similarity(t1: str, t2: str) -> float:
    """计算两段文本的相似度."""
    return SequenceMatcher(None, t1, t2).ratio()


def evaluate_homogeneity(texts: list) -> dict:
    """评估同一老师多条学情总结的同质化程度.

    同质化低 = 优秀（每条都有个性化内容）
    同质化高 = 及格（多条内容雷同）
    """
    if len(texts) < 2:
        return {
            "level": "仅1条",
            "avg_similarity": 0.0,
            "max_similarity": 0.0,
            "min_similarity": 0.0,
            "reason": "仅有1条学情总结，无法评估同质化"
        }

    # 计算所有配对的相似度
    similarities = []
    for i in range(len(texts)):
        for j in range(i + 1, len(texts)):
            sim = text_similarity(texts[i], texts[j])
            similarities.append(sim)

    avg_sim = sum(similarities) / len(similarities)
    max_sim = max(similarities)
    min_sim = min(similarities)

    if avg_sim < 0.35:
        level = "优秀"
        reason = f"同质化程度低，内容个性化程度高（平均相似度{avg_sim:.0%}）"
    elif avg_sim < 0.55:
        level = "及格"
        reason = f"同质化程度一般，部分内容有重复（平均相似度{avg_sim:.0%}）"
    else:
        level = "及格（偏高）"
        reason = f"同质化程度高，多条内容较雷同（平均相似度{avg_sim:.0%}）"

    return {
        "level": level,
        "avg_similarity": round(avg_sim, 3),
        "max_similarity": round(max_sim, 3),
        "min_similarity": round(min_sim, 3),
        "reason": reason
    }


# ══════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="学情总结抽查分析")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="输入Excel路径")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="输出Excel路径")
    parser.add_argument("--sample", type=int, default=DEFAULT_SAMPLE, help="抽样老师数量")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED, help="随机种子")
    args = parser.parse_args()

    input_file = Path(args.input)
    output_file = Path(args.output)
    sample_size = args.sample
    random.seed(args.seed)

    print("=" * 60)
    print("学情总结抽查分析")
    print("=" * 60)

    # 1. 读取数据
    print("\n[1] 读取数据...")
    df = pd.read_excel(input_file, header=HEADER_ROW)
    print(f"  总行数: {len(df)}")

    # 筛选有效学情总结（非空且非无效）
    valid_mask = df[XUEQING_COL].notna() & (df[XUEQING_COL].astype(str).str.strip() != '')
    valid_df = df[valid_mask].copy()
    print(f"  有学情总结的记录: {len(valid_df)}")

    # 2. 按老师分组
    print("\n[2] 按老师维度分组...")
    teacher_groups = valid_df.groupby(TEACHER_COL, dropna=False)
    teacher_stats = []
    for teacher, group in teacher_groups:
        if pd.isna(teacher) or str(teacher).strip() == '':
            continue
        teacher_stats.append({
            "teacher": teacher,
            "art_name": group[ART_NAME_COL].dropna().unique()[0] if len(group[ART_NAME_COL].dropna().unique()) > 0 else "",
            "total_entries": len(group),
            "valid_entries": len(group[group[XUEQING_COL].notna()]),
        })
    teacher_stats.sort(key=lambda x: x["valid_entries"], reverse=True)
    print(f"  共 {len(teacher_stats)} 位老师有学情总结")

    # 3. 抽样20位老师
    print(f"\n[3] 抽取 {sample_size} 位老师...")
    # 优先选择有较多学情总结的老师（评估更准确）
    eligible = [t for t in teacher_stats if t["valid_entries"] >= 2]
    if len(eligible) < sample_size:
        eligible = teacher_stats[:sample_size]
    sampled = random.sample(eligible, min(sample_size, len(eligible)))
    print(f"  已抽取 {len(sampled)} 位老师")

    # 4. 对每位老师进行评估
    print("\n[4] 逐老师评估学情总结...")
    results = []

    for item in sampled:
        teacher = item["teacher"]
        art_name = item["art_name"]
        entries = valid_df[valid_df[TEACHER_COL] == teacher].copy()

        # 评估每条学情总结
        entry_results = []
        valid_texts = []
        all_matched = True
        all_invalid = True
        stages = set()

        for _, row in entries.iterrows():
            text = str(row[XUEQING_COL]).strip()
            student_id = row[STUDENT_ID_COL]
            create_time = row[XUEQING_TIME_COL] if pd.notna(row[XUEQING_TIME_COL]) else ""

            match_result = match_template(text)

            entry_results.append({
                "student_id": student_id,
                "text": text,
                "create_time": str(create_time),
                "matched": match_result["matched"],
                "stage": match_result["stage"],
                "score": match_result["score"],
                "reason": match_result["reason"],
            })

            if match_result["matched"]:
                valid_texts.append(text)
                all_invalid = False
                if match_result["stage"] != "未明确阶段":
                    stages.add(match_result["stage"])
            else:
                all_matched = False

        # 评估同质化
        homogeneity = evaluate_homogeneity(valid_texts)

        # 综合评级
        if all_invalid:
            final_grade = "不及格"
            final_reason = "所有学情总结均不符合模板要求，疑似课后反馈内容"
        elif all_matched:
            if homogeneity["level"] in ("优秀", "仅1条"):
                final_grade = "优秀"
            else:
                final_grade = "及格"
            final_reason = f"符合模板 + {homogeneity['reason']}"
        else:
            matched_count = sum(1 for e in entry_results if e["matched"])
            total_count = len(entry_results)
            if matched_count / total_count >= 0.5:
                if homogeneity["level"] in ("优秀", "仅1条"):
                    final_grade = "及格（部分不符合）"
                else:
                    final_grade = "及格"
                final_reason = f"部分符合模板（{matched_count}/{total_count}条），{homogeneity['reason']}"
            else:
                final_grade = "不及格"
                final_reason = f"大部分不符合模板（仅{matched_count}/{total_count}条符合），内容疑似课后反馈"

        teacher_result = {
            "teacher": teacher,
            "art_name": art_name,
            "total_entries": len(entries),
            "matched_entries": sum(1 for e in entry_results if e["matched"]),
            "invalid_entries": sum(1 for e in entry_results if not e["matched"]),
            "stages": "、".join(stages) if stages else "未明确",
            "homogeneity_level": homogeneity["level"],
            "avg_similarity": homogeneity["avg_similarity"],
            "max_similarity": homogeneity["max_similarity"],
            "final_grade": final_grade,
            "final_reason": final_reason,
            "entries": entry_results,
        }
        results.append(teacher_result)
        print(f"  {teacher}({art_name}): {final_grade} | 符合{teacher_result['matched_entries']}/{len(entries)}条 | {homogeneity['level']}")

    # 5. 统计汇总
    print("\n[5] 统计汇总:")
    grade_counts = {}
    for r in results:
        g = r["final_grade"]
        grade_counts[g] = grade_counts.get(g, 0) + 1
    for grade, count in sorted(grade_counts.items()):
        print(f"  {grade}: {count}人")

    # 6. 输出Excel
    print(f"\n[6] 输出到 {output_file}...")
    _write_excel(results, valid_df, output_file)

    print("\n✅ 分析完成!")
    return OUTPUT_FILE


# ══════════════════════════════════════════════════════════════════
# Excel 输出
# ══════════════════════════════════════════════════════════════════
def _write_excel(results: list, original_df: pd.DataFrame, output_file: Path):
    """生成格式化的抽查结果 Excel."""

    wb = openpyxl.Workbook()

    # ── 颜色定义 ──
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="微软雅黑", size=10)
    BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

    # ═══════════════════════════════════════════════════════════════
    # Sheet 1: 抽查总览
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "抽查总览"

    # 标题
    ws1.merge_cells("A1:J1")
    ws1["A1"] = f"学情总结抽查结果 — {datetime.now().strftime('%Y年%m月%d日')}"
    ws1["A1"].font = Font(name="微软雅黑", size=14, bold=True)
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:J2")
    ws1["A2"] = f"抽查维度：老师维度 | 抽样数量：{len(results)}位老师 | 协作池：（益智）海外学管2026年5月协作池-豌豆-6月服务池"
    ws1["A2"].font = Font(name="微软雅黑", size=10, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # 表头
    headers = ["序号", "老师姓名", "老师艺名", "学情总结总数", "符合模板数",
               "不符合数", "适用阶段", "同质化评估", "平均相似度", "综合评级", "评估说明"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # 数据行
    for row_idx, r in enumerate(results, 5):
        ws1.cell(row=row_idx, column=1, value=row_idx - 4).font = BODY_FONT
        ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=2, value=r["teacher"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=3, value=r["art_name"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=4, value=r["total_entries"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=5, value=r["matched_entries"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=6, value=r["invalid_entries"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=7, value=r["stages"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=8, value=r["homogeneity_level"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=9, value=r["avg_similarity"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=9).number_format = '0.0%'

        grade_cell = ws1.cell(row=row_idx, column=10, value=r["final_grade"])
        grade_cell.font = BOLD_FONT
        grade_cell.alignment = Alignment(horizontal="center")
        if "优秀" in r["final_grade"]:
            grade_cell.fill = GREEN_FILL
        elif "及格" in r["final_grade"]:
            grade_cell.fill = YELLOW_FILL
        elif "不及格" in r["final_grade"]:
            grade_cell.fill = RED_FILL

        ws1.cell(row=row_idx, column=11, value=r["final_reason"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=11).alignment = WRAP_ALIGN

        for col_idx in range(1, 12):
            ws1.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    # 统计行
    stats_row = 5 + len(results) + 1
    ws1.cell(row=stats_row, column=1, value="统计").font = BOLD_FONT
    ws1.cell(row=stats_row, column=2, value=f"共 {len(results)} 位老师").font = BOLD_FONT

    grade_counts = {}
    for r in results:
        g = "优秀" if "优秀" in r["final_grade"] else ("及格" if "及格" in r["final_grade"] else "不及格")
        grade_counts[g] = grade_counts.get(g, 0) + 1
    ws1.cell(row=stats_row, column=3, value=f"优秀: {grade_counts.get('优秀', 0)} | 及格: {grade_counts.get('及格', 0)} | 不及格: {grade_counts.get('不及格', 0)}").font = BOLD_FONT

    # 列宽
    col_widths = [6, 12, 15, 12, 12, 10, 18, 15, 12, 15, 50]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════════════════
    # Sheet 2: 详细条目
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("详细条目")

    detail_headers = ["序号", "老师姓名", "老师艺名", "学员ID", "创建时间",
                      "学情总结内容", "是否匹配模板", "匹配阶段", "模板匹配度",
                      "评估说明", "老师综合评级"]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    row_num = 2
    seq = 0
    for r in results:
        for entry in r["entries"]:
            seq += 1
            ws2.cell(row=row_num, column=1, value=seq).font = BODY_FONT
            ws2.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")
            ws2.cell(row=row_num, column=2, value=r["teacher"]).font = BODY_FONT
            ws2.cell(row=row_num, column=3, value=r["art_name"]).font = BODY_FONT
            ws2.cell(row=row_num, column=4, value=entry["student_id"]).font = BODY_FONT
            ws2.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
            ws2.cell(row=row_num, column=5, value=entry["create_time"]).font = BODY_FONT
            ws2.cell(row=row_num, column=6, value=entry["text"]).font = BODY_FONT
            ws2.cell(row=row_num, column=6).alignment = WRAP_ALIGN

            match_cell = ws2.cell(row=row_num, column=7, value="是" if entry["matched"] else "否")
            match_cell.font = BODY_FONT
            match_cell.alignment = Alignment(horizontal="center")
            match_cell.fill = GREEN_FILL if entry["matched"] else RED_FILL

            ws2.cell(row=row_num, column=8, value=entry["stage"]).font = BODY_FONT
            ws2.cell(row=row_num, column=9, value=entry["score"]).font = BODY_FONT
            ws2.cell(row=row_num, column=9).number_format = '0%'
            ws2.cell(row=row_num, column=10, value=entry["reason"]).font = BODY_FONT
            ws2.cell(row=row_num, column=10).alignment = WRAP_ALIGN

            grade_cell = ws2.cell(row=row_num, column=11, value=r["final_grade"])
            grade_cell.font = BOLD_FONT
            if "优秀" in r["final_grade"]:
                grade_cell.fill = GREEN_FILL
            elif "及格" in r["final_grade"]:
                grade_cell.fill = YELLOW_FILL
            elif "不及格" in r["final_grade"]:
                grade_cell.fill = RED_FILL

            for col_idx in range(1, 12):
                ws2.cell(row=row_num, column=col_idx).border = THIN_BORDER

            row_num += 1

    # 列宽
    detail_widths = [6, 12, 15, 12, 22, 70, 12, 15, 12, 50, 15]
    for i, w in enumerate(detail_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════════════════
    # Sheet 3: 评估规则说明
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("评估规则")

    rules = [
        ["评估维度", "规则说明"],
        ["抽查维度", "以老师为维度，同一老师发送的多条信息作为一组进行评估"],
        ["模板匹配", "学情总结需包含编号条目（1. 2. 3...）和模板关键词（记忆犹新、较好知识点、提升知识点、家长痛点等）"],
        ["s1-3阶段", "学员兴趣：专注力、记忆犹新、学习兴趣/课堂活跃度、较好知识点、提升知识点、家长痛点、未续费顾虑点"],
        ["s4-6阶段", "学员习惯：学习习惯、记忆犹新、笔记草稿、较好知识点、提升知识点、家长痛点、未续费顾虑点"],
        ["s7-9阶段", "学员成绩：校内成绩、较好知识点、提升知识点、学习习惯、记忆犹新、家长痛点、未续费顾虑点"],
        ["评级-优秀", "符合模板要求 + 同质化程度低（平均相似度<35%，每条学情总结都有个性化内容）"],
        ["评级-及格", "符合模板要求 + 同质化程度高（平均相似度≥35%，多条内容有重复）"],
        ["评级-不及格", "发送内容不符合模板要求，且疑似课后反馈内容（如短文本、无编号结构、缺少关键词）"],
        ["同质化计算", "使用文本相似度算法(SequenceMatcher)计算同一老师多条学情总结两两之间的相似度，取平均值"],
    ]

    for row_idx, (col1, col2) in enumerate(rules, 1):
        cell1 = ws3.cell(row=row_idx, column=1, value=col1)
        cell2 = ws3.cell(row=row_idx, column=2, value=col2)
        if row_idx == 1:
            cell1.font = HEADER_FONT
            cell2.font = HEADER_FONT
            cell1.fill = HEADER_FILL
            cell2.fill = HEADER_FILL
        else:
            cell1.font = BOLD_FONT
            cell2.font = BODY_FONT
        cell1.alignment = WRAP_ALIGN
        cell2.alignment = WRAP_ALIGN
        cell1.border = THIN_BORDER
        cell2.border = THIN_BORDER

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 90

    wb.save(output_file)


if __name__ == "__main__":
    main()
