#!/usr/bin/env python3
"""学情总结抽查分析 - 按老师维度抽样、模板匹配、同质化评估.

支持完整端到端流程:
  --download           调用 export_report.py 重新从 BI 下载数据
  --monthly            按当月维度过滤(学情总结创建时间 ∈ [当月1号, end_date])
  --end-date           过滤截止日期, 默认今天
  --pool               当月协作池标签, 用于 BI 筛选
  --broadcast          抽查完成后将结果摘要发送到钉钉群

Usage:
    # 端到端: 下载 + 当月过滤 + 抽查 + 播报
    python3 xueqing_review.py --download --monthly --broadcast --sample 5

    # 仅抽查现有文件
    python3 xueqing_review.py --input <input.xlsx> --output <output.xlsx> [--sample 20] [--seed 42]
"""

import argparse
import json
import os
import subprocess
import sys
import urllib.request
import pandas as pd
import re
import random
from difflib import SequenceMatcher
from pathlib import Path
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# ══════════════════════════════════════════════════════════════════
# 默认配置
# ══════════════════════════════════════════════════════════════════
# WORKDIR 解析优先级（避免硬编码绝对路径，便于跨机器交接）：
#   1. CLI 参数 --workdir
#   2. 环境变量 XUEQING_WORKDIR
#   3. 用户主目录下的 Desktop/学情抽查/
#   4. 当前工作目录
SCRIPT_DIR = Path(__file__).resolve().parent


def _resolve_workdir(cli_workdir: str | None = None) -> Path:
    if cli_workdir:
        return Path(cli_workdir).expanduser().resolve()
    env_workdir = os.environ.get("XUEQING_WORKDIR")
    if env_workdir:
        return Path(env_workdir).expanduser().resolve()
    desktop_default = Path.home() / "Desktop" / "学情抽查"
    if desktop_default.exists():
        return desktop_default
    return Path.cwd()


WORKDIR = _resolve_workdir()
DEFAULT_INPUT_NAME = "教学&学管协作跟进明细.xlsx"
DEFAULT_OUTPUT_NAME = "学情总结抽查结果.xlsx"
DEFAULT_INPUT = WORKDIR / DEFAULT_INPUT_NAME
DEFAULT_OUTPUT = WORKDIR / DEFAULT_OUTPUT_NAME
# export_report.py 优先使用 WORKDIR 下的，找不到则用 skill 同目录
EXPORT_SCRIPT_CANDIDATES = [WORKDIR / "export_report.py", SCRIPT_DIR / "export_report.py"]
EXPORT_SCRIPT = next((p for p in EXPORT_SCRIPT_CANDIDATES if p.exists()), EXPORT_SCRIPT_CANDIDATES[0])
SECRETS_ENV = Path.home() / ".claude" / "secrets" / "intro_monitor.env"
DEFAULT_POOL = "（益智）海外学管2026年5月协作池-豌豆-6月服务池"
DEFAULT_SAMPLE = 20
DEFAULT_SEED = 42
HEADER_ROW = 9  # 数据表头所在行（第10行，索引9）
XUEQING_COL = "学情总结"
TEACHER_COL = "主讲姓名"
ART_NAME_COL = "上课老师艺名"
STUDENT_ID_COL = "学员id"
XUEQING_TIME_COL = "学情总结创建时间"

# 模板关键词（三个阶段通用关键词）
TEMPLATE_KEYWORDS = [
    "记忆犹新", "较好知识点", "需要提升知识点", "提升知识点",
    "家长痛点", "未续费顾虑点", "顾虑点"
]

# 各阶段特有关键词
STAGE_KEYWORDS = {
    "s1-3": ["专注力", "学习兴趣", "课堂活跃度"],
    "s4-6": ["学习习惯", "笔记草稿", "三率"],
    "s7-9": ["校内成绩", "校内学习成绩", "学校考试"],
}

# 不符合模板的短文本特征（无效内容，不属于学情总结也不属于特殊情况）
INVALID_PATTERNS = [
    r'^.{0,15}$',  # 太短的文本（15字以内）
    r'^(已加微|待跟进|待沟通|未接|已满)',
    r'^(rhya|.+课上很认真|.+很认真|孩子.+专注|宝贝.+认真)',
]

# 特殊情况（停课退费等），合格 ② 路径：不要求按模板反馈
SPECIAL_CASE_PATTERNS = [
    (r'^(停课|退费|请假|不续费|已流失|已转走|已续费|转介绍)', "停课/退费等特殊情况"),
]

# 详尽描述关键词：优秀必须命中至少 1 个（即包含具体维度反馈）
DETAIL_KEYWORDS = [
    "课堂参与度", "参与度",
    "兴趣度", "学习兴趣",
    "进步点", "进步",
    "专注度", "专注力",
]

# ══════════════════════════════════════════════════════════════════
# 反模式黑名单 (Anti-pattern Blacklist)
# 评估时显式拦截的常见错误用法，命中即直接判为「不符合模板」并在评估说明中标注原因。
# 抽查时若发现新模式，请把 (正则, 名称, 解释) 三元组追加到这里，并同步到
# references/evaluation-rules.md 的「反模式黑名单」章节。
# ══════════════════════════════════════════════════════════════════
ANTI_PATTERNS = [
    # 误把课后反馈模板当学情模板
    (r'(今日课堂|课堂表现|课堂笔记|课堂作业|课堂内容回顾|今天上课|本节课|本堂课表现)',
     "课后反馈模板",
     "不要用课后反馈模板替代学情总结模板：课后反馈聚焦单节课的课堂行为/作业完成，"
     "学情总结要按阶段输出『记忆犹新/较好知识点/提升知识点/家长痛点/未续费顾虑点』等结构化结论"),
    # 续费话术 / 销售模板
    (r'(续费优惠|限时活动|本周下单|名额有限|早鸟价|赠送课时)',
     "续费销售话术",
     "不要把续费话术当学情总结：销售话术不属于学情评估维度"),
    # 流水账 / 日报式描述
    (r'^(早上|中午|下午|晚上|今天).{0,30}(上课|做题|讲了|学了)',
     "日报流水账",
     "不要写成流水账：学情总结需要按阶段维度结构化输出，而不是按时间叙述"),
    # 仅复述题目/知识点而无评估
    (r'^(讲解了|复习了|学习了|完成了)[^。\n]{0,40}$',
     "无评估的知识点复述",
     "不要只罗列『讲解了 X / 学习了 Y』：必须给出『较好/需提升/家长痛点』等评估结论"),
    # 通用模板复制粘贴（同一段话用于所有学生）
    (r'(此学员|该学员|这个孩子|这位学员).{0,15}(表现良好|需要努力|继续加油|很棒)$',
     "万能套话",
     "不要用『表现良好/继续加油』收尾的万能套话：必须落到具体知识点和家长沟通点"),
]


def check_anti_pattern(text: str) -> dict | None:
    """命中反模式黑名单时返回 {name, reason}，否则返回 None."""
    for pattern, name, reason in ANTI_PATTERNS:
        if re.search(pattern, text):
            return {"name": name, "reason": reason}
    return None


# ══════════════════════════════════════════════════════════════════
# 模板匹配
# ══════════════════════════════════════════════════════════════════
def is_invalid_entry(text: str) -> bool:
    """判断是否为无效学情总结（非模板格式且非特殊情况）."""
    text = str(text).strip()
    if not text:
        return True
    for pattern in INVALID_PATTERNS:
        if re.match(pattern, text, re.IGNORECASE):
            return True
    return False


def check_special_case(text: str) -> str | None:
    """命中特殊情况（停课/退费等）时返回原因标签，否则 None."""
    text = str(text).strip()
    for pattern, label in SPECIAL_CASE_PATTERNS:
        if re.match(pattern, text):
            return label
    return None


def has_detail_keywords(text: str) -> list[str]:
    """返回文本命中的详尽描述关键词列表（用于优秀判定）."""
    return [kw for kw in DETAIL_KEYWORDS if kw in text]


def count_numbered_items(text: str) -> int:
    """统计文本中编号条目数量 (1. 2. 3. 等)."""
    matches = re.findall(r'(?:^|\n)\s*\d+[\.、．]\s*', text)
    return len(matches)


def match_template(text: str) -> dict:
    """评估学情总结是否匹配模板.

    返回:
        {
            "matched": bool,            # 是否符合模板
            "stage": str,               # 匹配的阶段
            "score": float,             # 匹配分数 0-1
            "reason": str,              # 评估说明
            "is_special_case": bool,    # 是否为停课退费等特殊情况
            "special_case_label": str,  # 特殊情况标签
            "detail_hits": list[str],   # 命中的详尽描述关键词
        }
    """
    text = str(text).strip()
    detail_hits = has_detail_keywords(text)

    # 0. 先检查是否特殊情况（合格 ② 路径，不要求模板）
    sc_label = check_special_case(text)
    if sc_label:
        return {
            "matched": False,
            "stage": "特殊情况",
            "score": 0.0,
            "reason": f"{sc_label}：合格 ② 情形，不要求按模板反馈",
            "is_special_case": True,
            "special_case_label": sc_label,
            "detail_hits": detail_hits,
        }

    # 1. 检查是否无效
    if is_invalid_entry(text):
        return {
            "matched": False,
            "stage": "无效",
            "score": 0.0,
            "reason": "不符合模板要求，疑似课后反馈或状态备注",
            "is_special_case": False,
            "special_case_label": "",
            "detail_hits": detail_hits,
        }

    # 1.5 反模式黑名单（命中即判不符合模板，且给出明确原因）
    anti = check_anti_pattern(text)
    if anti:
        return _finalize_match({
            "matched": False,
            "stage": f"反模式：{anti['name']}",
            "score": 0.0,
            "reason": f"命中反模式黑名单 [{anti['name']}]：{anti['reason']}",
        }, detail_hits)

    # 2. 检查是否有编号结构
    numbered_items = count_numbered_items(text)

    # 3. 检查关键词匹配
    matched_keywords = []
    for kw in TEMPLATE_KEYWORDS:
        if kw in text:
            matched_keywords.append(kw)

    # 4. 检查阶段匹配
    stage_matched = None
    for stage, keywords in STAGE_KEYWORDS.items():
        for kw in keywords:
            if kw in text:
                stage_matched = stage
                break
        if stage_matched:
            break

    # 5. 综合判断
    # 有编号结构 + 至少3个关键词 = 符合模板
    # 有编号结构 + 至少2个关键词 + 有阶段匹配 = 符合模板
    keyword_count = len(matched_keywords)

    if numbered_items >= 5 and keyword_count >= 3:
        stage = stage_matched or "未明确阶段"
        score = min(1.0, (numbered_items / 7) * 0.5 + (keyword_count / 5) * 0.5)
        return _finalize_match({
            "matched": True,
            "stage": stage,
            "score": round(score, 2),
            "reason": f"符合模板结构（{numbered_items}个编号条目，{keyword_count}个关键词命中：{', '.join(matched_keywords[:4])}）"
        }, detail_hits)
    elif numbered_items >= 3 and keyword_count >= 2 and stage_matched:
        score = min(1.0, (numbered_items / 7) * 0.4 + (keyword_count / 5) * 0.6)
        return _finalize_match({
            "matched": True,
            "stage": stage_matched,
            "score": round(score, 2),
            "reason": f"基本符合模板（{numbered_items}个编号条目，阶段{stage_matched}，{keyword_count}个关键词）"
        }, detail_hits)
    elif numbered_items >= 5 and keyword_count >= 1:
        return _finalize_match({
            "matched": True,
            "stage": stage_matched or "未明确阶段",
            "score": round(0.5, 2),
            "reason": f"有编号结构但关键词较少（{numbered_items}个编号条目，{keyword_count}个关键词），模板匹配度一般"
        }, detail_hits)
    else:
        return _finalize_match({
            "matched": False,
            "stage": "不符合",
            "score": 0.0,
            "reason": f"不符合模板要求（{numbered_items}个编号条目，{keyword_count}个关键词），内容缺少标准结构"
        }, detail_hits)


def _finalize_match(result: dict, detail_hits: list) -> dict:
    """补齐 match_template 各分支共有的字段（特殊情况标识 + 详尽描述命中）."""
    result.setdefault("is_special_case", False)
    result.setdefault("special_case_label", "")
    result["detail_hits"] = detail_hits
    return result


# ══════════════════════════════════════════════════════════════════
# 同质化评估
# ══════════════════════════════════════════════════════════════════
def text_similarity(t1: str, t2: str) -> float:
    """计算两段文本的相似度."""
    return SequenceMatcher(None, t1, t2).ratio()


def evaluate_homogeneity(texts: list) -> dict:
    """评估同一老师多条学情总结的同质化程度.

    同质化低 = 优秀（每条都有个性化内容）
    同质化高 = 及格（多条内容雷同）
    """
    if len(texts) < 2:
        return {
            "level": "仅1条",
            "avg_similarity": 0.0,
            "max_similarity": 0.0,
            "min_similarity": 0.0,
            "reason": "仅有1条学情总结，无法评估同质化"
        }

    # 计算所有配对的相似度
    similarities = []
    for i in range(len(texts)):
        for j in range(i + 1, len(texts)):
            sim = text_similarity(texts[i], texts[j])
            similarities.append(sim)

    avg_sim = sum(similarities) / len(similarities)
    max_sim = max(similarities)
    min_sim = min(similarities)

    if avg_sim < 0.35:
        level = "优秀"
        reason = f"同质化程度低，内容个性化程度高（平均相似度{avg_sim:.0%}）"
    elif avg_sim < 0.55:
        level = "及格"
        reason = f"同质化程度一般，部分内容有重复（平均相似度{avg_sim:.0%}）"
    else:
        level = "及格（偏高）"
        reason = f"同质化程度高，多条内容较雷同（平均相似度{avg_sim:.0%}）"

    return {
        "level": level,
        "avg_similarity": round(avg_sim, 3),
        "max_similarity": round(max_sim, 3),
        "min_similarity": round(min_sim, 3),
        "reason": reason
    }


# ══════════════════════════════════════════════════════════════════
# 端到端流程辅助函数（下载 / 月份过滤 / 钉钉播报）
# ══════════════════════════════════════════════════════════════════
def _self_check_coverage(sampled: list, teacher_stats: list, eligible: list,
                         min_coverage: float) -> dict:
    """抽样覆盖率自检：抽样老师数 / 全部有学情总结的老师数 是否 ≥ min_coverage.

    返回：
        {
            "ok": bool,             # 是否达标
            "coverage": float,      # 抽样占总老师比例
            "min_coverage": float,  # 阈值
            "sampled": int,
            "total_teachers": int,
            "eligible": int,
            "message": str,
        }
    """
    total = len(teacher_stats)
    coverage = (len(sampled) / total) if total else 0.0
    ok = coverage >= min_coverage
    msg = (
        f"抽样老师 {len(sampled)} / 全部 {total} = {coverage:.0%}（阈值 {min_coverage:.0%}）"
        f"，可抽样老师 {len(eligible)}"
    )
    print("\n[3.5] 抽样覆盖率自检")
    if ok:
        print(f"  ✅ 通过：{msg}")
    else:
        print(f"  ⚠ 未通过：{msg}")
        suggested = max(int(total * min_coverage + 0.999), len(sampled) + 1)
        print(f"     建议：将 --sample 调整到 ≥ {suggested}，或在 --min-coverage 参数中显式降低阈值。")
    return {
        "ok": ok,
        "coverage": round(coverage, 4),
        "min_coverage": min_coverage,
        "sampled": len(sampled),
        "total_teachers": total,
        "eligible": len(eligible),
        "message": msg,
    }


def _run_export(end_date: str, pool: str) -> int:
    """调用 export_report.py 重新从 BI 下载报表."""
    print("=" * 60)
    print(f"[Step] 通过 BI 重新下载报表 (截止 {end_date})")
    print("=" * 60)
    if not EXPORT_SCRIPT.exists():
        print(f"⚠ 未找到 {EXPORT_SCRIPT}，跳过下载")
        return 1
    cmd = [sys.executable, str(EXPORT_SCRIPT), "--end-date", end_date, "--pool", pool]
    print(f"运行: {' '.join(cmd)}")
    result = subprocess.run(cmd, cwd=str(WORKDIR))
    if result.returncode != 0:
        print(f"⚠ export_report.py 退出码 {result.returncode}，将尝试用已有文件继续")
    return result.returncode


def _filter_monthly(raw_file: Path, end_date: str) -> Path:
    """按 学情总结创建时间 ∈ [当月1号, end_date] 过滤记录,落到临时 xlsx,
    并保持 skill 期望的 9 行表头空白 + 第 10 行表头布局。"""
    print("=" * 60)
    print(f"[Step] 按当月过滤: {end_date[:7]}-01 ~ {end_date}")
    print("=" * 60)
    if not raw_file.exists():
        raise FileNotFoundError(f"原始文件不存在: {raw_file}")

    end_dt = pd.Timestamp(end_date) + pd.Timedelta(days=1)
    start_dt = pd.Timestamp(end_date[:7] + "-01")

    df = pd.read_excel(raw_file, header=HEADER_ROW, engine="openpyxl")
    print(f"  原始行数: {len(df)}")

    df["_xq_time"] = pd.to_datetime(df[XUEQING_TIME_COL], errors="coerce")
    mask = df["_xq_time"].between(start_dt, end_dt, inclusive="left") & df[XUEQING_COL].notna()
    filtered = df[mask].drop(columns=["_xq_time"]).copy()
    print(f"  当月有学情总结的记录: {len(filtered)}")

    out = WORKDIR / f"教学&学管协作跟进明细_当月_{end_date}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(HEADER_ROW):
        ws.append([])
    ws.append(list(filtered.columns))
    for row in filtered.itertuples(index=False):
        ws.append([
            v if not (isinstance(v, float) and pd.isna(v)) else None
            for v in row
        ])
    wb.save(out)
    print(f"  落地文件: {out}")
    return out


def _load_webhook_url() -> str | None:
    """从 ~/.claude/secrets/intro_monitor.env 读取 DINGTALK_WEBHOOK_URL."""
    if not SECRETS_ENV.exists():
        return None
    for line in SECRETS_ENV.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        if k.strip() == "DINGTALK_WEBHOOK_URL":
            return v.strip().strip('"').strip("'")
    return None


def _broadcast_to_dingtalk(results: list, end_date: str) -> bool:
    """把抽查结果汇总以 markdown 形式发送到钉钉群（仅播报结果）."""
    webhook = _load_webhook_url()
    if not webhook:
        print("⚠ 未配置 DINGTALK_WEBHOOK_URL（~/.claude/secrets/intro_monitor.env），跳过播报")
        return False

    grade_counts = {"优秀": 0, "合格": 0, "不合格": 0}
    for r in results:
        if r["final_grade"] in grade_counts:
            grade_counts[r["final_grade"]] += 1

    title = f"学情总结抽查结果（截止 {end_date}）"
    lines = [
        f"## {title}",
        "",
        f"- 抽样老师数量：**{len(results)}** 位",
        f"- 优秀：<font color=\"#52C41A\">**{grade_counts['优秀']}**</font> | "
        f"合格：<font color=\"#FAAD14\">**{grade_counts['合格']}**</font> | "
        f"不合格：<font color=\"#FF4D4F\">**{grade_counts['不合格']}**</font>",
        "",
        "### 评级明细",
        "",
        "| 老师 | 艺名 | 学情总结 | 符合数 | 评级 | 评估说明 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for r in results:
        grade = r["final_grade"]
        if grade == "优秀":
            grade_md = f"<font color=\"#52C41A\">**{grade}**</font>"
        elif grade == "不合格":
            grade_md = f"<font color=\"#FF4D4F\">**{grade}**</font>"
        else:
            grade_md = f"<font color=\"#FAAD14\">**{grade}**</font>"
        reason = (r["final_reason"] or "").replace("|", "/").replace("\n", " ")
        if len(reason) > 60:
            reason = reason[:60] + "…"
        lines.append(
            f"| {r['teacher']} | {r['art_name']} | {r['total_entries']} | "
            f"{r['matched_entries']} | {grade_md} | {reason} |"
        )

    payload = {
        "msgtype": "markdown",
        "markdown": {"title": title, "text": "\n".join(lines)},
    }
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        webhook,
        data=data,
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            body = resp.read().decode("utf-8", errors="replace")
        result = json.loads(body) if body else {}
        if result.get("errcode", 0) == 0:
            print(f"✅ 已播报到钉钉群（{len(results)} 位老师）")
            return True
        print(f"⚠ 钉钉返回: {body}")
        return False
    except Exception as e:
        print(f"⚠ 钉钉播报失败: {e}")
        return False


# ══════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════
def main():
    today = date.today().isoformat()
    parser = argparse.ArgumentParser(description="学情总结抽查分析")
    parser.add_argument("--input", default="", help="输入Excel路径")
    parser.add_argument("--output", default="", help="输出Excel路径（默认按日期生成）")
    parser.add_argument("--sample", type=int, default=DEFAULT_SAMPLE, help="抽样老师数量")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED, help="随机种子")
    parser.add_argument("--end-date", default=today, help="过滤截止日期 YYYY-MM-DD,默认今天")
    parser.add_argument("--pool", default=DEFAULT_POOL, help="当月协作池标签,用于 BI 筛选")
    parser.add_argument("--download", action="store_true", help="先调用 export_report.py 重新下载")
    parser.add_argument("--monthly", action="store_true", help="按当月维度过滤(学情总结创建时间)")
    parser.add_argument("--broadcast", action="store_true", help="抽查完成后播报到钉钉群")
    parser.add_argument("--workdir", default="", help="工作目录（默认 $XUEQING_WORKDIR 或 ~/Desktop/学情抽查）")
    parser.add_argument("--min-coverage", type=float, default=0.8,
                        help="抽样自检：抽样老师占可抽样老师的最低比例（默认 0.8 即 80%）")
    args = parser.parse_args()

    # 解析 workdir 后刷新依赖 WORKDIR 的默认路径
    global WORKDIR, DEFAULT_INPUT, DEFAULT_OUTPUT, EXPORT_SCRIPT
    WORKDIR = _resolve_workdir(args.workdir or None)
    DEFAULT_INPUT = WORKDIR / DEFAULT_INPUT_NAME
    DEFAULT_OUTPUT = WORKDIR / DEFAULT_OUTPUT_NAME
    EXPORT_SCRIPT = next(
        (p for p in [WORKDIR / "export_report.py", SCRIPT_DIR / "export_report.py"] if p.exists()),
        WORKDIR / "export_report.py",
    )
    print(f"  WORKDIR = {WORKDIR}")
    # --input 为空时回填为 WORKDIR 下的默认输入文件
    if not args.input:
        args.input = str(DEFAULT_INPUT)

    end_date = args.end_date
    sample_size = args.sample
    random.seed(args.seed)

    # 0. 可选: 重新下载
    if args.download:
        _run_export(end_date, args.pool)

    # 0.5 可选: 按当月过滤
    if args.monthly:
        input_file = _filter_monthly(Path(args.input), end_date)
    else:
        input_file = Path(args.input)

    if args.output:
        output_file = Path(args.output)
    elif args.monthly:
        output_file = WORKDIR / f"学情总结抽查结果_当月_{end_date}.xlsx"
    else:
        output_file = DEFAULT_OUTPUT

    print("=" * 60)
    print("学情总结抽查分析")
    print("=" * 60)

    # 1. 读取数据
    print("\n[1] 读取数据...")
    df = pd.read_excel(input_file, header=HEADER_ROW)
    print(f"  总行数: {len(df)}")

    # 筛选有效学情总结（非空且非无效）
    valid_mask = df[XUEQING_COL].notna() & (df[XUEQING_COL].astype(str).str.strip() != '')
    valid_df = df[valid_mask].copy()
    print(f"  有学情总结的记录: {len(valid_df)}")

    # 2. 按老师分组
    print("\n[2] 按老师维度分组...")
    teacher_groups = valid_df.groupby(TEACHER_COL, dropna=False)
    teacher_stats = []
    for teacher, group in teacher_groups:
        if pd.isna(teacher) or str(teacher).strip() == '':
            continue
        teacher_stats.append({
            "teacher": teacher,
            "art_name": group[ART_NAME_COL].dropna().unique()[0] if len(group[ART_NAME_COL].dropna().unique()) > 0 else "",
            "total_entries": len(group),
            "valid_entries": len(group[group[XUEQING_COL].notna()]),
        })
    teacher_stats.sort(key=lambda x: x["valid_entries"], reverse=True)
    print(f"  共 {len(teacher_stats)} 位老师有学情总结")

    # 3. 抽样20位老师
    print(f"\n[3] 抽取 {sample_size} 位老师...")
    # 优先选择有较多学情总结的老师（评估更准确）
    eligible = [t for t in teacher_stats if t["valid_entries"] >= 2]
    if len(eligible) < sample_size:
        eligible = teacher_stats[:sample_size]
    sampled = random.sample(eligible, min(sample_size, len(eligible)))
    print(f"  已抽取 {len(sampled)} 位老师")

    # 3.5 校验自检：抽样覆盖率 ≥ min_coverage
    coverage_check = _self_check_coverage(
        sampled=sampled,
        teacher_stats=teacher_stats,
        eligible=eligible,
        min_coverage=args.min_coverage,
    )

    # 4. 对每位老师进行评估
    print("\n[4] 逐老师评估学情总结...")
    results = []

    for item in sampled:
        teacher = item["teacher"]
        art_name = item["art_name"]
        entries = valid_df[valid_df[TEACHER_COL] == teacher].copy()

        # 评估每条学情总结
        entry_results = []
        matched_texts = []
        stages = set()

        for _, row in entries.iterrows():
            text = str(row[XUEQING_COL]).strip()
            student_id = row[STUDENT_ID_COL]
            create_time = row[XUEQING_TIME_COL] if pd.notna(row[XUEQING_TIME_COL]) else ""

            match_result = match_template(text)

            entry_results.append({
                "student_id": student_id,
                "text": text,
                "create_time": str(create_time),
                "matched": match_result["matched"],
                "stage": match_result["stage"],
                "score": match_result["score"],
                "reason": match_result["reason"],
                "is_special_case": match_result["is_special_case"],
                "special_case_label": match_result["special_case_label"],
                "detail_hits": match_result["detail_hits"],
            })

            if match_result["matched"]:
                matched_texts.append(text)
                if match_result["stage"] != "未明确阶段":
                    stages.add(match_result["stage"])

        # 评估同质化（仅基于符合模板的条目，避免特殊情况文本干扰）
        homogeneity = evaluate_homogeneity(matched_texts)

        # ── 新评级标准（与 references/evaluation-rules.md 对齐）──
        #   优秀: 全部按模板反馈 + 1-3 点详尽描述命中（参与度/兴趣度/进步点/专注度）+ 个性化达标
        #   合格 ①: 全部按模板反馈 + 个性化达标
        #   合格 ②: 全部为停课退费等特殊情况
        #   不合格 ①: 存在未按模板反馈的条目（且非特殊情况）
        #   不合格 ②: 同质化高（与历史/学员高度相似，相似度 ≥ 0.55）
        #   不合格 ③: 未完成个性化反馈（同质化条件下与 ② 等价）
        total_count = len(entry_results)
        matched_count = sum(1 for e in entry_results if e["matched"])
        special_case_count = sum(1 for e in entry_results if e["is_special_case"])
        non_match_non_special = total_count - matched_count - special_case_count

        # 个性化判定：≥2 条匹配且平均相似度 < 0.55 视为完成个性化反馈
        avg_sim = homogeneity["avg_similarity"]
        high_similarity = matched_count >= 2 and avg_sim >= 0.55
        personalized_ok = matched_count <= 1 or avg_sim < 0.55

        # 详尽描述命中：所有匹配条目都有至少 1 个 DETAIL_KEYWORDS 命中
        matched_entries = [e for e in entry_results if e["matched"]]
        all_have_detail = bool(matched_entries) and all(e["detail_hits"] for e in matched_entries)

        if non_match_non_special > 0:
            final_grade = "不合格"
            final_reason = (
                f"不合格①：{non_match_non_special}/{total_count} 条未按模板反馈"
                + (f"（其中 {special_case_count} 条为特殊情况不计）" if special_case_count else "")
            )
        elif high_similarity:
            final_grade = "不合格"
            final_reason = (
                f"不合格②/③：同质化高，未完成个性化反馈（平均相似度 {avg_sim:.0%} ≥ 55%）"
            )
        elif matched_count == 0 and special_case_count == total_count and total_count > 0:
            final_grade = "合格"
            final_reason = f"合格②：全部 {total_count} 条均为停课退费等特殊情况"
        elif all_have_detail and personalized_ok and matched_count >= 1:
            sample_hits = matched_entries[0]["detail_hits"][:3] if matched_entries else []
            final_grade = "优秀"
            sim_part = f"，平均相似度 {avg_sim:.0%}" if matched_count >= 2 else ""
            final_reason = (
                f"优秀：{matched_count}/{total_count} 条按模板反馈，含详尽描述"
                f"（命中关键词如：{ '、'.join(sample_hits) or '—' }）{sim_part}"
            )
        else:
            sim_part = f"（平均相似度 {avg_sim:.0%}）" if matched_count >= 2 else ""
            missing_detail = matched_count > 0 and not all_have_detail
            extra = "；缺少详尽维度描述（参与度/兴趣度/进步点/专注度）" if missing_detail else ""
            final_reason = (
                f"合格①：{matched_count}/{total_count} 条按模板反馈，个性化达标{sim_part}{extra}"
            )
            final_grade = "合格"

        teacher_result = {
            "teacher": teacher,
            "art_name": art_name,
            "total_entries": len(entries),
            "matched_entries": matched_count,
            "special_case_entries": special_case_count,
            "invalid_entries": non_match_non_special,
            "stages": "、".join(stages) if stages else "未明确",
            "homogeneity_level": homogeneity["level"],
            "avg_similarity": homogeneity["avg_similarity"],
            "max_similarity": homogeneity["max_similarity"],
            "final_grade": final_grade,
            "final_reason": final_reason,
            "entries": entry_results,
        }
        results.append(teacher_result)
        print(
            f"  {teacher}({art_name}): {final_grade} | "
            f"匹配{matched_count} / 特殊{special_case_count} / 未达标{non_match_non_special} "
            f"| 共{len(entries)}条 | {homogeneity['level']}"
        )

    # 5. 统计汇总
    print("\n[5] 统计汇总:")
    grade_counts = {}
    for r in results:
        g = r["final_grade"]
        grade_counts[g] = grade_counts.get(g, 0) + 1
    for grade, count in sorted(grade_counts.items()):
        print(f"  {grade}: {count}人")

    # 6. 输出Excel
    print(f"\n[6] 输出到 {output_file}...")
    _write_excel(results, valid_df, output_file, coverage_check)

    print("\n✅ 分析完成!")

    if args.broadcast:
        print("\n[7] 播报抽查结果到钉钉群...")
        _broadcast_to_dingtalk(results, end_date)

    return output_file


# ══════════════════════════════════════════════════════════════════
# Excel 输出
# ══════════════════════════════════════════════════════════════════
def _write_excel(results: list, original_df: pd.DataFrame, output_file: Path,
                 coverage_check: dict | None = None):
    """生成格式化的抽查结果 Excel."""

    wb = openpyxl.Workbook()

    # ── 颜色定义 ──
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="微软雅黑", size=10)
    BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

    # ═══════════════════════════════════════════════════════════════
    # Sheet 1: 抽查总览
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "抽查总览"

    # 标题
    ws1.merge_cells("A1:L1")
    ws1["A1"] = f"学情总结抽查结果 — {datetime.now().strftime('%Y年%m月%d日')}"
    ws1["A1"].font = Font(name="微软雅黑", size=14, bold=True)
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:L2")
    ws1["A2"] = f"抽查维度：老师维度 | 抽样数量：{len(results)}位老师 | 协作池：（益智）海外学管2026年5月协作池-豌豆-6月服务池"
    ws1["A2"].font = Font(name="微软雅黑", size=10, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # 表头
    headers = ["序号", "老师姓名", "老师艺名", "学情总结总数", "符合模板数",
               "不符合数", "适用阶段", "同质化评估", "平均相似度", "综合评级", "评估说明",
               "特殊情况数"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # 数据行
    for row_idx, r in enumerate(results, 5):
        ws1.cell(row=row_idx, column=1, value=row_idx - 4).font = BODY_FONT
        ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=2, value=r["teacher"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=3, value=r["art_name"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=4, value=r["total_entries"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=5, value=r["matched_entries"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=6, value=r["invalid_entries"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=7, value=r["stages"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=8, value=r["homogeneity_level"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=9, value=r["avg_similarity"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=9).number_format = '0.0%'

        grade_cell = ws1.cell(row=row_idx, column=10, value=r["final_grade"])
        grade_cell.font = BOLD_FONT
        grade_cell.alignment = Alignment(horizontal="center")
        if r["final_grade"] == "优秀":
            grade_cell.fill = GREEN_FILL
        elif r["final_grade"] == "不合格":
            grade_cell.fill = RED_FILL
        elif r["final_grade"] == "合格":
            grade_cell.fill = YELLOW_FILL

        ws1.cell(row=row_idx, column=11, value=r["final_reason"]).font = BODY_FONT
        ws1.cell(row=row_idx, column=11).alignment = WRAP_ALIGN

        ws1.cell(row=row_idx, column=12, value=r.get("special_case_entries", 0)).font = BODY_FONT
        ws1.cell(row=row_idx, column=12).alignment = Alignment(horizontal="center")

        for col_idx in range(1, 13):
            ws1.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    # 统计行
    stats_row = 5 + len(results) + 1
    ws1.cell(row=stats_row, column=1, value="统计").font = BOLD_FONT
    ws1.cell(row=stats_row, column=2, value=f"共 {len(results)} 位老师").font = BOLD_FONT

    grade_counts = {"优秀": 0, "合格": 0, "不合格": 0}
    for r in results:
        g = r["final_grade"]
        if g in grade_counts:
            grade_counts[g] += 1
    ws1.cell(row=stats_row, column=3, value=f"优秀: {grade_counts['优秀']} | 合格: {grade_counts['合格']} | 不合格: {grade_counts['不合格']}").font = BOLD_FONT

    # 列宽
    col_widths = [6, 12, 15, 12, 12, 10, 18, 15, 12, 15, 50, 12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════════════════
    # Sheet 2: 详细条目
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("详细条目")

    detail_headers = ["序号", "老师姓名", "老师艺名", "学员ID", "创建时间",
                      "学情总结内容", "是否匹配模板", "匹配阶段", "模板匹配度",
                      "评估说明", "老师综合评级", "详尽描述命中", "特殊情况"]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    row_num = 2
    seq = 0
    for r in results:
        for entry in r["entries"]:
            seq += 1
            ws2.cell(row=row_num, column=1, value=seq).font = BODY_FONT
            ws2.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")
            ws2.cell(row=row_num, column=2, value=r["teacher"]).font = BODY_FONT
            ws2.cell(row=row_num, column=3, value=r["art_name"]).font = BODY_FONT
            ws2.cell(row=row_num, column=4, value=entry["student_id"]).font = BODY_FONT
            ws2.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
            ws2.cell(row=row_num, column=5, value=entry["create_time"]).font = BODY_FONT
            ws2.cell(row=row_num, column=6, value=entry["text"]).font = BODY_FONT
            ws2.cell(row=row_num, column=6).alignment = WRAP_ALIGN

            if entry.get("is_special_case"):
                match_label = "特殊情况"
                match_fill = YELLOW_FILL
            elif entry["matched"]:
                match_label = "是"
                match_fill = GREEN_FILL
            else:
                match_label = "否"
                match_fill = RED_FILL
            match_cell = ws2.cell(row=row_num, column=7, value=match_label)
            match_cell.font = BODY_FONT
            match_cell.alignment = Alignment(horizontal="center")
            match_cell.fill = match_fill

            ws2.cell(row=row_num, column=8, value=entry["stage"]).font = BODY_FONT
            ws2.cell(row=row_num, column=9, value=entry["score"]).font = BODY_FONT
            ws2.cell(row=row_num, column=9).number_format = '0%'
            ws2.cell(row=row_num, column=10, value=entry["reason"]).font = BODY_FONT
            ws2.cell(row=row_num, column=10).alignment = WRAP_ALIGN

            grade_cell = ws2.cell(row=row_num, column=11, value=r["final_grade"])
            grade_cell.font = BOLD_FONT
            if r["final_grade"] == "优秀":
                grade_cell.fill = GREEN_FILL
            elif r["final_grade"] == "不合格":
                grade_cell.fill = RED_FILL
            elif r["final_grade"] == "合格":
                grade_cell.fill = YELLOW_FILL

            ws2.cell(
                row=row_num, column=12,
                value="、".join(entry.get("detail_hits") or []) or "—",
            ).font = BODY_FONT
            ws2.cell(row=row_num, column=12).alignment = Alignment(horizontal="center", wrap_text=True)

            ws2.cell(
                row=row_num, column=13,
                value=entry.get("special_case_label") or ("是" if entry.get("is_special_case") else "—"),
            ).font = BODY_FONT
            ws2.cell(row=row_num, column=13).alignment = Alignment(horizontal="center", wrap_text=True)

            for col_idx in range(1, 14):
                ws2.cell(row=row_num, column=col_idx).border = THIN_BORDER

            row_num += 1

    # 列宽
    detail_widths = [6, 12, 15, 12, 22, 70, 12, 15, 12, 50, 15, 18, 15]
    for i, w in enumerate(detail_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════════════════
    # Sheet 3: 评估规则说明（详细规则下沉到 references/evaluation-rules.md）
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("评估规则")

    # 顶部：本次抽查的运行自检结果
    rules: list[tuple[str, str]] = [("评估维度", "规则说明 / 本次自检")]
    if coverage_check is not None:
        cov_status = "✅ 通过" if coverage_check["ok"] else "⚠ 未通过"
        rules.append((
            "抽样覆盖率自检",
            f"{cov_status}：{coverage_check['message']}"
            f"（如需调整阈值用 --min-coverage，提升覆盖率用 --sample）",
        ))
    rules.extend([
        ("评估规则来源",
         "完整规则维护在 references/evaluation-rules.md（与本 skill 同目录），"
         "包含模板匹配标准、各阶段关键词、反模式黑名单、评级标准、同质化计算方法。"
         "本 sheet 仅给概要，避免代码与规则重复维护。"),
        ("抽查维度", "以老师为维度，同一老师发送的多条信息作为一组进行评估"),
        ("评级-优秀",
         "全部按模板反馈 + 1-3 点详尽描述（命中『课堂参与度/兴趣度/进步点/专注度』等关键词）+ 个性化达标"),
        ("评级-合格①", "全部按模板反馈 + 个性化达标（平均相似度 < 55%）"),
        ("评级-合格②", "全部为停课退费等特殊情况，无需按模板反馈"),
        ("评级-不合格①", "存在未按模板反馈的条目（且非特殊情况）"),
        ("评级-不合格②/③",
         "同质化高（与历史/其他学员学情总结相似度 ≥ 55%）= 未完成个性化反馈"),
        ("特殊情况识别",
         "停课/退费/请假/不续费/已流失/已转走/已续费/转介绍 开头 → 特殊情况，走合格②路径"),
        ("反模式黑名单",
         "命中即判不符合，含：课后反馈模板、续费销售话术、日报流水账、无评估的知识点复述、"
         "万能套话。详见 references/evaluation-rules.md。"),
        ("同质化计算",
         "SequenceMatcher 计算两两相似度取均值；< 55% = 个性化达标，"
         "< 35% 是优秀候选的更高门槛参考"),
    ])

    for row_idx, (col1, col2) in enumerate(rules, 1):
        cell1 = ws3.cell(row=row_idx, column=1, value=col1)
        cell2 = ws3.cell(row=row_idx, column=2, value=col2)
        if row_idx == 1:
            cell1.font = HEADER_FONT
            cell2.font = HEADER_FONT
            cell1.fill = HEADER_FILL
            cell2.fill = HEADER_FILL
        else:
            cell1.font = BOLD_FONT
            cell2.font = BODY_FONT
        cell1.alignment = WRAP_ALIGN
        cell2.alignment = WRAP_ALIGN
        cell1.border = THIN_BORDER
        cell2.border = THIN_BORDER

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 90

    wb.save(output_file)


if __name__ == "__main__":
    main()
