# -*- coding: utf-8 -*-
"""
BI 报表指标提取器
从4张服务绩效核算报表中提取关键指标
"""

import glob
from pathlib import Path
from typing import Dict, Any, Optional
import openpyxl


def find_file(directory: Path, pattern: str) -> Optional[Path]:
    """在目录中按模式查找文件，排除Excel临时文件(~$开头)。"""
    matches = [p for p in directory.glob(pattern)
               if not p.name.startswith("~$") and p.suffix == ".xlsx"]
    if not matches:
        return None
    # 返回最新修改的文件
    return max(matches, key=lambda p: p.stat().st_mtime)


def find_column_by_header(ws, header_text: str, header_row: int = 2,
                           column_index_hint: int = 0) -> Optional[int]:
    """
    在指定行按列头文字查找列号。
    支持模糊匹配，返回第一个匹配或按 hint 取第 N 个匹配。
    """
    matches = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        val = cell.value
        if val and header_text in str(val):
            matches.append(col)
    if not matches:
        return None
    idx = column_index_hint if column_index_hint < len(matches) else 0
    return matches[idx]


def find_row_by_conditions(ws, conditions: Dict[str, str],
                            start_row: int = 3, max_row: int = 50) -> Optional[int]:
    """
    按条件查找行号。conditions: {列字母: 期望值}
    例如 {"B": "海外团队", "C": "总计"}
    """
    for row in range(start_row, min(max_row, ws.max_row + 1)):
        all_match = True
        for col_letter, expected in conditions.items():
            col = openpyxl.utils.column_index_from_string(col_letter)
            val = ws.cell(row=row, column=col).value
            if val is None or str(val).strip() != expected.strip():
                all_match = False
                break
        if all_match:
            return row
    return None


def extract_metric(wb_path: Path, sheet_name: str,
                   column_header: str, row_conditions: Dict[str, str],
                   header_row: int = 2, column_index_hint: int = 0,
                   data_row_offset: int = 0) -> Any:
    """
    从Excel文件中提取单个指标值。
    data_row_offset: 找到匹配行后，向下偏移多少行取数据（用于合并单元格场景）
    """
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {wb_path}")
    ws = wb[sheet_name]

    # 定位列
    col = find_column_by_header(ws, column_header, header_row, column_index_hint)
    if col is None:
        raise ValueError(f"Column header '{column_header}' not found in sheet '{sheet_name}'")

    # 定位行
    row = find_row_by_conditions(ws, row_conditions)
    if row is None:
        raise ValueError(f"Row matching {row_conditions} not found")

    # 应用偏移
    data_row = row + data_row_offset

    # 读取值
    val = ws.cell(row=data_row, column=col).value
    wb.close()
    return val


def extract_all_metrics(bi_dir: Path, metric_sources: Dict[str, Dict]) -> Dict[str, Any]:
    """
    根据配置提取所有指标。
    返回: {指标名: 值}
    """
    results = {}
    errors = []

    for metric_name, config in metric_sources.items():
        try:
            # 查找文件
            pattern = config["file_pattern"]
            file_path = find_file(bi_dir, pattern)
            if file_path is None:
                errors.append(f"[{metric_name}] 未找到匹配文件: {pattern}")
                continue

            # 提取指标
            locate = config["locate_by"]
            hint = locate.get("column_index_hint", 0)
            header_row = locate.get("header_row", 2)
            data_row_offset = locate.get("data_row_offset", 0)

            val = extract_metric(
                wb_path=file_path,
                sheet_name=config["sheet"],
                column_header=locate["column_header"],
                row_conditions=locate["row_match"],
                header_row=header_row,
                column_index_hint=hint,
                data_row_offset=data_row_offset,
            )

            # 格式化值
            if val is not None:
                # 百分比值转为小数
                if isinstance(val, str) and "%" in val:
                    val = float(val.replace("%", "")) / 100
                results[metric_name] = val
                print(f"  ✓ [{metric_name}] = {val} (来自 {file_path.name})")
            else:
                errors.append(f"[{metric_name}] 值为空")

        except Exception as e:
            errors.append(f"[{metric_name}] {e}")

    if errors:
        print("\n⚠ 提取警告:")
        for e in errors:
            print(f"  - {e}")

    return results


def resolve_metric_for_incentive_item(item_text: str,
                                       metric_map: Dict[str, str],
                                       extracted_metrics: Dict[str, Any]) -> Optional[Any]:
    """
    根据激励完成情况的文本，匹配对应的指标值。
    支持模糊匹配。
    """
    item_text = str(item_text).strip()

    # 精确匹配
    if item_text in metric_map:
        metric_name = metric_map[item_text]
        return extracted_metrics.get(metric_name)

    # 模糊匹配
    for key, metric_name in metric_map.items():
        if key in item_text or item_text in key:
            return extracted_metrics.get(metric_name)

    return None


if __name__ == "__main__":
    from config import BI_REPORT_DIR, METRIC_SOURCES

    print("=== 提取BI报表指标 ===")
    metrics = extract_all_metrics(BI_REPORT_DIR, METRIC_SOURCES)
    print(f"\n提取完成，共 {len(metrics)} 个指标:")
    for k, v in metrics.items():
        print(f"  {k}: {v}")
