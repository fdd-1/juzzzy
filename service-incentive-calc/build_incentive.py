# -*- coding: utf-8 -*-
"""
服务绩效核算 - 主构建模块
读取参考文件 + BI报表 → 输出当月激励核算Excel
"""

import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from extract_metrics import extract_all_metrics, resolve_metric_for_incentive_item
from config import (
    BI_REPORT_DIR, REFERENCE_INCENTIVE_FILE, OUTPUT_DIR,
    METRIC_SOURCES, INCENTIVE_ITEM_METRIC_MAP, INCENTIVE_ITEM_SHEET,
    INCENTIVE_THRESHOLDS
)


def parse_incentive_items(wb) -> Dict[int, Dict[str, Any]]:
    """
    解析激励项 sheet 的结构，找出每个激励细项所在行。
    返回: {行号: {"激励方向": ..., "激励完成情况文本": ..., "数据行": ...}}
    """
    cfg = INCENTIVE_ITEM_SHEET
    ws = wb[cfg["name"]]
    header_row = cfg["header_row"]
    col_map = cfg["columns"]

    items = {}
    # 遍历激励项 sheet，找包含"激励完成情况"内容的行
    for row in range(header_row + 1, ws.max_row + 1):
        direction = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col_map["激励方向"])).value
        status_text = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col_map["激励完成情况"])).value
        total = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col_map["激励总额"])).value

        if status_text and total is not None:
            items[row] = {
                "激励方向": direction,
                "激励完成情况文本": str(status_text).strip(),
                "激励总额": total,
                "数据行": row,
            }
            # 下一行通常是实际数值行（激励完成情况下面一行）
            data_row = row + 1
            val = ws.cell(row=data_row, column=openpyxl.utils.column_index_from_string(col_map["激励完成情况"])).value
            if val is not None:
                items[row]["数据行"] = data_row

    return items


def fill_incentive_data(wb, metrics: Dict[str, Any],
                        metric_map: Dict[str, str]) -> Dict[str, Any]:
    """
    将提取的指标填入激励项 sheet 的激励完成情况列。
    同时计算并填入实际激励金额（处理合并单元格）。
    返回填充摘要。
    """
    cfg = INCENTIVE_ITEM_SHEET
    ws = wb[cfg["name"]]
    col_map = cfg["columns"]
    status_col = openpyxl.utils.column_index_from_string(col_map["激励完成情况"])
    total_col = openpyxl.utils.column_index_from_string(col_map["激励总额"])
    amount_col = openpyxl.utils.column_index_from_string(col_map["实际激励金额"])

    items = parse_incentive_items(wb)
    filled = []
    not_found = []

    # 使用配置中的阈值
    thresholds = INCENTIVE_THRESHOLDS

    for row, item in items.items():
        text = item["激励完成情况文本"]
        data_row = item.get("数据行", row + 1)
        total = item.get("激励总额")

        # 匹配指标
        val = resolve_metric_for_incentive_item(text, metric_map, metrics)

        if val is not None:
            # 填入激励完成情况到 F列 (data_row)
            cell = ws.cell(row=data_row, column=status_col)
            cell.value = val
            cell.number_format = '0.00%' if isinstance(val, float) and val < 1 else 'General'

            # 计算实际激励金额
            if total is not None and isinstance(val, (int, float)):
                # 根据指标名称确定阈值
                threshold = None
                for key, t in thresholds.items():
                    if key in text:
                        threshold = t
                        break

                if threshold is not None:
                    # 计算比例（不超过1）
                    ratio = min(val / threshold, 1.0)
                    amount = round(total * ratio, 2)

                    # J列是合并单元格，需要取消合并后写入值
                    # J列的合并范围是 J{row}:J{data_row}（row是header行，data_row是数据行）
                    amount_row = row  # 公式所在行（也是合并单元格的主单元格）

                    # 检查是否需要取消合并
                    merge_range_str = f"J{amount_row}:J{data_row}"
                    merge_ranges = [str(r) for r in ws.merged_cells.ranges]
                    if merge_range_str in merge_ranges:
                        ws.unmerge_cells(merge_range_str)

                    # 写入实际激励金额
                    ws.cell(row=amount_row, column=amount_col).value = amount

                    # 重新合并
                    ws.merge_cells(f"J{amount_row}:J{data_row}")

            filled.append({
                "激励方向": item["激励方向"],
                "文本": text,
                "值": val,
                "行": data_row,
                "激励总额": total,
            })
        else:
            not_found.append({
                "激励方向": item["激励方向"],
                "文本": text,
                "行": data_row,
            })

    return {"filled": filled, "not_found": not_found}


def update_summary_formulas(wb):
    """
    确保激励汇总 sheet 的公式指向正确的行。
    （如果激励项行号变化，可能需要调整，目前保持原公式）
    """
    # 目前参考文件的公式是相对引用，直接保留即可
    pass


def build_incentive_excel(
    bi_dir: Path = BI_REPORT_DIR,
    reference_file: Path = REFERENCE_INCENTIVE_FILE,
    output_dir: Path = OUTPUT_DIR,
    month: str = "",
) -> Path:
    """
    主构建流程。
    month: 月份标识，如 "5月"，默认取当月。
    返回输出文件路径。
    """
    if not month:
        month = f"{date.today().month}月"

    print(f"=" * 60)
    print(f"  服务绩效核算 - {month}")
    print(f"=" * 60)

    # ── 1. 提取BI报表指标 ──
    print("\n[1/4] 提取BI报表指标...")
    metrics = extract_all_metrics(bi_dir, METRIC_SOURCES)
    print(f"  成功提取 {len(metrics)} 个指标")

    # ── 2. 读取参考文件 ──
    print(f"\n[2/4] 读取参考激励文件: {reference_file.name}")
    wb = openpyxl.load_workbook(reference_file)
    print(f"  Sheets: {wb.sheetnames}")

    # ── 3. 填入指标数据 ──
    print(f"\n[3/4] 填入激励完成情况...")
    result = fill_incentive_data(wb, metrics, INCENTIVE_ITEM_METRIC_MAP)

    print(f"  ✓ 已填充 {len(result['filled'])} 项:")
    total_amount = 0
    for item in result['filled']:
        val_str = f"{item['值']:.4f}" if isinstance(item['值'], float) else str(item['值'])
        incentive = item.get("激励总额", 0)
        print(f"    - [{item['激励方向']}] {item['文本']}: {val_str} (激励总额: {incentive})")
        total_amount += incentive if incentive else 0

    print(f"  总激励金额: {total_amount}")

    if result['not_found']:
        print(f"  ⚠ 未匹配 {len(result['not_found'])} 项:")
        for item in result['not_found']:
            print(f"    - [{item['激励方向']}] {item['文本']} (行{item['行']})")

    # ── 4. 输出文件 ──
    print(f"\n[4/4] 输出当月激励核算文件...")
    month_dir = output_dir / month
    month_dir.mkdir(parents=True, exist_ok=True)
    output_file = month_dir / f"{month}服务激励_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb.save(output_file)
    wb.close()
    print(f"  ✅ 已保存: {output_file}")

    return output_file


if __name__ == "__main__":
    build_incentive_excel()
