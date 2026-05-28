"""生成脱敏样例数据。

来源：03_output/<期次>/积分汇总.xlsx 的真实结构（学员 ID / 姓名 / 课件名 / 课包 ID 全部替换为占位）
输出：_release/学情积分核算/samples/积分汇总_sample.xlsx
"""
import random
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

random.seed(42)
HERE = Path(__file__).parent.resolve()
OUT = HERE / "积分汇总_sample.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()

# Sheet 1: 汇总
ws1 = wb.active
ws1.title = "汇总"
ws1.append(["学生ID", "当前LP姓名", "当前小组", "总积分", "上课时间"])
SAMPLE_LP = ["LP_001", "LP_002", "LP_003", "LP_004"]
SAMPLE_GROUP = ["小组A", "小组B", "小组C"]
for i in range(20):
    sid = 10000000 + i  # 占位 ID
    lp = SAMPLE_LP[i % len(SAMPLE_LP)]
    grp = SAMPLE_GROUP[i % len(SAMPLE_GROUP)]
    total = random.choice([500, 1000, 1500, 2000])
    times = ", ".join(
        (datetime(2026, 5, 1) + timedelta(days=random.randint(0, 14), hours=random.randint(8, 19))).strftime("%Y-%m-%d %H:%M:%S")
        for _ in range(random.randint(1, 4))
    )
    ws1.append([sid, lp, grp, total, times])

# Sheet 2: 获得积分明细
ws2 = wb.create_sheet("获得积分明细")
ws2.append(["学生ID", "积分数量", "上课日期时间", "课件名称"])
SAMPLE_LESSON = ["示例课件_01", "示例课件_02", "示例课件_03", "示例课件_04"]
for i in range(40):
    sid = 10000000 + (i % 20)
    amt = 500
    dt = datetime(2026, 5, 1) + timedelta(days=random.randint(0, 14), hours=random.randint(8, 19))
    lesson = random.choice(SAMPLE_LESSON)
    ws2.append([sid, amt, dt, lesson])

# Sheet 3: 学情课包ID池
ws3 = wb.create_sheet("学情课包ID池")
ws3.append(["学员ID", "当前课包ID"])
for i in range(50):
    sid = 10000000 + (i % 30)
    pkg = 20000000 + i  # 占位课包 ID
    ws3.append([sid, pkg])

wb.save(OUT)
print(f"[OK] 生成脱敏样例: {OUT}")
print(f"  Sheets: {wb.sheetnames}")
