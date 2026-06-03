"""
学情积分核算 - 数据处理脚本

读取 BI 下载的两张报表，执行:
1. 从报表1(续费规划表)筛选"学情"课包 → 更新累计课包池
2. 在报表2(上课明细)中匹配课包池 → 计算3列(是否学情课包/是否符合发放条件/发放积分数量)
3. 生成输出Excel: 续费筛选数据、带标注的上课明细、获得积分明细、汇总

用法:
  python process_xueqing.py --report1 "续费规划表.xlsx" --report2 "上课明细.xlsx" --pool "历史课包池.xlsx" --output "输出.xlsx"
  python process_xueqing.py --report1 "续费规划表.xlsx" --report2 "上课明细.xlsx" --pool-sheet "25-26年5月15号学情课包ID" --output "输出.xlsx"
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, numbers

sys.stdout.reconfigure(encoding="utf-8")


def find_header_row(ws, keywords=("学员ID", "学生ID", "课包ID", "课时包ID"), max_scan=20, min_cols=15):
    """在前 max_scan 行中找到包含关键字的表头行，且至少有 min_cols 个非空列"""
    for r in range(1, max_scan + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 100))]
        non_null_count = sum(1 for v in row_vals if v is not None)
        if non_null_count < min_cols:
            continue
        row_text = "|".join(str(v or "") for v in row_vals)
        if any(kw in row_text for kw in keywords):
            return r
    return 1


def read_sheet_as_dicts(ws, header_row=None):
    """读取工作表为字典列表"""
    if header_row is None:
        header_row = find_header_row(ws)
    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        headers.append(str(h).strip() if h else f"_col{c}")
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            row[h] = ws.cell(r, c).value
        if any(v is not None for v in row.values()):
            rows.append(row)
    return headers, rows


def load_pool_from_sheet(wb, sheet_name):
    """从Excel的指定sheet加载课包池 (学员ID, 课包ID) 集合"""
    ws = wb[sheet_name]
    pairs = set()
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        pid = ws.cell(r, 2).value
        if sid is not None and pid is not None:
            pairs.add((sid, pid))
    return pairs


def load_pool_from_file(pool_path):
    """从独立Excel文件加载课包池"""
    wb = openpyxl.load_workbook(pool_path, data_only=True)
    ws = wb.active
    pairs = set()
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        pid = ws.cell(r, 2).value
        if sid is not None and pid is not None:
            pairs.add((sid, pid))
    return pairs


def filter_xueqing_from_report1(report1_path):
    """从续费规划表中筛选课包名称含'学情'的行，返回 (headers, rows, pool_pairs)"""
    wb = openpyxl.load_workbook(report1_path, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws, keywords=("学员ID", "当前课包名称", "当前课包ID"))
    headers, all_rows = read_sheet_as_dicts(ws, header_row)

    xueqing_rows = []
    pool_pairs = set()
    for row in all_rows:
        pkg_name = str(row.get("当前课包名称", "") or "")
        if "学情" in pkg_name:
            xueqing_rows.append(row)
            sid = row.get("学员ID")
            pid = row.get("当前课包ID")
            if sid is not None and pid is not None:
                pool_pairs.add((sid, pid))

    return headers, xueqing_rows, pool_pairs


def process_report2(report2_path, pool_pairs):
    """
    处理上课明细:
    - 匹配课包池 → 是否学情课包
    - 判断发放条件 → 是否符合发放条件 (4个条件: 1.在池中 2.是否预习=1 3.线上作业提交状态=已提交 4.消耗课时>=1)
    - 计算积分 → 发放积分数量
    返回 (headers, all_rows_annotated, qualified_rows)
    """
    wb = openpyxl.load_workbook(report2_path, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws, keywords=("学生ID", "课时包ID", "考勤状态", "上课日期时间"))
    headers, all_rows = read_sheet_as_dicts(ws, header_row)

    extra_cols = ["是否学情课包", "是否符合发放条件", "发放积分数量"]
    out_headers = headers + extra_cols

    annotated = []
    qualified = []

    for row in all_rows:
        sid = row.get("学生ID")
        pid = row.get("课时包ID")

        # 条件1: 是否在课包池中
        in_pool = (sid, pid) in pool_pairs
        is_xueqing = pid if in_pool else ""

        # 条件2: 是否预习 = 1
        is_preview = row.get("是否预习")
        preview_ok = (is_preview == 1 or str(is_preview).strip() == "1")

        # 条件3: 线上作业提交状态 = 已提交
        homework_status = str(row.get("线上作业提交状态", "") or "").strip()
        homework_ok = (homework_status == "已提交")

        # 条件4: 基础课时消耗课时 + 赠送课时消耗课时 >= 1
        base_consume = row.get("基础课时消耗课时") or 0
        gift_consume = row.get("赠送课时消耗课时") or 0
        try:
            base_consume = float(base_consume)
            gift_consume = float(gift_consume)
        except (ValueError, TypeError):
            base_consume = 0
            gift_consume = 0

        total_consume = base_consume + gift_consume
        consume_ok = (total_consume >= 1)

        # 综合判断: 4个条件都满足才符合发放条件
        is_qualified = 1 if (in_pool and preview_ok and homework_ok and consume_ok) else 0

        # 计算积分
        score = int(total_consume * 500) if is_qualified else 0

        row["是否学情课包"] = is_xueqing
        row["是否符合发放条件"] = is_qualified
        row["发放积分数量"] = score
        annotated.append(row)

        if is_qualified:
            qualified.append(row)

    return out_headers, annotated, qualified


def build_detail_sheet(qualified_rows):
    """构建获得积分明细数据: 学生ID, 积分数量, 上课日期时间, 课件名称"""
    detail = []
    for row in qualified_rows:
        detail.append({
            "学生ID": row.get("学生ID"),
            "积分数量": row.get("发放积分数量", 0),
            "上课日期时间": row.get("上课日期时间"),
            "课件名称": row.get("课件名称"),
        })
    return detail


def build_summary(qualified_rows, report1_rows):
    """
    构建汇总数据:
    - 每个学生: LP姓名, 小组, 总积分, 上课时间列表
    """
    student_info = {}
    for row in qualified_rows:
        sid = row.get("学生ID")
        if sid not in student_info:
            student_info[sid] = {
                "学生ID": sid,
                "当前LP姓名": row.get("当前LP姓名", ""),
                "当前小组": row.get("当前小组", ""),
                "总积分": 0,
                "上课时间": [],
            }
        student_info[sid]["总积分"] += row.get("发放积分数量", 0)
        class_time = row.get("上课日期时间")
        if class_time:
            student_info[sid]["上课时间"].append(str(class_time))

    for info in student_info.values():
        info["上课时间"] = sorted(info["上课时间"])

    return list(student_info.values())


def write_output_split(output_dir, report1_headers, xueqing_rows, report2_headers,
                       annotated_rows, detail_data, summary_data, pool_pairs):
    """生成输出到文件夹，拆分为多个Excel避免单文件过大"""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    total_score = sum(item["总积分"] for item in summary_data)

    # 文件1: 汇总+积分明细+课包池
    wb1 = openpyxl.Workbook()
    ws_summary = wb1.active
    ws_summary.title = "汇总"
    summary_headers = ["学生ID", "当前LP姓名", "当前小组", "总积分", "上课时间"]
    for c, h in enumerate(summary_headers, 1):
        ws_summary.cell(1, c, h)
        ws_summary.cell(1, c).font = Font(bold=True)
    for r, item in enumerate(summary_data, 2):
        ws_summary.cell(r, 1, item["学生ID"])
        ws_summary.cell(r, 2, item["当前LP姓名"])
        ws_summary.cell(r, 3, item["当前小组"])
        ws_summary.cell(r, 4, item["总积分"])
        ws_summary.cell(r, 5, ", ".join(item["上课时间"]))
    last_row = len(summary_data) + 2
    ws_summary.cell(last_row, 3, "总计")
    ws_summary.cell(last_row, 3).font = Font(bold=True)
    ws_summary.cell(last_row, 4, total_score)
    ws_summary.cell(last_row, 4).font = Font(bold=True)

    ws_detail = wb1.create_sheet("获得积分明细")
    detail_headers = ["学生ID", "积分数量", "上课日期时间", "课件名称"]
    for c, h in enumerate(detail_headers, 1):
        ws_detail.cell(1, c, h)
        ws_detail.cell(1, c).font = Font(bold=True)
    for r, item in enumerate(detail_data, 2):
        ws_detail.cell(r, 1, item["学生ID"])
        ws_detail.cell(r, 2, item["积分数量"])
        ws_detail.cell(r, 3, item["上课日期时间"])
        ws_detail.cell(r, 4, item["课件名称"])
    detail_last = len(detail_data) + 2
    ws_detail.cell(detail_last, 1, "总计")
    ws_detail.cell(detail_last, 1).font = Font(bold=True)
    ws_detail.cell(detail_last, 2, total_score)
    ws_detail.cell(detail_last, 2).font = Font(bold=True)

    ws_pool = wb1.create_sheet("学情课包ID池")
    ws_pool.cell(1, 1, "学员ID")
    ws_pool.cell(1, 2, "当前课包ID")
    ws_pool.cell(1, 1).font = Font(bold=True)
    ws_pool.cell(1, 2).font = Font(bold=True)
    for r, (sid, pid) in enumerate(sorted(pool_pairs), 2):
        ws_pool.cell(r, 1, sid)
        ws_pool.cell(r, 2, pid)

    f1 = output_dir / "积分汇总.xlsx"
    wb1.save(f1)

    # 文件2: 上课明细(带标注列)
    wb2 = openpyxl.Workbook()
    ws_class = wb2.active
    ws_class.title = "海外思维学员上课明细"
    for c, h in enumerate(report2_headers, 1):
        ws_class.cell(1, c, h)
        ws_class.cell(1, c).font = Font(bold=True)
    for r, row in enumerate(annotated_rows, 2):
        for c, h in enumerate(report2_headers, 1):
            ws_class.cell(r, c, row.get(h))

    f2 = output_dir / "上课明细_带标注.xlsx"
    wb2.save(f2)

    # 文件3: 续费规划表筛选
    wb3 = openpyxl.Workbook()
    ws_renew = wb3.active
    ws_renew.title = "海外思维续费规划表_学情筛选"
    for c, h in enumerate(report1_headers, 1):
        ws_renew.cell(1, c, h)
        ws_renew.cell(1, c).font = Font(bold=True)
    for r, row in enumerate(xueqing_rows, 2):
        for c, h in enumerate(report1_headers, 1):
            ws_renew.cell(r, c, row.get(h))

    f3 = output_dir / "续费规划表_学情筛选.xlsx"
    wb3.save(f3)

    print(f"\n[OK] 输出文件夹: {output_dir}")
    print(f"  1. 积分汇总.xlsx — 汇总({len(summary_data)}名学生, 总积分{total_score}) + 积分明细({len(detail_data)}条) + 课包池({len(pool_pairs)}对)")
    print(f"  2. 上课明细_带标注.xlsx — {len(annotated_rows)}行 (含是否学情课包/是否符合发放条件/发放积分数量)")
    print(f"  3. 续费规划表_学情筛选.xlsx — {len(xueqing_rows)}行")


def main():
    parser = argparse.ArgumentParser(description="学情积分核算 - 数据处理")
    parser.add_argument("--report1", required=True, help="报表1(续费规划表)Excel路径")
    parser.add_argument("--report2", required=True, help="报表2(上课明细)Excel路径")
    parser.add_argument("--pool", help="历史课包池Excel文件路径(独立文件)")
    parser.add_argument("--pool-sheet", help="从参考Excel中读取课包池的sheet名")
    parser.add_argument("--pool-source", help="包含pool-sheet的Excel文件路径(默认同report1)")
    parser.add_argument("--output", required=True, help="输出文件夹路径")
    args = parser.parse_args()

    print("=" * 60)
    print("学情积分核算 - 数据处理")
    print("=" * 60)

    # 1. 从报表1筛选学情
    print("\n[1/4] 从续费规划表筛选学情课包...")
    r1_headers, xueqing_rows, new_pool_pairs = filter_xueqing_from_report1(args.report1)
    print(f"  续费规划表中学情课包: {len(xueqing_rows)} 行")
    print(f"  本期新增课包池对: {len(new_pool_pairs)} 对")

    # 2. 加载/合并课包池
    print("\n[2/4] 加载课包池...")
    if args.pool:
        pool_pairs = load_pool_from_file(args.pool)
        print(f"  从文件加载历史池: {len(pool_pairs)} 对")
    elif args.pool_sheet:
        source = args.pool_source or args.report1
        wb_pool = openpyxl.load_workbook(source, data_only=True)
        pool_pairs = load_pool_from_sheet(wb_pool, args.pool_sheet)
        print(f"  从sheet '{args.pool_sheet}' 加载池: {len(pool_pairs)} 对")
    else:
        pool_pairs = set()
        print("  无历史池，仅使用本期数据")

    pool_pairs = pool_pairs | new_pool_pairs
    print(f"  合并后课包池总计: {len(pool_pairs)} 对")

    # 3. 处理上课明细
    print("\n[3/4] 处理上课明细，匹配课包池并计算积分...")
    r2_headers, annotated, qualified = process_report2(args.report2, pool_pairs)
    print(f"  上课明细总行: {len(annotated)}")
    print(f"  符合发放条件: {len(qualified)} 行")
    total_score = sum(r.get("发放积分数量", 0) for r in qualified)
    print(f"  总积分: {total_score}")

    # 4. 生成输出
    print("\n[4/4] 生成输出Excel...")
    detail_data = build_detail_sheet(qualified)
    summary_data = build_summary(qualified, xueqing_rows)

    write_output_split(args.output, r1_headers, xueqing_rows, r2_headers,
                       annotated, detail_data, summary_data, pool_pairs)


if __name__ == "__main__":
    main()
