"""检查报表2结构 - 非read_only模式"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from pathlib import Path
import openpyxl

path = r"C:\Users\fengjianyi\Desktop\方法积分核算\01_bi_exports\海外思维学员上课明细.xlsx"
print(f"文件大小: {Path(path).stat().st_size / 1024 / 1024:.1f} MB")

# 用非read_only模式检查前几行
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active
print(f"Sheet: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}")

# 检查合并单元格
print(f"合并单元格数: {len(ws.merged_cells.ranges)}")
if ws.merged_cells.ranges:
    for mc in list(ws.merged_cells.ranges)[:5]:
        print(f"  {mc}")

# 检查第10行
print(f"\n第10行 (非空列):")
row10_vals = [(c, ws.cell(10, c).value) for c in range(1, min(ws.max_column + 1, 80)) if ws.cell(10, c).value is not None]
print(f"  非空列数: {len(row10_vals)}")
print(f"  前20: {row10_vals[:20]}")

# 检查第11行
print(f"\n第11行 (非空列):")
row11_vals = [(c, ws.cell(11, c).value) for c in range(1, min(ws.max_column + 1, 80)) if ws.cell(11, c).value is not None]
print(f"  非空列数: {len(row11_vals)}")
print(f"  前20: {row11_vals[:20]}")
