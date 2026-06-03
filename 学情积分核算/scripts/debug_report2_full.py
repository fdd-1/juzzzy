"""检查报表2的完整结构"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

path = r"C:\Users\fengjianyi\Desktop\方法积分核算\01_bi_exports\海外思维学员上课明细.xlsx"
print(f"文件大小: {Path(path).stat().st_size / 1024 / 1024:.1f} MB")

wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Max row: {ws.max_row}")
print(f"Max col: {ws.max_column}")

# 检查第10行的所有列
print("\n第10行(前50列):")
row_idx = 0
for row in ws.iter_rows(min_row=10, max_row=10, values_only=False):
    vals = [(c.column, c.value) for c in row[:50] if c.value is not None]
    print(f"  非空列数: {len(vals)}")
    print(f"  前30列: {vals[:30]}")

# 检查第11行
print("\n第11行(前50列):")
for row in ws.iter_rows(min_row=11, max_row=11, values_only=False):
    vals = [(c.column, c.value) for c in row[:50] if c.value is not None]
    print(f"  非空列数: {len(vals)}")
    print(f"  前30列: {vals[:30]}")

wb.close()

from pathlib import Path
