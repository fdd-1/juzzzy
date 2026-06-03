"""详细检查表头"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

# 报表1 - 检查第14-15行
print("=== 报表1 第14行表头 ===")
wb1 = openpyxl.load_workbook(
    r"C:\Users\fengjianyi\Desktop\方法积分核算\01_bi_exports\海外思维续费规划表_新版_26年启用.xlsx",
    data_only=True
)
ws1 = wb1.active
row14 = [ws1.cell(14, c).value for c in range(1, ws1.max_column + 1)]
non_null_14 = [(i+1, v) for i, v in enumerate(row14) if v is not None]
print(f"  非空列数: {len(non_null_14)}")
for col, val in non_null_14[:20]:
    print(f"    C{col}: {val}")
print("  ...")
for col, val in non_null_14[20:45]:
    print(f"    C{col}: {val}")

# 检查第15行(数据行)
print("\n=== 报表1 第15行(首条数据) ===")
row15_sample = [(c, ws1.cell(15, c).value) for c in range(1, 50) if ws1.cell(15, c).value is not None]
print(f"  {row15_sample[:15]}")

# 报表2 - 检查第8-12行
print("\n=== 报表2 第8-12行 ===")
wb2 = openpyxl.load_workbook(
    r"C:\Users\fengjianyi\Desktop\方法积分核算\01_bi_exports\海外思维学员上课明细.xlsx",
    data_only=True, read_only=True
)
ws2 = wb2.active
row_idx = 0
for row in ws2.iter_rows(min_row=8, max_row=12, values_only=False):
    row_idx += 1
    actual_row = 7 + row_idx
    vals = [(c.column, c.value) for c in row if c.value is not None]
    if vals:
        print(f"  R{actual_row}: {vals[:20]}")

wb2.close()
