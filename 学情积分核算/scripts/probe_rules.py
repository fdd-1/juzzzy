import openpyxl
import sys
from collections import Counter, defaultdict

sys.stdout.reconfigure(encoding="utf-8")

xl_path = r"C:\Users\fengjianyi\Desktop\其余活动-跟进\学情课包\5.1-5.15\20260501-20260515学情积分发放明细-业务.xlsx"
wb = openpyxl.load_workbook(xl_path, data_only=True)

ws_r = wb["海外思维续费规划表_新版_26年启用"]
header_row = 15
headers = [ws_r.cell(header_row, c).value for c in range(1, ws_r.max_column + 1)]
sid_col = headers.index("学员ID") + 1
pkgid_col = headers.index("当前课包ID") + 1
name_col = headers.index("当前课包名称") + 1
sign_ym_col = headers.index("当前课包签单年月") + 1

print("续费表 行数:", ws_r.max_row, "数据行起始:", header_row + 1)
print("总数据行:", ws_r.max_row - header_row)

name_counter = Counter()
sign_ym_counter = Counter()
xueqing_pkg_ids = set()
xueqing_rows = 0
for r in range(header_row + 1, ws_r.max_row + 1):
    nm = ws_r.cell(r, name_col).value
    sym = ws_r.cell(r, sign_ym_col).value
    if nm:
        name_counter[nm] += 1
        if "学情" in nm:
            xueqing_rows += 1
            xueqing_pkg_ids.add(ws_r.cell(r, pkgid_col).value)
    sign_ym_counter[str(sym)[:7] if sym else None] += 1

print(f"\n续费表中名称含[学情]行数: {xueqing_rows}, 唯一课包ID: {len(xueqing_pkg_ids)}")
print("前10课包名称:")
for n, c in name_counter.most_common(10):
    print(f"  {c} -> {n}")
print("签单年月分布(前15):", sign_ym_counter.most_common(15))

# 课包池
ws_pkg = wb["25-26年5月15号学情课包ID"]
pkg_set = set()
for r in range(2, ws_pkg.max_row + 1):
    pkg_set.add(ws_pkg.cell(r, 2).value)

inter = xueqing_pkg_ids & pkg_set
diff_in_renew_only = xueqing_pkg_ids - pkg_set
diff_in_pool_only = pkg_set - xueqing_pkg_ids
print(f"\n续费表[学情]课包 ∩ 5.15池: {len(inter)} / 续费表={len(xueqing_pkg_ids)} / 池子={len(pkg_set)}")
print(f"仅续费表有: {len(diff_in_renew_only)}")
print(f"仅池子有: {len(diff_in_pool_only)}")

pkg2name = {}
pkg2signym = {}
for r in range(header_row + 1, ws_r.max_row + 1):
    pid = ws_r.cell(r, pkgid_col).value
    pkg2name[pid] = ws_r.cell(r, name_col).value
    pkg2signym[pid] = ws_r.cell(r, sign_ym_col).value

print("\n仅池子有的课包，对应续费表名称(前10):")
for pid in list(diff_in_pool_only)[:10]:
    print(f"  {pid} -> {pkg2name.get(pid, '<不在续费表>')}")

print("\n仅续费表有的[学情]课包，对应名称+签单年月(前10):")
for pid in list(diff_in_renew_only)[:10]:
    print(f"  {pid} -> {pkg2name.get(pid)} | 签单年月={pkg2signym.get(pid)}")

# 上课明细
ws_c = wb["海外思维学员上课明细"]
pkg_score = defaultdict(set)
for r in range(11, ws_c.max_row + 1):
    pkg_score[ws_c.cell(r, 16).value].add(ws_c.cell(r, 74).value)
pkg_2000 = {p for p, s in pkg_score.items() if 2000 in s}

print("\n2000积分课包来自续费表的名称+签单年月:")
for pid in pkg_2000:
    print(f"  {pid} -> {pkg2name.get(pid, '<不在续费表>')} | 签单年月={pkg2signym.get(pid)}")
