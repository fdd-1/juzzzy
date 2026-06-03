import openpyxl
import sys
from collections import Counter

sys.stdout.reconfigure(encoding="utf-8")

xl_path = r"C:\Users\fengjianyi\Desktop\其余活动-跟进\学情课包\5.1-5.15\20260501-20260515学情积分发放明细-业务.xlsx"
wb = openpyxl.load_workbook(xl_path, data_only=True)
ws = wb["海外思维学员上课明细"]

headers = [ws.cell(10, c).value for c in range(1, ws.max_column + 1)]
ci = {h: i + 1 for i, h in enumerate(headers) if h}

print(f"基础课时消耗课时 列={ci['基础课时消耗课时']}")
print(f"赠送课时消耗课时 列={ci['赠送课时消耗课时']}")

mismatches = []
match = 0
total = 0
score_consume_dist = Counter()
for r in range(11, ws.max_row + 1):
    base = ws.cell(r, ci["基础课时消耗课时"]).value or 0
    gift = ws.cell(r, ci["赠送课时消耗课时"]).value or 0
    consume = base + gift
    score = ws.cell(r, ci["发放积分数量"]).value
    expected = consume * 500
    total += 1
    score_consume_dist[(consume, score)] += 1
    if score == expected:
        match += 1
    elif len(mismatches) < 5:
        mismatches.append({
            "row": r,
            "学生ID": ws.cell(r, ci["学生ID"]).value,
            "基础消耗": base,
            "赠送消耗": gift,
            "总消耗": consume,
            "实发积分": score,
            "预期(消耗*500)": expected,
            "考勤状态": ws.cell(r, ci["考勤状态"]).value,
        })

print(f"\n积分 = (基础消耗+赠送消耗)*500 验证: {match}/{total} 匹配")
print(f"消耗课时 vs 积分 二维分布: {dict(score_consume_dist)}")
if mismatches:
    print("\n不匹配样本:")
    for m in mismatches:
        print(f"  {m}")
