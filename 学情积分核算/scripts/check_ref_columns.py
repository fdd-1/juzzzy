"""检查参考Excel中是否学情课包列的实际值"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

ref_path = r"C:\Users\fengjianyi\Desktop\其余活动-跟进\学情课包\5.1-5.15\20260501-20260515学情积分发放明细-业务.xlsx"
wb = openpyxl.load_workbook(ref_path, data_only=True)
ws = wb["海外思维学员上课明细"]

# 找到列索引
headers = [ws.cell(10, c).value for c in range(1, ws.max_column + 1)]
ci = {h: i + 1 for i, h in enumerate(headers) if h}

print("前10行的是否学情课包列值:")
for r in range(11, 21):
    val = ws.cell(r, ci["是否学情课包"]).value
    sid = ws.cell(r, ci["学生ID"]).value
    pid = ws.cell(r, ci["课时包ID"]).value
    print(f"  行{r}: 学生ID={sid}, 课时包ID={pid}, 是否学情课包={val} (type={type(val).__name__})")
