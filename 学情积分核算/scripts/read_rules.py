import openpyxl
import sys

sys.stdout.reconfigure(encoding="utf-8")

xl_path = r"C:\Users\fengjianyi\Desktop\其余活动-跟进\学情课包\5.1-5.15\20260501-20260515学情积分发放明细-业务.xlsx"
wb = openpyxl.load_workbook(xl_path, data_only=True)
ws = wb["规则"]

for r in range(1, ws.max_row + 1):
    row_data = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    non_null = [x for x in row_data if x is not None]
    if non_null:
        print(f"R{r}: {non_null}")