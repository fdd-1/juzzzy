"""检查报表1的第11-15行"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

path = r"C:\Users\fengjianyi\Desktop\方法积分核算\01_bi_exports\海外思维续费规划表_新版_26年启用.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

for r in range(8, 16):
    vals = [(c, ws.cell(r, c).value) for c in range(1, 50) if ws.cell(r, c).value is not None]
    if vals:
        print(f"R{r} ({len(vals)} non-null): {vals[:15]}")
    else:
        print(f"R{r}: (empty)")
