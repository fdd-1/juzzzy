"""检查下载文件的表头结构"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

# 报表1
print("=" * 60)
print("报表1: 续费规划表")
print("=" * 60)
wb1 = openpyxl.load_workbook(
    r"C:\Users\fengjianyi\Desktop\方法积分核算\01_bi_exports\海外思维续费规划表_新版_26年启用.xlsx",
    data_only=True
)
print(f"Sheets: {wb1.sheetnames}")
ws1 = wb1.active
print(f"Active sheet: {ws1.title}, rows={ws1.max_row}, cols={ws1.max_column}")

# 打印前5行
for r in range(1, 6):
    vals = [ws1.cell(r, c).value for c in range(1, min(ws1.max_column + 1, 50))]
    non_null = [(i+1, v) for i, v in enumerate(vals) if v is not None]
    if non_null:
        print(f"  R{r}: {non_null[:10]}")

# 找学情
pkg_name_col = None
for c in range(1, ws1.max_column + 1):
    for r in range(1, 20):
        v = ws1.cell(r, c).value
        if v and "课包名称" in str(v):
            pkg_name_col = c
            print(f"\n  '课包名称' found at R{r}C{c}: {v}")
            break
    if pkg_name_col:
        break

if pkg_name_col:
    xueqing_count = 0
    for r in range(1, ws1.max_row + 1):
        v = ws1.cell(r, pkg_name_col).value
        if v and "学情" in str(v):
            xueqing_count += 1
    print(f"  含'学情'的行数: {xueqing_count}")

print("\n" + "=" * 60)
print("报表2: 上课明细")
print("=" * 60)
wb2 = openpyxl.load_workbook(
    r"C:\Users\fengjianyi\Desktop\方法积分核算\01_bi_exports\海外思维学员上课明细.xlsx",
    data_only=True, read_only=True
)
print(f"Sheets: {wb2.sheetnames}")
ws2 = wb2.active
print(f"Active sheet: {ws2.title}")

# 打印前3行
row_count = 0
for row in ws2.iter_rows(min_row=1, max_row=3, values_only=False):
    row_count += 1
    vals = [(c.column, c.value) for c in row if c.value is not None]
    print(f"  R{row_count}: {vals[:15]}")

# 找学生ID和课时包ID列
headers_row = None
for row in ws2.iter_rows(min_row=1, max_row=15, values_only=False):
    vals = [c.value for c in row]
    text = "|".join(str(v) for v in vals if v)
    if "学生ID" in text or "课时包ID" in text:
        headers_row = row[0].row
        print(f"\n  表头行: R{headers_row}")
        header_vals = [(c.column, c.value) for c in row if c.value is not None]
        print(f"  表头(前20): {header_vals[:20]}")
        break

wb2.close()
