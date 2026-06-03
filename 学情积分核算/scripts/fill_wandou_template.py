#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""把 <output-dir>/积分汇总.xlsx 的「获得积分明细」写入「发放豌豆币文档填写模板」中：
   - 第 1 列：学员ID
   - 第 3 列：获得积分
   - 其他列从模板示例行下拉复制

输出文件名带当天日期：发放豌豆币文档填写模板_YYYYMMDD.xlsx

用法:
  python fill_wandou_template.py --output-dir <期次目录>
  python fill_wandou_template.py --src 积分汇总.xlsx --template 发放豌豆币文档填写模板.xlsx --dest 发放豌豆币文档填写模板_20260528.xlsx
"""
import argparse
import io
import shutil
import sys
from datetime import date
from pathlib import Path

import openpyxl

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

BASE_DIR = Path(r"c:\Users\fengjianyi\Desktop\学情积分核算")
DEFAULT_OUTPUT_ROOT = BASE_DIR / "03_output"
DEFAULT_TEMPLATE = DEFAULT_OUTPUT_ROOT / "发放豌豆币文档填写模板.xlsx"


def find_col(hdr, *names):
    for n in names:
        if n in hdr:
            return hdr.index(n) + 1
    raise ValueError(f"未找到列: {names}, 实际表头: {hdr}")


def fill_template(src_xlsx: Path, template: Path, dest: Path) -> int:
    """从 src_xlsx 的「获得积分明细」生成 dest（基于 template 拷贝）。返回写入行数。"""
    print(f"[STEP 1] 复制模板 -> {dest}")
    shutil.copyfile(template, dest)

    print(f"[STEP 2] 读取源 {src_xlsx} sheet=获得积分明细")
    src_wb = openpyxl.load_workbook(src_xlsx, data_only=True)
    if "获得积分明细" not in src_wb.sheetnames:
        raise SystemExit(f"[ERROR] 没有「获得积分明细」sheet, 现有: {src_wb.sheetnames}")
    src_ws = src_wb["获得积分明细"]

    hdr = [src_ws.cell(1, c).value for c in range(1, src_ws.max_column + 1)]
    print(f"  源表头: {hdr}")
    col_id = find_col(hdr, "学生ID", "学员ID")
    col_amt = find_col(hdr, "积分数量", "获得积分")

    rows = []
    for r in range(2, src_ws.max_row + 1):
        sid = src_ws.cell(r, col_id).value
        amt = src_ws.cell(r, col_amt).value
        if sid is None or amt is None:
            continue
        if isinstance(sid, str) and ("总计" in sid or "合计" in sid):
            continue
        rows.append((sid, amt))
    print(f"  有效数据行: {len(rows)}")

    print(f"[STEP 3] 写入 {dest}")
    dst_wb = openpyxl.load_workbook(dest)
    dst_ws = dst_wb["Sheet1"]
    sample = {c: dst_ws.cell(2, c).value for c in range(1, dst_ws.max_column + 1)}
    print(f"  示例行: {sample}")

    for i, (sid, amt) in enumerate(rows):
        r = 2 + i
        for c in range(1, dst_ws.max_column + 1):
            if c == 1:
                dst_ws.cell(r, c).value = sid
            elif c == 3:
                dst_ws.cell(r, c).value = amt
            else:
                dst_ws.cell(r, c).value = sample[c]

    print(f"[STEP 4] 总写入 {len(rows)} 行")
    dst_wb.save(dest)
    print(f"[OK] 完成: {dest}")
    return len(rows)


def main() -> int:
    ap = argparse.ArgumentParser(description="生成发放豌豆币文档填写模板")
    ap.add_argument("--output-dir", help="期次目录（含 积分汇总.xlsx），输出会写到该目录下")
    ap.add_argument("--src", help="积分汇总.xlsx 路径（覆盖 --output-dir 推导）")
    ap.add_argument("--template", help="原始空模板路径（默认 03_output/发放豌豆币文档填写模板.xlsx）")
    ap.add_argument("--dest", help="目标文件路径（默认 <output-dir>/发放豌豆币文档填写模板_YYYYMMDD.xlsx）")
    ap.add_argument("--date", help="文件名日期后缀 YYYYMMDD（默认今天）")
    args = ap.parse_args()

    today = (args.date or date.today().strftime("%Y%m%d")).replace("-", "")

    template = Path(args.template) if args.template else DEFAULT_TEMPLATE
    if not template.exists():
        print(f"[ERROR] 找不到原始模板: {template}")
        return 1

    if args.src:
        src = Path(args.src)
        out_dir = Path(args.dest).parent if args.dest else src.parent
    elif args.output_dir:
        out_dir = Path(args.output_dir)
        src = out_dir / "积分汇总.xlsx"
    else:
        print("[ERROR] 至少指定 --output-dir 或 --src")
        return 1

    if not src.exists():
        print(f"[ERROR] 找不到积分汇总: {src}")
        return 1

    dest = Path(args.dest) if args.dest else out_dir / f"发放豌豆币文档填写模板_{today}.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)

    fill_template(src, template, dest)
    # 输出最终路径，供上层脚本捕获
    print(f"[DEST] {dest}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
