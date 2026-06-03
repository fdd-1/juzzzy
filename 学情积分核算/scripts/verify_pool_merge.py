import openpyxl
import sys

sys.stdout.reconfigure(encoding="utf-8")

xl_path = r"C:\Users\fengjianyi\Desktop\其余活动-跟进\学情课包\5.1-5.15\20260501-20260515学情积分发放明细-业务.xlsx"
wb = openpyxl.load_workbook(xl_path, data_only=True)

# 旧池子 表头
ws_old = wb["2025.11-2026.3.15课包ID"]
print("旧池子表头:", [ws_old.cell(1, c).value for c in range(1, ws_old.max_column + 1)])
print("旧池子前5行:")
for r in range(2, 7):
    print(f"  {[ws_old.cell(r, c).value for c in range(1, ws_old.max_column + 1)]}")

# 新池子 表头
ws_new = wb["25-26年5月15号学情课包ID"]
print("\n新池子表头:", [ws_new.cell(1, c).value for c in range(1, ws_new.max_column + 1)])
print("新池子前5行:")
for r in range(2, 7):
    print(f"  {[ws_new.cell(r, c).value for c in range(1, ws_new.max_column + 1)]}")

# 旧池子 (学员ID, 课包ID) pairs
old_pairs = set()
for r in range(2, ws_old.max_row + 1):
    old_pairs.add((ws_old.cell(r, 1).value, ws_old.cell(r, 2).value))

new_pairs = set()
for r in range(2, ws_new.max_row + 1):
    new_pairs.add((ws_new.cell(r, 1).value, ws_new.cell(r, 2).value))

print(f"\n旧池子 (sid,pkg) 对数: {len(old_pairs)}")
print(f"新池子 (sid,pkg) 对数: {len(new_pairs)}")
print(f"旧 ⊂ 新: {old_pairs.issubset(new_pairs)}")
print(f"新增 (sid,pkg) 对: {len(new_pairs - old_pairs)}")

# 续费规划表 (5月份只有199条 - 当期新增)
ws_r = wb["海外思维续费规划表_新版_26年启用"]
header_row = 15
headers = [ws_r.cell(header_row, c).value for c in range(1, ws_r.max_column + 1)]
sid_col = headers.index("学员ID") + 1
pkgid_col = headers.index("当前课包ID") + 1
name_col = headers.index("当前课包名称") + 1

current_xueqing_pairs = set()
for r in range(header_row + 1, ws_r.max_row + 1):
    nm = ws_r.cell(r, name_col).value
    if nm and "学情" in nm:
        current_xueqing_pairs.add((ws_r.cell(r, sid_col).value, ws_r.cell(r, pkgid_col).value))

print(f"\n本期续费表[学情]课包对数: {len(current_xueqing_pairs)}")
print(f"旧 ∪ 本期 = {len(old_pairs | current_xueqing_pairs)}, vs 新池子 {len(new_pairs)}")

# 新池子 - (旧 ∪ 本期)
extra = new_pairs - (old_pairs | current_xueqing_pairs)
print(f"新池子额外的对(不在旧也不在本期): {len(extra)}")
if extra:
    print("额外对样本(前5):")
    for p in list(extra)[:5]:
        print(f"  {p}")
