import openpyxl
import sys
from collections import Counter, defaultdict

sys.stdout.reconfigure(encoding="utf-8")

xl_path = r"C:\Users\fengjianyi\Desktop\其余活动-跟进\学情课包\5.1-5.15\20260501-20260515学情积分发放明细-业务.xlsx"
wb = openpyxl.load_workbook(xl_path, data_only=True)

ws = wb["海外思维学员上课明细"]
# 表头在第10行
headers = [ws.cell(10, c).value for c in range(1, ws.max_column + 1)]

# 找关键列索引
col_idx = {}
for i, h in enumerate(headers):
    if h:
        col_idx[h] = i + 1

print("关键列索引:")
for key in ["学生ID", "课时包ID", "考勤状态", "上课状态", "课节类型", "学生班级身份",
            "是否寒暑班", "课件名称", "上课日期时间", "是否学情课包", "是否符合发放条件", "发放积分数量"]:
    print(f"  {key}: 列{col_idx.get(key)}")

# 对比 500 vs 2000 行的各字段分布
features_500 = defaultdict(Counter)
features_2000 = defaultdict(Counter)
check_cols = ["考勤状态", "上课状态", "课节类型", "学生班级身份", "是否寒暑班"]

for r in range(11, ws.max_row + 1):
    score = ws.cell(r, col_idx["发放积分数量"]).value
    target = features_2000 if score == 2000 else features_500
    for col_name in check_cols:
        val = ws.cell(r, col_idx[col_name]).value
        target[col_name][val] += 1

print("\n=== 2000积分行 字段分布 ===")
for col_name in check_cols:
    print(f"  {col_name}: {dict(features_2000[col_name])}")

print("\n=== 500积分行 字段分布 ===")
for col_name in check_cols:
    print(f"  {col_name}: {dict(features_500[col_name])}")

# 看每个学生的上课次数 vs 积分
student_rows = defaultdict(list)
for r in range(11, ws.max_row + 1):
    sid = ws.cell(r, col_idx["学生ID"]).value
    score = ws.cell(r, col_idx["发放积分数量"]).value
    student_rows[sid].append(score)

# 2000积分学生的总行数
print("\n=== 2000积分学生的行数分布 ===")
for sid, scores in student_rows.items():
    if 2000 in scores:
        print(f"  学生{sid}: 总行数={len(scores)}, 积分列表={scores}")
