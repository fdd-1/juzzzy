"""
验证 process_xueqing.py 的逻辑是否与参考文档一致。
使用参考Excel中的数据模拟完整流程，对比结果。
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl

ref_path = r"C:\Users\fengjianyi\Desktop\其余活动-跟进\学情课包\5.1-5.15\20260501-20260515学情积分发放明细-业务.xlsx"
wb = openpyxl.load_workbook(ref_path, data_only=True)

# 1. 加载参考课包池
ws_pool = wb["25-26年5月15号学情课包ID"]
pool_pairs = set()
for r in range(2, ws_pool.max_row + 1):
    sid = ws_pool.cell(r, 1).value
    pid = ws_pool.cell(r, 2).value
    if sid is not None and pid is not None:
        pool_pairs.add((sid, pid))
print(f"课包池: {len(pool_pairs)} 对")

# 2. 处理上课明细
ws = wb["海外思维学员上课明细"]
headers = [ws.cell(10, c).value for c in range(1, ws.max_column + 1)]
ci = {h: i + 1 for i, h in enumerate(headers) if h}

total = 0
qualified_count = 0
total_score = 0
mismatches = []

for r in range(11, ws.max_row + 1):
    sid = ws.cell(r, ci["学生ID"]).value
    pid = ws.cell(r, ci["课时包ID"]).value
    base = ws.cell(r, ci["基础课时消耗课时"]).value or 0
    gift = ws.cell(r, ci["赠送课时消耗课时"]).value or 0

    in_pool = (sid, pid) in pool_pairs
    is_xueqing = pid if in_pool else ""
    consume = float(base) + float(gift)
    score = int(consume * 500)
    is_qualified = 1 if (in_pool and score > 0) else 0
    final_score = score if is_qualified else 0

    # 对比参考值
    ref_xueqing = ws.cell(r, ci["是否学情课包"]).value
    ref_qualified = ws.cell(r, ci["是否符合发放条件"]).value
    ref_score = ws.cell(r, ci["发放积分数量"]).value

    total += 1
    if is_qualified:
        qualified_count += 1
        total_score += final_score

    if (is_xueqing != ref_xueqing or is_qualified != ref_qualified or final_score != ref_score):
        if len(mismatches) < 10:
            mismatches.append({
                "row": r, "sid": sid,
                "calc": (is_xueqing, is_qualified, final_score),
                "ref": (ref_xueqing, ref_qualified, ref_score),
            })

print(f"\n总行数: {total}")
print(f"符合发放条件: {qualified_count}")
print(f"总积分: {total_score}")
print(f"不匹配行数: {len(mismatches)}")
if mismatches:
    print("\n不匹配样本:")
    for m in mismatches:
        print(f"  行{m['row']} 学生{m['sid']}: 计算={m['calc']} vs 参考={m['ref']}")
else:
    print("\n[OK] 所有行计算结果与参考完全一致!")

# 3. 对比获得积分明细
ws_detail = wb["获得积分明细"]
ref_detail_count = ws_detail.max_row - 1  # 减去表头
# 最后一行可能是总计
last_val = ws_detail.cell(ws_detail.max_row, 1).value
if last_val == "总计" or (isinstance(last_val, str) and "总" in str(last_val)):
    ref_detail_count -= 1

print(f"\n获得积分明细: 计算={qualified_count} vs 参考={ref_detail_count}")

# 4. 对比汇总
ws_summary = wb["汇总"]
ref_total_in_summary = ws_summary.cell(1, 4).value  # D1 是总积分
print(f"汇总总积分: 计算={total_score} vs 参考={ref_total_in_summary}")
