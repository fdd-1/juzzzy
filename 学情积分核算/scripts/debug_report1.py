"""Debug: 模拟 process_xueqing.py 对报表1的处理"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

path = r"C:\Users\fengjianyi\Desktop\方法积分核算\01_bi_exports\海外思维续费规划表_新版_26年启用.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

# 模拟 find_header_row
keywords = ("学员ID", "当前课包名称", "当前课包ID")
header_row = None
for r in range(1, 21):
    row_vals = [str(ws.cell(r, c).value or "") for c in range(1, min(ws.max_column + 1, 100))]
    row_text = "|".join(row_vals)
    if any(kw in row_text for kw in keywords):
        header_row = r
        print(f"找到表头行: R{r}")
        break

if not header_row:
    print("未找到表头行!")
    # 打印前20行的关键内容
    for r in range(1, 21):
        row_vals = [str(ws.cell(r, c).value or "") for c in range(1, 10)]
        print(f"  R{r}: {row_vals}")
    sys.exit(1)

# 读取表头
headers = []
for c in range(1, ws.max_column + 1):
    h = ws.cell(header_row, c).value
    headers.append(str(h).strip() if h else f"_col{c}")

print(f"表头列数: {len(headers)}")
print(f"'当前课包名称' in headers: {'当前课包名称' in headers}")
if '当前课包名称' in headers:
    idx = headers.index('当前课包名称')
    print(f"  位置: {idx} (col {idx+1})")

# 读取数据并筛选
xueqing_count = 0
total_rows = 0
sample_pkg_names = []
for r in range(header_row + 1, ws.max_row + 1):
    row = {}
    for c, h in enumerate(headers, 1):
        row[h] = ws.cell(r, c).value
    if any(v is not None for v in row.values()):
        total_rows += 1
        pkg_name = str(row.get("当前课包名称", "") or "")
        if "学情" in pkg_name:
            xueqing_count += 1
        if total_rows <= 5:
            sample_pkg_names.append(row.get("当前课包名称"))

print(f"\n数据行数: {total_rows}")
print(f"学情行数: {xueqing_count}")
print(f"前5行课包名称: {sample_pkg_names}")

# 检查学员ID和课包ID
if xueqing_count > 0:
    sid_col = headers.index("学员ID") if "学员ID" in headers else None
    pid_col = headers.index("当前课包ID") if "当前课包ID" in headers else None
    print(f"学员ID列: {sid_col}, 当前课包ID列: {pid_col}")
