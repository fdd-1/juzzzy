import openpyxl
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")

xl_path = r"C:\Users\fengjianyi\Desktop\其余活动-跟进\学情课包\5.1-5.15\20260501-20260515学情积分发放明细-业务.xlsx"
wb = openpyxl.load_workbook(xl_path, data_only=True)

# 旧池子
ws_old = wb["2025.11-2026.3.15课包ID"]
old_set = set()
for r in range(2, ws_old.max_row + 1):
    old_set.add(ws_old.cell(r, 2).value)

# 新池子
ws_new = wb["25-26年5月15号学情课包ID"]
new_set = set()
for r in range(2, ws_new.max_row + 1):
    new_set.add(ws_new.cell(r, 2).value)

# 上课明细
ws_c = wb["海外思维学员上课明细"]
pkg_score = defaultdict(set)
for r in range(11, ws_c.max_row + 1):
    pkg_score[ws_c.cell(r, 16).value].add(ws_c.cell(r, 74).value)
pkg_2000 = {p for p, s in pkg_score.items() if 2000 in s}
pkg_500 = {p for p, s in pkg_score.items() if 500 in s and 2000 not in s}

# 2000积分课包 vs 旧池子
in_old = pkg_2000 & old_set
not_in_old = pkg_2000 - old_set
print(f"2000积分课包({len(pkg_2000)}) 在旧池子: {len(in_old)}, 不在旧池子: {len(not_in_old)}")
print(f"500积分课包({len(pkg_500)}) 在旧池子: {len(pkg_500 & old_set)}, 不在旧池子: {len(pkg_500 - old_set)}")

# 旧池子 vs 新池子
print(f"\n旧池子大小: {len(old_set)}, 新池子大小: {len(new_set)}")
print(f"旧 & 新: {len(old_set & new_set)}")
print(f"仅旧: {len(old_set - new_set)}, 仅新: {len(new_set - old_set)}")

# 上课明细中命中旧池子的课包
old_in_data = set()
for r in range(11, ws_c.max_row + 1):
    pid = ws_c.cell(r, 16).value
    if pid in old_set:
        old_in_data.add(pid)
print(f"\n上课明细中命中旧池子的课包数: {len(old_in_data)}")

# 旧池子课包积分分布
old_scores = defaultdict(int)
new_scores = defaultdict(int)
for r in range(11, ws_c.max_row + 1):
    pid = ws_c.cell(r, 16).value
    score = ws_c.cell(r, 74).value
    if pid in old_set:
        old_scores[score] += 1
    else:
        new_scores[score] += 1
print(f"旧池子课包积分分布: {dict(old_scores)}")
print(f"非旧池子课包积分分布: {dict(new_scores)}")

# 看看 not_in_old 的2000课包是否在新池子里
print(f"\n2000积分但不在旧池子的课包: {not_in_old}")
for pid in not_in_old:
    print(f"  {pid} 在新池子: {pid in new_set}")
