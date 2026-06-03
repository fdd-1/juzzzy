import openpyxl
import sys

sys.stdout.reconfigure(encoding="utf-8")

xl_path = r"C:\Users\fengjianyi\Desktop\其余活动-跟进\学情课包\5.1-5.15\20260501-20260515学情积分发放明细-业务.xlsx"
wb = openpyxl.load_workbook(xl_path, data_only=True)

ws_new = wb["25-26年5月15号学情课包ID"]
new_pairs = set()
new_pkg_ids = set()
for r in range(2, ws_new.max_row + 1):
    sid = ws_new.cell(r, 1).value
    pid = ws_new.cell(r, 2).value
    new_pairs.add((sid, pid))
    new_pkg_ids.add(pid)

ws = wb["海外思维学员上课明细"]
headers = [ws.cell(10, c).value for c in range(1, ws.max_column + 1)]
ci = {h: i + 1 for i, h in enumerate(headers) if h}

pair_match = 0
pkg_only_match = 0
no_match = 0
total = 0
mismatch_samples = []
for r in range(11, ws.max_row + 1):
    sid = ws.cell(r, ci["学生ID"]).value
    pid = ws.cell(r, ci["课时包ID"]).value
    total += 1
    if (sid, pid) in new_pairs:
        pair_match += 1
    elif pid in new_pkg_ids:
        pkg_only_match += 1
        if len(mismatch_samples) < 3:
            mismatch_samples.append((r, sid, pid))
    else:
        no_match += 1

print(f"上课明细总行: {total}")
print(f"  按 (学生ID, 课包ID) 在池子里: {pair_match}")
print(f"  仅 课包ID 在池子里(学生ID不匹配): {pkg_only_match}")
print(f"  完全不在池子: {no_match}")
if mismatch_samples:
    print("  仅课包匹配样本:")
    for s in mismatch_samples:
        print(f"    {s}")
