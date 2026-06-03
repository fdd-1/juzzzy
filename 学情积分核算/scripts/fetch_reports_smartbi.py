#!/usr/bin/env python3
"""从 BI 系统下载学情积分核算所需的两张报表（smartbi-data-cli 方式）。

使用 SmartBI browser export 方法，通过 Playwright 打开 SIMPLE_REPORT，
设置筛选条件后导出 Excel。

报表1: 海外思维续费规划表_新版_26年启用
  report_id: I2c928087019b236723675f9c019b353f6027505b
  筛选: 当前课包签单年月开始/结束, 当前课包签单时间开始/结束

报表2: 海外思维学员上课明细
  report_id: I2c9280870198767976798e4f0198889e7cc27654
  筛选: 开始日期, 结束日期

用法:
  python fetch_reports_smartbi.py --start 2026-05-01 --end 2026-05-15

环境变量:
  SMARTBI_USERNAME  BI 账号
  SMARTBI_PASSWORD  BI 密码
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))

from smartbi_browser_export import export_simple_report, SmartbiBrowserExportError

CONFIG_PATH = PROJECT_ROOT / "configs" / "smartbi_simple_report_tasks.json"
OUTPUT_DIR = PROJECT_ROOT / "01_bi_exports"

USERNAME = os.environ.get("SMARTBI_USERNAME", "76218")
PASSWORD = os.environ.get("SMARTBI_PASSWORD", "123456")


def load_config() -> dict:
    with CONFIG_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def build_filters_for_task(task: dict, start_date: str, end_date: str) -> list[list[str]]:
    """根据 task config 的 date_mapping 构建 [alias, value, displayValue] 列表。"""
    mapping = task.get("filters", {}).get("date_mapping", {})
    filters: list[list[str]] = []
    for field_name, role in mapping.items():
        if role == "start_date":
            filters.append([field_name, start_date, start_date])
        elif role == "end_date":
            filters.append([field_name, end_date, end_date])
    return filters


async def fetch_one_report(
    task_name: str,
    task: dict,
    start_date: str,
    end_date: str,
    output_dir: Path,
    headless: bool = True,
) -> dict[str, Any]:
    """下载单张报表，返回结果 dict。

    若 task 配置了 split_days，则自动按 N 天分段下载，再合并为单个 Excel
    （用于上课明细这类9w+行、浏览器JS堆易OOM的大报表）。
    """
    report = task["report"]
    report_id = report["id"]
    filename_template = task.get("output", {}).get("filename", f"{task_name}.xlsx")
    # 在文件名后加上时间段后缀，避免不同时间段的下载相互覆盖
    start_short = start_date.replace("-", "")
    end_short = end_date.replace("-", "")
    stem, ext = filename_template.rsplit(".", 1) if "." in filename_template else (filename_template, "xlsx")
    filename = f"{stem}_{start_short}-{end_short}.{ext}"
    max_rows = task.get("max_rows", 200000)
    split_days = task.get("split_days", 0)  # 0 = 不分段
    final_output = output_dir / filename

    print(f"  报表: {report.get('path', task_name)}")
    print(f"  report_id: {report_id}")

    if split_days and split_days > 0:
        # 分段下载并合并
        segments = build_date_segments(start_date, end_date, split_days)
        print(f"  分段: {len(segments)} 段（每段 {split_days} 天）")
        segment_files = []
        total_bytes = 0
        t0 = time.time()
        for i, (seg_start, seg_end) in enumerate(segments, 1):
            seg_filename = f"{stem}_{seg_start.replace('-','')}-{seg_end.replace('-','')}_part{i}.{ext}"
            seg_path = output_dir / seg_filename
            # 已存在且非空则复用，避免重复下载
            if seg_path.exists() and seg_path.stat().st_size > 1000:
                print(f"\n  [段{i}/{len(segments)}] {seg_start} ~ {seg_end} (已存在 {seg_path.stat().st_size} bytes，复用)")
                segment_files.append(seg_path)
                total_bytes += seg_path.stat().st_size
                continue
            seg_filters = build_filters_for_task(task, seg_start, seg_end)
            print(f"\n  [段{i}/{len(segments)}] {seg_start} ~ {seg_end}")
            print(f"    筛选: {seg_filters}")
            print(f"    输出: {seg_path}")
            seg_result = await export_simple_report(
                username=USERNAME,
                password=PASSWORD,
                report_id=report_id,
                output_path=seg_path,
                max_rows=max_rows,
                browser_channel="chrome",
                headless=headless,
                filters=seg_filters,
            )
            seg_bytes = seg_result.get("bytes", 0)
            total_bytes += seg_bytes
            segment_files.append(seg_path)
            print(f"    完成: {seg_bytes} bytes")
        # 合并段文件
        print(f"\n  合并 {len(segment_files)} 段为 {final_output} ...")
        merge_xlsx_files(segment_files, final_output)
        merged_size = final_output.stat().st_size
        # 验证合并是否成功（合并文件应该接近段文件总大小）
        if merged_size < total_bytes * 0.3:
            raise SmartbiBrowserExportError(
                f"合并后文件异常: 段文件总计 {total_bytes} bytes，"
                f"合并后只有 {merged_size} bytes。段文件已保留以便重试: "
                f"{[str(f) for f in segment_files]}"
            )
        # 合并成功，清理段文件
        for f in segment_files:
            try:
                f.unlink()
            except Exception:
                pass
        elapsed = time.time() - t0
        print(f"  完成! {merged_size} bytes (合并自{len(segments)}段), {elapsed:.1f}s")
        return {
            "output": str(final_output),
            "bytes": merged_size,
            "segments": len(segments),
            "elapsed_sec": round(elapsed, 1),
        }

    # 单次下载
    filters = build_filters_for_task(task, start_date, end_date)
    print(f"  筛选: {filters}")
    print(f"  输出: {final_output}")
    t0 = time.time()
    result = await export_simple_report(
        username=USERNAME,
        password=PASSWORD,
        report_id=report_id,
        output_path=final_output,
        max_rows=max_rows,
        browser_channel="chrome",
        headless=headless,
        filters=filters,
    )
    elapsed = time.time() - t0
    result["elapsed_sec"] = round(elapsed, 1)
    print(f"  完成! {result.get('bytes', 0)} bytes, {elapsed:.1f}s")
    return result


def build_date_segments(start_date: str, end_date: str, segment_days: int) -> list[tuple[str, str]]:
    """把日期区间拆成多个 segment_days 天的连续段。"""
    sd = date.fromisoformat(start_date)
    ed = date.fromisoformat(end_date)
    segments = []
    cur = sd
    while cur <= ed:
        seg_end = min(cur + timedelta(days=segment_days - 1), ed)
        segments.append((cur.isoformat(), seg_end.isoformat()))
        cur = seg_end + timedelta(days=1)
    return segments


def merge_xlsx_files(input_files: list[Path], output_path: Path) -> None:
    """把多个 xlsx 文件合并成一个（追加行；以第一个文件的表头为准）。

    SmartBI 导出的 xlsx 可能不完全符合标准（比如 read_only 模式下 iter_rows 返回空），
    所以这里使用标准模式（非 read_only）读取，确保数据被正确提取。
    """
    import openpyxl
    if not input_files:
        raise ValueError("merge_xlsx_files: input_files is empty")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    header_written = False
    total_data_rows = 0
    for idx, f in enumerate(input_files):
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        seg_data_rows = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                if not header_written:
                    ws_out.append(list(row))
                    header_written = True
                # 跳过后续段的表头行
                continue
            ws_out.append(list(row))
            seg_data_rows += 1
        total_data_rows += seg_data_rows
        print(f"    [合并] 段{idx+1} {f.name}: {seg_data_rows} 行数据")
        wb.close()
    print(f"    [合并] 总数据行数: {total_data_rows}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    wb_out.close()


async def fetch_all(start_date: str, end_date: str, headless: bool = True) -> dict:
    """下载所有启用的报表任务。"""
    config = load_config()
    tasks = config.get("tasks", {})
    output_dir = OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    results = {}
    for task_name, task in tasks.items():
        if not task.get("enabled", True):
            continue
        print(f"\n{'='*60}")
        print(f"下载: {task.get('description', task_name)}")
        print(f"{'='*60}")
        try:
            result = await fetch_one_report(
                task_name, task, start_date, end_date, output_dir, headless
            )
            results[task_name] = {"status": "ok", **result}
        except SmartbiBrowserExportError as e:
            print(f"  [ERROR] {e}")
            results[task_name] = {"status": "error", "error": str(e)}
        except Exception as e:
            print(f"  [ERROR] {type(e).__name__}: {e}")
            results[task_name] = {"status": "error", "error": str(e)}

    success = sum(1 for r in results.values() if r["status"] == "ok")
    total = len(results)
    print(f"\n{'='*60}")
    print(f"结果: {success}/{total} 成功")
    if success == total:
        print(f"[OK] 所有报表下载完成，文件在: {output_dir}")
    else:
        print("[WARN] 部分报表下载失败，请检查上方日志")
    return results


def main():
    parser = argparse.ArgumentParser(
        description="下载学情积分核算所需 BI 报表（smartbi-data-cli 方式）"
    )
    parser.add_argument("--start", required=True, help="开始日期 YYYY-MM-DD")
    parser.add_argument("--end", required=True, help="结束日期 YYYY-MM-DD")
    parser.add_argument("--headful", action="store_true", help="显示浏览器窗口（调试用）")
    parser.add_argument("--json", action="store_true", help="输出 JSON 结果")
    args = parser.parse_args()

    results = asyncio.run(fetch_all(args.start, args.end, headless=not args.headful))

    if args.json:
        print(json.dumps(results, ensure_ascii=False, indent=2, default=str))

    failed = [k for k, v in results.items() if v["status"] != "ok"]
    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
